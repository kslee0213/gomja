"""valuation_verdict.py — 기업가치 대비 주가의 과소/과대평가 '결론'을 낸다.

두 추출 플러그인(dart-kospi-financials / us-stocks-financials)이 만든 캐시 또는
워크북을 입력받아, 여러 가치평가 방법을 같은 잣대로 계산하고 하나의 결론으로
합친다:

  1. 방법별 적정주가(주당)
       A 자산가치   = 지배주주 BPS × 정당 PBR( (ROE−g)/(r−g), 하한·상한 클램프 )
       B 수익가치   = 정상화 EPS(최근 3년 양수 평균) × 목표 PER(과거 밴드 중앙값 또는 기본값)
       C 현금흐름   = 정상화 FCF(OCF−CAPEX, 3년 평균) 2단계 DCF(5년 성장 → 5년 페이드 → 영구)
       D 잔여이익   = BPS + Σ (ROE_t − r)·B_{t−1} / (1+r)^t  (ROE가 10년에 걸쳐 r로 수렴)
       E 배당가치   = DPS×(1+g)/(r−g)  (참고용, 기본 가중치 0)
  2. 가중 평균 적정주가(가중치는 기업 성격에 따라 자동 조정, 사용자 재지정 가능)
  3. 시나리오(비관/기준/낙관): 할인율·성장률·목표 PER를 동시에 흔든 목표주가 범위
  4. 판정: 괴리율(현재가 vs 적정주가) 구간 + 신뢰도(방법 간 분산·데이터 완결성·추정연도 여부)
  5. 산출물: "가치평가_결론" 시트(가정 셀은 파란색 입력값, 계산은 전부 수식) + JSON + 마크다운 요약

정직성 규칙:
  - 데이터가 없으면 그 방법을 "해당 없음"으로 뺀다(0으로 지어내지 않음).
  - 할인율·성장률·목표 PER 등 리서치 없이 기본값을 쓴 항목은 결론에 그대로 열거한다.
  - 판정은 "현재 가정 하에서의 결론"이며, 주가 예측이 아님을 시트와 요약에 명시한다.

사용법:
  python valuation_verdict.py --source dart --cache-dir <dart cache> --corp-code 00126380 [--xlsx 파일] [--assumptions a.json]
  python valuation_verdict.py --source sec  --cache-dir <sec cache>  --ticker AAPL [--xlsx 파일]
  python valuation_verdict.py --source xlsx --xlsx <기존 워크북>      # 워크북 값을 직접 읽어 같은 파일에 시트 추가
  python valuation_verdict.py --source json --input normalized.json --out-xlsx 결과.xlsx
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import statistics
import sys
from dataclasses import dataclass, field, asdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from adapters import FinInput, load_dart, load_json, load_sec, load_xlsx  # noqa: E402

SHEET = "가치평가_결론"
N1, N2 = 5, 5  # DCF 1단계 성장 연수, 페이드 연수


# --------------------------------------------------------------------------
# 가정
# --------------------------------------------------------------------------
@dataclass
class Assumptions:
    cost_of_equity_pct: float | None = None   # None이면 시장 기본값
    rf_pct: float | None = None               # rf·beta·erp를 주면 r = rf + beta×erp
    beta: float | None = None
    erp_pct: float | None = None
    terminal_growth_pct: float | None = None
    stage1_growth_pct: float | None = None    # None이면 매출 CAGR(3년)에서 자동(0~12% 클램프)
    target_per: float | None = None           # None이면 과거 PER 밴드 중앙값 → 없으면 시장 기본값
    payout_ratio: float | None = None         # None이면 DPS·배당지급액에서 추정 → 없으면 0.25
    pbr_floor: float = 0.3
    pbr_cap: float = 4.0
    weights: dict[str, float] | None = None   # {"asset":..,"earnings":..,"dcf":..,"rim":..,"ddm":..}
    normalize_years: int = 3
    convergence_years: int = 3                # 기대수익률 계산 시 적정가 수렴 가정 연수
    bear: dict = field(default_factory=lambda: {"r_add": 1.5, "g1_mult": 0.5, "per_mult": 0.8})
    bull: dict = field(default_factory=lambda: {"r_add": -1.0, "g1_mult": 1.3, "per_mult": 1.2})
    defaults_used: list[str] = field(default_factory=list)

    @classmethod
    def from_file(cls, path: str | None) -> "Assumptions":
        a = cls()
        if path:
            d = json.loads(Path(path).read_text(encoding="utf-8"))
            for k, v in d.items():
                if hasattr(a, k):
                    setattr(a, k, v)
        return a


MARKET_DEFAULTS = {
    "KR": {"r": 9.0, "g": 2.0, "per": 10.0},
    "US": {"r": 9.0, "g": 2.5, "per": 18.0},
}


# --------------------------------------------------------------------------
# 계산 유틸
# --------------------------------------------------------------------------
def _last(vals: list[float | None]) -> float | None:
    return next((v for v in reversed(vals) if v is not None), None)


def _tail(vals: list[float | None], n: int) -> list[float]:
    return [v for v in vals[-n:] if v is not None]


def _cagr(vals: list[float | None], n: int) -> float | None:
    v = [x for x in vals if x is not None]
    if len(v) < 2:
        return None
    v = v[-(n + 1):]
    a, b = v[0], v[-1]
    if a <= 0 or b <= 0:
        return None
    return (b / a) ** (1 / (len(v) - 1)) - 1


@dataclass
class MethodResult:
    key: str
    name: str
    value: float | None
    applicable: bool
    weight: float
    detail: dict
    note: str = ""


@dataclass
class Scenario:
    name: str
    r_pct: float
    g1_pct: float
    gT_pct: float
    target_per: float
    methods: list[MethodResult]
    fair_value: float | None
    upside_pct: float | None


def resolve_params(fi: FinInput, a: Assumptions) -> dict:
    """가정값을 확정한다(기본값 사용 여부를 기록)."""
    md = MARKET_DEFAULTS.get(fi.market, MARKET_DEFAULTS["KR"])
    p: dict = {}
    if a.rf_pct is not None and a.beta is not None and a.erp_pct is not None:
        p["r"] = a.rf_pct + a.beta * a.erp_pct
    elif a.cost_of_equity_pct is not None:
        p["r"] = a.cost_of_equity_pct
    else:
        p["r"] = md["r"]
        a.defaults_used.append(f"자기자본비용 r={md['r']}% (시장 기본값, 리서치 안 됨)")
    p["gT"] = a.terminal_growth_pct if a.terminal_growth_pct is not None else md["g"]
    if a.terminal_growth_pct is None:
        a.defaults_used.append(f"영구성장률 g={md['g']}% (시장 기본값)")

    rev_cagr = _cagr(fi.series("revenue"), a.normalize_years)
    if a.stage1_growth_pct is not None:
        p["g1"] = a.stage1_growth_pct
    else:
        p["g1"] = max(0.0, min(12.0, (rev_cagr or 0.0) * 100))
        a.defaults_used.append(f"1단계 성장률 g1={p['g1']:.1f}% (최근 {a.normalize_years}년 매출 CAGR {((rev_cagr or 0) * 100):.1f}%를 0~12%로 클램프)")

    # 과거 PER 밴드 (KR: 결산일 종가 × 주식수 ÷ 지배주주순이익)
    ni_p = _parent(fi, "ni")
    hist_per = []
    if fi.shares:
        for px, ni in zip(fi.hist_price, ni_p):
            if px and ni and ni > 0:
                hist_per.append(px * fi.shares / (ni * fi.amount_unit))
    p["hist_per"] = hist_per
    if a.target_per is not None:
        p["per"] = a.target_per
    elif len(hist_per) >= 3 and 5 <= statistics.median(hist_per) <= 40:
        p["per"] = statistics.median(hist_per)
        a.defaults_used.append(f"목표 PER={p['per']:.1f}배 (회사의 과거 {len(hist_per)}개년 PER 중앙값 — 업종 비교는 안 됨)")
    else:
        p["per"] = md["per"]
        a.defaults_used.append(f"목표 PER={md['per']}배 (시장 기본값 — 과거 PER 밴드가 없거나 5~40배를 벗어남)")

    # 배당성향
    if a.payout_ratio is not None:
        p["payout"] = a.payout_ratio
    else:
        ni_last = _last(ni_p)
        div_last = _last(fi.series("dividends_paid"))
        if fi.dps and fi.shares and ni_last and ni_last > 0:
            p["payout"] = min(1.0, max(0.0, fi.dps * fi.shares / (ni_last * fi.amount_unit)))
        elif div_last and ni_last and ni_last > 0:
            p["payout"] = min(1.0, max(0.0, div_last / ni_last))
        else:
            p["payout"] = 0.25
            a.defaults_used.append("배당성향 25% (배당 데이터 없음)")
    p["rev_cagr"] = rev_cagr
    return p


def _parent(fi: FinInput, what: str) -> list[float | None]:
    """지배주주 값이 있으면 그것, 없으면 전체 값."""
    par = fi.series(f"{what}_parent")
    tot = fi.series(what)
    if any(v is not None for v in par):
        return [pv if pv is not None else tv for pv, tv in zip(par, tot)]
    return tot


def compute_methods(fi: FinInput, p: dict, a: Assumptions, r_pct: float, g1_pct: float, gT_pct: float, per: float) -> list[MethodResult]:
    r, g1, gT = r_pct / 100, g1_pct / 100, gT_pct / 100
    unit, sh = fi.amount_unit, fi.shares
    ni_p, eq_p = _parent(fi, "ni"), _parent(fi, "equity")
    ocf, capex = fi.series("ocf"), fi.series("capex")
    fcf = [(o - c) if (o is not None and c is not None) else None for o, c in zip(ocf, capex)]
    ny = a.normalize_years

    bps = (_last(eq_p) * unit / sh) if (sh and _last(eq_p) is not None) else None
    roe_hist = [(n / e) for n, e in zip(ni_p, eq_p) if n is not None and e not in (None, 0) and e > 0]
    roe_norm = statistics.mean(roe_hist[-ny:]) if roe_hist else None
    eps_hist = [(n * unit / sh) for n in ni_p if n is not None] if sh else []
    eps_pos = [e for e in eps_hist[-ny:] if e > 0]
    eps_norm = statistics.mean(eps_pos) if (eps_pos and eps_hist and eps_hist[-1] > 0) else None
    fcf_tail = _tail(fcf, ny)
    fcf_norm = statistics.mean(fcf_tail) if fcf_tail else None
    out: list[MethodResult] = []

    # A 자산가치
    if bps is not None and bps > 0 and roe_norm is not None and r > gT:
        jpbr = (roe_norm - gT) / (r - gT)
        jpbr_c = max(a.pbr_floor, min(a.pbr_cap, jpbr))
        out.append(MethodResult("asset", "A. 자산가치(BPS×정당PBR)", bps * jpbr_c, True, 0,
                                {"BPS": bps, "ROE_norm": roe_norm, "justified_PBR_raw": jpbr, "justified_PBR": jpbr_c}))
    else:
        out.append(MethodResult("asset", "A. 자산가치(BPS×정당PBR)", None, False, 0, {"BPS": bps, "ROE_norm": roe_norm}, "BPS≤0 또는 ROE/주식수 없음"))

    # B 수익가치
    if eps_norm is not None and eps_norm > 0:
        out.append(MethodResult("earnings", "B. 수익가치(정상화EPS×목표PER)", eps_norm * per, True, 0,
                                {"EPS_norm": eps_norm, "EPS_latest": eps_hist[-1] if eps_hist else None, "target_PER": per, "hist_PER": p.get("hist_per", [])}))
    else:
        out.append(MethodResult("earnings", "B. 수익가치(정상화EPS×목표PER)", None, False, 0,
                                {"EPS_latest": eps_hist[-1] if eps_hist else None}, "최근 순이익이 적자이거나 EPS 산출 불가"))

    # C DCF (FCF)
    dcf_detail: dict = {"FCF_norm": fcf_norm, "FCF_hist": fcf}
    if fcf_norm is not None and fcf_norm > 0 and sh and r > gT:
        pv_sum, f, growths = 0.0, fcf_norm, []
        for t in range(1, N1 + N2 + 1):
            g = g1 if t <= N1 else g1 + (gT - g1) * (t - N1) / N2
            f *= (1 + g)
            growths.append(g)
            pv_sum += f / (1 + r) ** t
        tv = f * (1 + gT) / (r - gT)
        pv_tv = tv / (1 + r) ** (N1 + N2)
        eq_val = pv_sum + pv_tv
        dcf_detail.update({"PV_stage": pv_sum, "PV_TV": pv_tv, "TV_share_pct": pv_tv / eq_val * 100 if eq_val else None, "growth_path": growths, "equity_value": eq_val})
        out.append(MethodResult("dcf", "C. 현금흐름가치(FCF 2단계 DCF)", eq_val * unit / sh, True, 0, dcf_detail))
    else:
        out.append(MethodResult("dcf", "C. 현금흐름가치(FCF 2단계 DCF)", None, False, 0, dcf_detail, "정상화 FCF≤0 이거나 OCF/CAPEX/주식수 없음"))

    # D 잔여이익(RIM)
    if bps is not None and bps > 0 and roe_norm is not None:
        b_prev, pv_ri, path = bps, 0.0, []
        for t in range(1, 11):
            roe_t = roe_norm + (r - roe_norm) * t / 10
            ri = (roe_t - r) * b_prev
            pv_ri += ri / (1 + r) ** t
            path.append(roe_t)
            b_prev = b_prev * (1 + roe_t * (1 - p["payout"]))
        out.append(MethodResult("rim", "D. 잔여이익가치(RIM)", bps + pv_ri, True, 0,
                                {"BPS": bps, "ROE_norm": roe_norm, "PV_RI": pv_ri, "roe_path": path, "payout": p["payout"]},
                                "ROE가 10년에 걸쳐 r로 수렴한다고 가정(그 이후 초과이익 0)"))
    else:
        out.append(MethodResult("rim", "D. 잔여이익가치(RIM)", None, False, 0, {}, "BPS≤0 또는 ROE 없음"))

    # E 배당(DDM) — 참고
    if fi.dps and fi.dps > 0 and r > gT:
        out.append(MethodResult("ddm", "E. 배당가치(DDM, 참고)", fi.dps * (1 + gT) / (r - gT), True, 0, {"DPS": fi.dps, "g": gT}))
    else:
        out.append(MethodResult("ddm", "E. 배당가치(DDM, 참고)", None, False, 0, {}, "배당 없음"))
    return out


def assign_weights(fi: FinInput, methods: list[MethodResult], a: Assumptions, p: dict) -> tuple[list[MethodResult], str]:
    by = {m.key: m for m in methods}
    if a.weights:
        w = dict(a.weights)
        why = "사용자 지정 가중치"
    else:
        base = {"KR": {"asset": .25, "earnings": .30, "dcf": .25, "rim": .20, "ddm": 0},
                "US": {"asset": .10, "earnings": .35, "dcf": .35, "rim": .20, "ddm": 0}}[fi.market if fi.market in ("KR", "US") else "KR"]
        w = dict(base)
        why = f"{fi.market} 기본 가중치"
        roe_norm = by["asset"].detail.get("ROE_norm")
        bps = by["asset"].detail.get("BPS")
        pbr_now = (fi.price / bps) if (fi.price and bps) else None
        if not by["earnings"].applicable:
            w = {"asset": .5, "earnings": 0, "dcf": .3, "rim": .2, "ddm": 0}
            why = "최근 순이익 적자 → 수익가치 제외, 자산가치 비중 확대"
        elif pbr_now is not None and pbr_now < 0.7 and roe_norm is not None and roe_norm < 0.06:
            w = {"asset": .40, "earnings": .20, "dcf": .20, "rim": .20, "ddm": 0}
            why = "PBR<0.7·ROE<6%(자산주 성격) → 자산가치 비중 확대"
        elif by["dcf"].applicable and (by["dcf"].detail.get("TV_share_pct") or 0) > 80:
            cut = w["dcf"] * 0.4
            w["dcf"] -= cut
            w["earnings"] += cut  # 영구가치 의존이 큰 DCF 비중을 줄여 수익가치로 이전
            why += " / DCF 영구가치 비중>80% → DCF 가중치 40% 축소"
    tot = sum(w.get(m.key, 0) for m in methods if m.applicable)
    for m in methods:
        m.weight = (w.get(m.key, 0) / tot) if (m.applicable and tot > 0) else 0.0
    return methods, why


def fair_value(methods: list[MethodResult]) -> float | None:
    vals = [(m.value, m.weight) for m in methods if m.applicable and m.weight > 0 and m.value is not None]
    if not vals:
        return None
    return sum(v * w for v, w in vals)


def run_scenario(name: str, fi: FinInput, p: dict, a: Assumptions, r_pct: float, g1_pct: float, gT_pct: float, per: float) -> Scenario:
    ms = compute_methods(fi, p, a, r_pct, g1_pct, gT_pct, per)
    ms, _ = assign_weights(fi, ms, a, p)
    fv = fair_value(ms)
    up = (fv / fi.price - 1) * 100 if (fv and fi.price) else None
    return Scenario(name, r_pct, g1_pct, gT_pct, per, ms, fv, up)


# --------------------------------------------------------------------------
# 판정
# --------------------------------------------------------------------------
def verdict_band(upside_pct: float | None) -> tuple[str, str]:
    if upside_pct is None:
        return "판정 불가", "현재가 또는 적정주가가 없음"
    if upside_pct >= 30:
        return "저평가(강)", "적정주가가 현재가보다 30% 이상 높음"
    if upside_pct >= 10:
        return "저평가", "적정주가가 현재가보다 10~30% 높음"
    if upside_pct > -10:
        return "적정", "현재가가 적정주가 ±10% 이내"
    if upside_pct > -30:
        return "고평가", "현재가가 적정주가보다 10~30% 높음"
    return "고평가(강)", "현재가가 적정주가보다 30% 이상 높음"


def confidence(fi: FinInput, base: Scenario, p: dict) -> tuple[str, list[str]]:
    reasons: list[str] = []
    app = [m for m in base.methods if m.applicable and m.weight > 0 and m.value]
    level = 3  # 3=높음 2=보통 1=낮음
    if base.fair_value and len(app) >= 2:
        disp = (max(m.value for m in app) - min(m.value for m in app)) / base.fair_value
        if disp >= 0.8:
            level -= 2
            reasons.append(f"방법 간 분산 큼(최대−최소 = 적정주가의 {disp * 100:.0f}%)")
        elif disp >= 0.4:
            level -= 1
            reasons.append(f"방법 간 분산 보통(최대−최소 = 적정주가의 {disp * 100:.0f}%)")
    if len(app) <= 2:
        level -= 1
        reasons.append(f"적용 가능한 방법이 {len(app)}개뿐")
    if not any(m.key == "earnings" and m.applicable for m in base.methods):
        level -= 1
        reasons.append("최근 순이익 적자 → 수익가치 미적용(자산가치 의존)")
    if fi.estimated_last:
        level -= 1
        reasons.append("마지막 연도가 추정치(E)")
    ocf, ni = fi.series("ocf"), _parent(fi, "ni")
    q = [o / n for o, n in zip(ocf[-3:], ni[-3:]) if o is not None and n and n > 0]
    if q and statistics.mean(q) < 0.8:
        level -= 1
        reasons.append(f"이익의 질 낮음(OCF/순이익 3년 평균 {statistics.mean(q):.2f})")
    if fi.price_date:
        try:
            age = (dt.date.today() - dt.datetime.strptime(fi.price_date, "%Y%m%d").date()).days
            if age > 60:
                level -= 1
                reasons.append(f"주가 기준일이 {age}일 전(오래됨)")
        except ValueError:
            pass
    if any(m.key == "dcf" and m.applicable and (m.detail.get("TV_share_pct") or 0) > 75 for m in base.methods):
        reasons.append("DCF 가치의 75% 이상이 영구가치(가정 민감)")
    label = {3: "높음", 2: "보통"}.get(max(1, level), "낮음")
    if not reasons:
        reasons.append("특이 감점 요인 없음")
    return label, reasons


# --------------------------------------------------------------------------
# 전체 실행
# --------------------------------------------------------------------------
def evaluate(fi: FinInput, a: Assumptions) -> dict:
    p = resolve_params(fi, a)
    base = run_scenario("기준", fi, p, a, p["r"], p["g1"], p["gT"], p["per"])
    bear = run_scenario("비관", fi, p, a, p["r"] + a.bear["r_add"], p["g1"] * a.bear["g1_mult"], p["gT"], p["per"] * a.bear["per_mult"])
    bull = run_scenario("낙관", fi, p, a, p["r"] + a.bull["r_add"], min(p["g1"] * a.bull["g1_mult"], 20.0), p["gT"], p["per"] * a.bull["per_mult"])
    _, why = assign_weights(fi, base.methods, a, p)
    band, band_why = verdict_band(base.upside_pct)
    conf, conf_why = confidence(fi, base, p)

    dy = (fi.dps / fi.price * 100) if (fi.dps and fi.price) else 0.0
    exp_ret = None
    if base.fair_value and fi.price:
        exp_ret = ((base.fair_value / fi.price) ** (1 / a.convergence_years) - 1) * 100 + dy
    liq_ps = (fi.liquidation_value * fi.amount_unit / fi.shares) if (fi.liquidation_value is not None and fi.shares) else None
    if liq_ps is not None and liq_ps < 0:
        fi.warnings.append("청산가치(간이)가 음수 — 조정자산이 총부채에 못 미침(청산 시 주주 몫 없음)")
    app_vals = [m.value for m in base.methods if m.applicable and m.weight > 0 and m.value]
    if fi.price is None:
        conf = "낮음"
        conf_why.insert(0, "현재가가 없어 괴리율·판정을 낼 수 없음")

    return {
        "company": fi.company, "market": fi.market, "currency": fi.currency, "source": fi.source,
        "generated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "price": fi.price, "price_date": fi.price_date, "shares": fi.shares, "years": fi.years, "estimated_last": fi.estimated_last,
        "params": {k: v for k, v in p.items() if k != "hist_per"}, "hist_per": p.get("hist_per", []),
        "weights_reason": why,
        "base": _scen_dict(base), "bear": _scen_dict(bear), "bull": _scen_dict(bull),
        "fair_value": base.fair_value,
        "upside_pct": base.upside_pct,
        "premium_discount_pct": ((fi.price / base.fair_value - 1) * 100) if (fi.price and base.fair_value) else None,
        "range_methods": {"min": min(app_vals) if app_vals else None, "max": max(app_vals) if app_vals else None},
        "target_range": {"bear": bear.fair_value, "base": base.fair_value, "bull": bull.fair_value},
        "downside_refs": {"liquidation_per_share": liq_ps, "bear_case": bear.fair_value},
        "expected_annual_return_pct": exp_ret, "dividend_yield_pct": dy, "convergence_years": a.convergence_years,
        "verdict": band, "verdict_reason": band_why, "confidence": conf, "confidence_reasons": conf_why,
        "defaults_used": a.defaults_used, "notes": fi.notes, "warnings": fi.warnings,
        "disclaimer": ("이 결론은 위 가정(할인율·성장률·목표 PER)에 종속된 '조건부 추정'이며 주가 예측이 아닙니다. "
                       "기본값을 쓴 가정이 하나라도 있으면 그 부분에 한해 '이 정보들은 정확하지 않습니다'."),
    }


def _scen_dict(s: Scenario) -> dict:
    d = asdict(s)
    return d


# --------------------------------------------------------------------------
# 엑셀 시트 (가정=파란색 입력, 계산=수식)
# --------------------------------------------------------------------------
def write_sheet(xlsx_path: str, fi: FinInput, res: dict, a: Assumptions) -> str:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter as L

    FN = "맑은 고딕"
    TITLE, LABEL, BODY = Font(name=FN, bold=True, size=15), Font(name=FN, bold=True), Font(name=FN, size=10)
    BLUE, NOTE = Font(name=FN, color="0000FF"), Font(name=FN, italic=True, size=9, color="808080")
    SEC_FONT, SEC_FILL = Font(name=FN, bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F4E79")
    HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
    BIG = Font(name=FN, bold=True, size=13)
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    WRAP = Alignment(wrap_text=True, vertical="top")

    fmtv = (lambda v: "-" if v is None else (f"{v:,.0f}" if fi.currency == "KRW" else f"{v:,.2f}"))  # noqa: E731
    path = Path(xlsx_path)
    if path.exists():
        wb = load_workbook(path)
        if SHEET in wb.sheetnames:
            del wb[SHEET]
    else:
        wb = Workbook()
        wb.remove(wb.active)
    ws = wb.create_sheet(SHEET, 0)
    ws.sheet_view.showGridLines = False
    n = len(fi.years)
    cur = fi.currency
    amt_fmt = "#,##0.0;(#,##0.0);-"
    px_fmt = "#,##0" if cur == "KRW" else "#,##0.00"

    def box(r1, r2, c1, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).border = BORDER

    def section(r, text, width=12):
        ws.cell(r, 1, text).font = SEC_FONT
        for c in range(1, width + 1):
            ws.cell(r, c).fill = SEC_FILL
        return r + 1

    def inp(r, label, value, fmt, note=""):
        ws.cell(r, 1, label).font = BODY
        c = ws.cell(r, 2, value)
        c.font, c.number_format = BLUE, fmt
        if note:
            ws.cell(r, 3, note).font = NOTE
        box(r, r, 1, 2)
        return f"$B${r}"

    # ---- 헤더 ----
    ws["A1"] = f"가치평가 결론 — {fi.company}"
    ws["A1"].font = TITLE
    ws["A2"] = (f"시장 {fi.market} · 통화 {cur} · 금액 단위 {fi.unit_label} · 데이터 {fi.source} · 생성 {res['generated_at']} · "
                f"주가 기준일 {fi.price_date or '미상'}")
    ws["A2"].font = NOTE

    # ---- 0. 결론 요약(자리만 잡아두고 맨 마지막에 채운다) ----
    SUMMARY_ROW = 4
    r = SUMMARY_ROW + 8

    # ---- 1. 가정(입력) ----
    r = section(r, "1. 가정 (파란색 셀을 바꾸면 아래 기준 시나리오가 전부 재계산됩니다)")
    p = res["params"]
    R = inp(r, "자기자본비용 r", p["r"] / 100, "0.0%", "rf+β×ERP 또는 시장 기본값"); r += 1
    GT = inp(r, "영구성장률 g", p["gT"] / 100, "0.0%", "장기 명목성장률 이하"); r += 1
    G1 = inp(r, f"1단계 성장률 g1 ({N1}년)", p["g1"] / 100, "0.0%", f"이후 {N2}년간 g로 선형 페이드"); r += 1
    PER = inp(r, "목표 PER", p["per"], "0.0", "과거 밴드 중앙값/시장 기본값/사용자 지정"); r += 1
    PAY = inp(r, "배당성향", p["payout"], "0.0%", "RIM의 장부가 성장에 사용"); r += 1
    PBF = inp(r, "정당 PBR 하한", a.pbr_floor, "0.00"); r += 1
    PBC = inp(r, "정당 PBR 상한", a.pbr_cap, "0.00"); r += 1
    SH = inp(r, "주식수", fi.shares, "#,##0", "자기주식 포함 발행주식수"); r += 1
    PX = inp(r, f"현재가({cur})", fi.price, px_fmt, f"기준일 {fi.price_date or '미상'}"); r += 1
    UNIT = inp(r, "금액 단위 배수", fi.amount_unit, "#,##0", f"{fi.unit_label} → {cur}"); r += 1
    DPS = inp(r, f"DPS({cur})", fi.dps if fi.dps is not None else 0, px_fmt); r += 1
    base = res["base"]
    W = {}
    for m in base["methods"]:
        W[m["key"]] = inp(r, f"가중치 · {m['name']}", round(m["weight"], 4), "0.0%", "" if m["applicable"] else f"해당 없음: {m['note']}")
        r += 1
    ws.cell(r, 1, f"가중치 근거: {res['weights_reason']}").font = NOTE
    r += 2

    # ---- 2. 입력 데이터 ----
    r = section(r, f"2. 입력 데이터 (단위: {fi.unit_label}, 지배주주 귀속값 우선)")
    hdr = r
    ws.cell(r, 1, "항목").font = LABEL
    for i, y in enumerate(fi.years):
        ws.cell(r, 3 + i, y).font = LABEL
        ws.cell(r, 3 + i).fill = HDR_FILL
    ws.cell(r, 1).fill = HDR_FILL
    r += 1
    rows = {}
    ni_p, eq_p = _parent(fi, "ni"), _parent(fi, "equity")
    data_rows = [
        ("revenue", "매출액", fi.series("revenue")), ("op", "영업이익", fi.series("op")),
        ("ni", "순이익(지배주주)", ni_p), ("equity", "자본(지배주주)", eq_p),
        ("assets", "자산총계", fi.series("assets")), ("liabilities", "부채총계", fi.series("liabilities")),
        ("ocf", "영업활동현금흐름", fi.series("ocf")), ("capex", "CAPEX(유형자산취득)", fi.series("capex")),
        ("cash", "현금및현금성자산", fi.series("cash")), ("borrowings", "차입금(합산)", fi.series("borrowings")),
        ("hist_price", f"결산일 종가({cur})", fi.hist_price),
    ]
    for key, label, vals in data_rows:
        ws.cell(r, 1, label).font = BODY
        rows[key] = r
        for i in range(n):
            v = vals[i] if i < len(vals) else None
            c = ws.cell(r, 3 + i, v)
            c.font, c.number_format = BLUE, (px_fmt if key == "hist_price" else amt_fmt)
        r += 1
    col = lambda i: L(3 + i)  # noqa: E731
    lastc, first3 = col(n - 1), col(max(0, n - a.normalize_years))
    derived = [
        ("fcf", "FCF = OCF − CAPEX", lambda i: f"=IF(OR({col(i)}{rows['ocf']}=\"\",{col(i)}{rows['capex']}=\"\"),\"\",{col(i)}{rows['ocf']}-{col(i)}{rows['capex']})", amt_fmt),
        ("eps", f"EPS({cur})", lambda i: f"=IF({col(i)}{rows['ni']}=\"\",\"\",{col(i)}{rows['ni']}*{UNIT}/{SH})", px_fmt),
        ("bps", f"BPS({cur})", lambda i: f"=IF({col(i)}{rows['equity']}=\"\",\"\",{col(i)}{rows['equity']}*{UNIT}/{SH})", px_fmt),
        ("roe", "ROE", lambda i: f"=IFERROR({col(i)}{rows['ni']}/{col(i)}{rows['equity']},\"\")", "0.0%"),
        ("per_h", "과거 PER(결산일)", lambda i: f"=IFERROR(IF({col(i)}{rows['ni']}<=0,\"\",{col(i)}{rows['hist_price']}/{col(i)}{rows['eps'] if 'eps' in rows else 0}),\"\")", "0.0"),
        ("pbr_h", "과거 PBR(결산일)", lambda i: f"=IFERROR({col(i)}{rows['hist_price']}/{col(i)}{rows['bps']},\"\")", "0.00"),
        ("ocf_ni", "OCF/순이익(이익의 질)", lambda i: f"=IFERROR({col(i)}{rows['ocf']}/{col(i)}{rows['ni']},\"\")", "0.00"),
    ]
    for key, label, fn, fmt in derived:
        ws.cell(r, 1, label).font = BODY
        rows[key] = r
        for i in range(n):
            c = ws.cell(r, 3 + i, fn(i))
            c.number_format = fmt
        r += 1
    box(hdr, r - 1, 1, 2 + n)
    if fi.estimated_last:
        ws.cell(r, 1, f"※ {fi.years[-1]}는 추정치입니다(사업보고서 미공시). 정상화 값에는 포함되지만 신뢰도 감점 요인입니다.").font = NOTE
        r += 1
    r += 1

    # ---- 3. 방법별 적정주가 (수식) ----
    r = section(r, f"3. 방법별 적정주가 (주당, {cur}) — 기준 시나리오, 전부 수식")
    ws.cell(r, 1, "지표").font = LABEL; ws.cell(r, 2, "값").font = LABEL
    for c in (1, 2):
        ws.cell(r, c).fill = HDR_FILL
    r += 1
    M: dict[str, str] = {}

    def calc(label, formula, fmt, key=None, note=""):
        nonlocal r
        ws.cell(r, 1, label).font = BODY
        c = ws.cell(r, 2, formula)
        c.number_format = fmt
        if note:
            ws.cell(r, 3, note).font = NOTE
        box(r, r, 1, 2)
        ref = f"$B${r}"
        if key:
            M[key] = ref
        r += 1
        return ref

    ni_rng, eq_rng = f"{first3}{rows['ni']}:{lastc}{rows['ni']}", f"{first3}{rows['equity']}:{lastc}{rows['equity']}"
    roe_rng, eps_rng, fcf_rng = f"{first3}{rows['roe']}:{lastc}{rows['roe']}", f"{first3}{rows['eps']}:{lastc}{rows['eps']}", f"{first3}{rows['fcf']}:{lastc}{rows['fcf']}"
    BPS = calc("BPS(최근, 지배주주)", f"={lastc}{rows['bps']}", px_fmt)
    ROE = calc(f"정상화 ROE(최근 {a.normalize_years}년 평균)", f"=IFERROR(AVERAGE({roe_rng}),\"\")", "0.0%")
    JPBR = calc("정당 PBR = (ROE−g)/(r−g), 클램프", f"=IFERROR(MAX({PBF},MIN({PBC},({ROE}-{GT})/({R}-{GT}))),\"\")", "0.00")
    A_ = calc("▶ A. 자산가치 = BPS×정당PBR", f"=IFERROR({BPS}*{JPBR},\"\")", px_fmt, "asset")
    EPSN = calc(f"정상화 EPS(최근 {a.normalize_years}년 중 양수 평균, 최근연도 흑자일 때만)", f"=IFERROR(IF({lastc}{rows['eps']}>0,AVERAGEIF({eps_rng},\">0\"),\"\"),\"\")", px_fmt)
    B_ = calc("▶ B. 수익가치 = 정상화EPS×목표PER", f"=IFERROR(IF({EPSN}=\"\",\"\",{EPSN}*{PER}),\"\")", px_fmt, "earnings")
    FCFN = calc(f"정상화 FCF(최근 {a.normalize_years}년 평균, {fi.unit_label})", f"=IFERROR(AVERAGE({fcf_rng}),\"\")", amt_fmt)
    r += 1
    # DCF 표
    ws.cell(r, 1, "DCF 연차").font = LABEL
    for t in range(1, N1 + N2 + 1):
        ws.cell(r, 2 + t, t).font = LABEL
        ws.cell(r, 2 + t).fill = HDR_FILL
    ws.cell(r, 1).fill = HDR_FILL
    hdr_dcf = r; r += 1
    ws.cell(r, 1, "성장률").font = BODY
    g_row = r
    for t in range(1, N1 + N2 + 1):
        f = f"={G1}" if t <= N1 else f"={G1}+({GT}-{G1})*{t - N1}/{N2}"
        ws.cell(r, 2 + t, f).number_format = "0.0%"
    r += 1
    ws.cell(r, 1, f"FCF 예측({fi.unit_label})").font = BODY
    f_row = r
    for t in range(1, N1 + N2 + 1):
        prev = FCFN if t == 1 else f"{L(2 + t - 1)}{r}"
        ws.cell(r, 2 + t, f"=IFERROR({prev}*(1+{L(2 + t)}{g_row}),\"\")").number_format = amt_fmt
    r += 1
    ws.cell(r, 1, "현재가치").font = BODY
    pv_row = r
    for t in range(1, N1 + N2 + 1):
        ws.cell(r, 2 + t, f"=IFERROR({L(2 + t)}{f_row}/(1+{R})^{t},\"\")").number_format = amt_fmt
    r += 1
    box(hdr_dcf, r - 1, 1, 2 + N1 + N2)
    lastdc = L(2 + N1 + N2)
    PVS = calc(f"예측기간 PV 합({fi.unit_label})", f"=IFERROR(SUM(C{pv_row}:{lastdc}{pv_row}),\"\")", amt_fmt)
    TV = calc("영구가치 PV = FCF10×(1+g)/(r−g)/(1+r)^10", f"=IFERROR({lastdc}{f_row}*(1+{GT})/({R}-{GT})/(1+{R})^{N1 + N2},\"\")", amt_fmt)
    calc("영구가치 비중", f"=IFERROR({TV}/({PVS}+{TV}),\"\")", "0%", note="75% 넘으면 가정 민감")
    C_ = calc("▶ C. 현금흐름가치 = (PV합+TV)×단위/주식수", f"=IFERROR(IF({FCFN}<=0,\"\",({PVS}+{TV})*{UNIT}/{SH}),\"\")", px_fmt, "dcf")
    r += 1
    # RIM 표
    ws.cell(r, 1, "RIM 연차").font = LABEL
    for t in range(1, 11):
        ws.cell(r, 2 + t, t).font = LABEL
        ws.cell(r, 2 + t).fill = HDR_FILL
    ws.cell(r, 1).fill = HDR_FILL
    hdr_rim = r; r += 1
    ws.cell(r, 1, "ROE_t (r로 선형 수렴)").font = BODY
    roe_row = r
    for t in range(1, 11):
        ws.cell(r, 2 + t, f"=IFERROR({ROE}+({R}-{ROE})*{t}/10,\"\")").number_format = "0.0%"
    r += 1
    ws.cell(r, 1, f"기초 BPS({cur})").font = BODY
    b_row = r
    for t in range(1, 11):
        prev = BPS if t == 1 else f"{L(2 + t - 1)}{r}*(1+{L(2 + t - 1)}{roe_row}*(1-{PAY}))"
        ws.cell(r, 2 + t, f"=IFERROR({prev},\"\")").number_format = px_fmt
    r += 1
    ws.cell(r, 1, "잔여이익 PV").font = BODY
    ri_row = r
    for t in range(1, 11):
        ws.cell(r, 2 + t, f"=IFERROR(({L(2 + t)}{roe_row}-{R})*{L(2 + t)}{b_row}/(1+{R})^{t},\"\")").number_format = px_fmt
    r += 1
    box(hdr_rim, r - 1, 1, 12)
    D_ = calc("▶ D. 잔여이익가치 = BPS + Σ잔여이익PV", f"=IFERROR({BPS}+SUM(C{ri_row}:L{ri_row}),\"\")", px_fmt, "rim")
    E_ = calc("▶ E. 배당가치 = DPS×(1+g)/(r−g)", f"=IFERROR(IF({DPS}<=0,\"\",{DPS}*(1+{GT})/({R}-{GT})),\"\")", px_fmt, "ddm")
    r += 1

    # ---- 4. 결론 ----
    r = section(r, "4. 결론 (기준 시나리오는 수식, 비관/낙관은 스크립트 계산값)")
    wsum = "+".join(f"IF(ISNUMBER({M[k]}),{W[k]},0)" for k in M)
    wprod = "+".join(f"IF(ISNUMBER({M[k]}),{M[k]}*{W[k]},0)" for k in M)
    FV = calc(f"★ 적정주가(가중평균, {cur})", f"=IFERROR(({wprod})/({wsum}),\"\")", px_fmt, note="가중치는 1번 섹션에서 조정")
    ws.cell(r - 1, 1).font = BIG; ws.cell(r - 1, 2).font = BIG
    calc("현재가", f"={PX}", px_fmt)
    UP = calc("상승여력 = 적정주가/현재가 − 1", f"=IFERROR({FV}/{PX}-1,\"\")", "0.0%")
    calc("괴리율 = 현재가/적정주가 − 1 (양수=고평가)", f"=IFERROR({PX}/{FV}-1,\"\")", "0.0%")
    VD = calc("판정", f"=IF({UP}=\"\",\"판정 불가\",IF({UP}>=0.3,\"저평가(강)\",IF({UP}>=0.1,\"저평가\",IF({UP}>-0.1,\"적정\",IF({UP}>-0.3,\"고평가\",\"고평가(강)\")))))", "General",
              note="상승여력 +30%↑ 저평가(강) / +10~30 저평가 / ±10 적정 / −10~−30 고평가 / −30↓ 고평가(강)")
    ws.cell(r - 1, 2).font = BIG
    calc(f"기대 연평균수익률({a.convergence_years}년 수렴 가정, 배당 포함)", f"=IFERROR(({FV}/{PX})^(1/{a.convergence_years})-1+IF({PX}>0,{DPS}/{PX},0),\"\")", "0.0%",
         note="적정주가로 수렴한다는 '가정'의 산술 결과이지 예측이 아님")
    vals = [M[k] for k in M]
    calc("방법별 최솟값", f"=IFERROR(MIN({','.join(vals)}),\"\")", px_fmt)
    calc("방법별 최댓값", f"=IFERROR(MAX({','.join(vals)}),\"\")", px_fmt)
    liq = res["downside_refs"]["liquidation_per_share"]
    calc(f"하방 참고: 청산가치/주({cur})", liq if liq is not None else "산출 불가", px_fmt)
    r += 1
    ws.cell(r, 1, "시나리오").font = LABEL
    for i, h in enumerate(["r", "g1", "목표PER", f"목표주가({cur})", "상승여력"]):
        ws.cell(r, 2 + i, h).font = LABEL
    for c in range(1, 7):
        ws.cell(r, c).fill = HDR_FILL
    sh_ = r; r += 1
    for key in ("bear", "base", "bull"):
        s = res[key]
        ws.cell(r, 1, s["name"]).font = BODY
        ws.cell(r, 2, s["r_pct"] / 100).number_format = "0.0%"
        ws.cell(r, 3, s["g1_pct"] / 100).number_format = "0.0%"
        ws.cell(r, 4, s["target_per"]).number_format = "0.0"
        if key == "base":
            ws.cell(r, 5, f"={FV}").number_format = px_fmt
            ws.cell(r, 6, f"={UP}").number_format = "0.0%"
        else:
            ws.cell(r, 5, s["fair_value"]).number_format = px_fmt
            ws.cell(r, 6, (s["upside_pct"] / 100) if s["upside_pct"] is not None else None).number_format = "0.0%"
            for c in (5, 6):
                ws.cell(r, c).font = BLUE
        r += 1
    box(sh_, r - 1, 1, 6)
    ws.cell(r, 1, f"비관: r+{a.bear['r_add']}%p, g1×{a.bear['g1_mult']}, PER×{a.bear['per_mult']} / 낙관: r{a.bull['r_add']:+}%p, g1×{a.bull['g1_mult']}(상한 20%), PER×{a.bull['per_mult']}").font = NOTE
    r += 2

    # ---- 5. 신뢰도·가정 고지 ----
    r = section(r, "5. 신뢰도 · 불확실성 고지")
    ws.cell(r, 1, "신뢰도").font = LABEL
    ws.cell(r, 2, res["confidence"]).font = BIG
    r += 1
    for t in res["confidence_reasons"]:
        ws.cell(r, 1, f"- {t}").font = BODY
        r += 1
    r += 1
    ws.cell(r, 1, "리서치 없이 기본값/자동 추정을 쓴 가정 (이 정보들은 정확하지 않습니다):").font = LABEL
    r += 1
    for t in res["defaults_used"] or ["(없음 — 모든 가정을 사용자가 지정)"]:
        ws.cell(r, 1, f"- {t}").font = BODY
        r += 1
    if res["warnings"]:
        r += 1
        ws.cell(r, 1, "데이터 경고:").font = LABEL
        r += 1
        for t in res["warnings"]:
            ws.cell(r, 1, f"- {t}").font = Font(name=FN, color="C00000", size=10)
            r += 1
    if res["notes"]:
        r += 1
        ws.cell(r, 1, "데이터 메모:").font = LABEL
        r += 1
        for t in res["notes"]:
            ws.cell(r, 1, f"- {t}").font = NOTE
            r += 1
    r += 1
    ws.cell(r, 1, res["disclaimer"]).font = NOTE
    ws.cell(r, 1).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=12)

    # ---- 0. 결론 요약 채우기 ----
    sr = section(SUMMARY_ROW, "0. 결론 요약")
    summary = [
        ("판정", f"={VD}", "General"), (f"적정주가(기준, {cur})", f"={FV}", px_fmt), (f"현재가({cur})", f"={PX}", px_fmt),
        ("상승여력", f"={UP}", "0.0%"),
        (f"목표주가 범위 비관 / 낙관({cur})", f"{fmtv(res['target_range']['bear'])} / {fmtv(res['target_range']['bull'])}" if res["target_range"]["bear"] is not None else "-", "General"),
        ("신뢰도", res["confidence"], "General"),
    ]
    for i, (label, val, fmt) in enumerate(summary):
        ws.cell(sr + i, 1, label).font = LABEL
        c = ws.cell(sr + i, 2, val)
        c.number_format = fmt
        c.font = BIG if i in (0, 1) else BODY
        box(sr + i, sr + i, 1, 2)
    ws.cell(sr, 3, "세부 계산과 가정은 아래 1~5번 섹션. 결론은 가정에 종속된 조건부 추정이며 주가 예측이 아님.").font = NOTE

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 16
    for c in range(3, 16):
        ws.column_dimensions[L(c)].width = 13
    ws.freeze_panes = "A4"
    wb.save(path)
    return str(path)


# --------------------------------------------------------------------------
# 요약(markdown)
# --------------------------------------------------------------------------
def render_markdown(res: dict, fi: FinInput) -> str:
    cur = fi.currency
    f = (lambda v: "-" if v is None else (f"{v:,.0f}" if cur == "KRW" else f"{v:,.2f}"))
    pct = lambda v: "-" if v is None else f"{v:+.1f}%"  # noqa: E731
    lines = [
        f"## {fi.company} 가치평가 결론 ({res['generated_at']})",
        f"- 현재가: {f(res['price'])} {cur} (기준일 {res['price_date'] or '미상'})",
        f"- **적정주가(기준): {f(res['fair_value'])} {cur}** → 상승여력 {pct(res['upside_pct'])}, 괴리율 {pct(res['premium_discount_pct'])}",
        f"- **판정: {res['verdict']}** ({res['verdict_reason']}) · 신뢰도 **{res['confidence']}**",
        f"- 목표주가 범위: 비관 {f(res['target_range']['bear'])} / 기준 {f(res['target_range']['base'])} / 낙관 {f(res['target_range']['bull'])}",
        f"- 하방 참고: 청산가치/주 {f(res['downside_refs']['liquidation_per_share'])}, 방법별 최솟값 {f(res['range_methods']['min'])}",
        f"- 기대 연평균수익률({res['convergence_years']}년 수렴 가정, 배당 {res['dividend_yield_pct']:.1f}% 포함): {pct(res['expected_annual_return_pct'])}",
        "",
        "| 방법 | 적정주가 | 가중치 | 비고 |", "|---|---:|---:|---|",
    ]
    for m in res["base"]["methods"]:
        lines.append(f"| {m['name']} | {f(m['value'])} | {m['weight'] * 100:.0f}% | {m['note'] or ''} |")
    lines += ["", f"가중치 근거: {res['weights_reason']}", "", "신뢰도 근거:"] + [f"- {t}" for t in res["confidence_reasons"]]
    lines += ["", "기본값/자동 추정 가정 (이 정보들은 정확하지 않습니다):"] + [f"- {t}" for t in res["defaults_used"]]
    if res["warnings"]:
        lines += ["", "데이터 경고:"] + [f"- {t}" for t in res["warnings"]]
    lines += ["", f"> {res['disclaimer']}"]
    return "\n".join(lines)


# --------------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser(description="기업가치 대비 주가 과소/과대평가 결론")
    ap.add_argument("--source", choices=["dart", "sec", "xlsx", "json"], required=True)
    ap.add_argument("--cache-dir", help="dart/sec: 추출 플러그인의 cache 폴더")
    ap.add_argument("--corp-code", help="dart: 8자리 corp_code")
    ap.add_argument("--ticker", help="sec: 티커")
    ap.add_argument("--input", help="json: 정규화 입력 파일")
    ap.add_argument("--xlsx", help="시트를 추가할 기존 워크북(xlsx 소스면 필수)")
    ap.add_argument("--out-xlsx", help="워크북이 없을 때 새로 만들 파일 경로")
    ap.add_argument("--assumptions", help="가정 JSON(assumptions.example.json 참고)")
    ap.add_argument("--years", type=int, default=5)
    ap.add_argument("--no-estimate", action="store_true", help="dart: 진행연도(E) 추정 컬럼을 만들지 않음")
    ap.add_argument("--json-out", help="결과 JSON 저장 경로(기본: 워크북 옆 verdict_{회사}.json)")
    args = ap.parse_args()

    if args.source == "dart":
        if not (args.cache_dir and args.corp_code):
            ap.error("--source dart 는 --cache-dir 와 --corp-code 가 필요합니다")
        fi = load_dart(args.cache_dir, args.corp_code, args.years, estimate=not args.no_estimate)
    elif args.source == "sec":
        if not (args.cache_dir and args.ticker):
            ap.error("--source sec 는 --cache-dir 와 --ticker 가 필요합니다")
        fi = load_sec(args.cache_dir, args.ticker, args.years)
    elif args.source == "xlsx":
        if not args.xlsx:
            ap.error("--source xlsx 는 --xlsx 가 필요합니다")
        fi = load_xlsx(args.xlsx, args.years)
    else:
        if not args.input:
            ap.error("--source json 은 --input 이 필요합니다")
        fi = load_json(args.input)

    a = Assumptions.from_file(args.assumptions)
    res = evaluate(fi, a)

    target = args.xlsx or args.out_xlsx
    if not target:
        target = f"{fi.company}_가치평가결론_{dt.date.today().strftime('%Y%m%d')}.xlsx"
    saved = write_sheet(target, fi, res, a)
    jpath = args.json_out or str(Path(saved).with_name(f"verdict_{fi.company}.json"))
    Path(jpath).write_text(json.dumps(res, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(render_markdown(res, fi))
    print()
    print(json.dumps({"saved": saved, "json": jpath, "verdict": res["verdict"], "confidence": res["confidence"],
                      "fair_value": res["fair_value"], "price": res["price"], "upside_pct": res["upside_pct"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()

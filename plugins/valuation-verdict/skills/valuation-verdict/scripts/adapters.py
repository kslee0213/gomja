"""adapters.py — 여러 소스(DART 캐시 / SEC 캐시 / 기존 xlsx / 정규화 JSON)를
`FinInput`(정규화된 재무 입력) 하나로 바꾼다.

valuation_verdict.py는 이 FinInput만 보고 계산하므로, 한국·미국 어느 쪽이든
같은 결론 로직을 탄다. 새 데이터 소스를 붙이려면 여기에 어댑터 하나만 추가하면 된다.

원칙:
  - 값을 지어내지 않는다. 못 찾은 항목은 None으로 두고 warnings에 사유를 남긴다.
  - 금액은 FinInput.amount_unit(원 단위 환산 배수: 억원=1e8, 백만달러=1e6) 기준으로
    저장하고, 주당 값으로 바꿀 때만 amount_unit을 곱한다.
  - 지배주주 귀속 순이익/자본(ni_parent/equity_parent)이 있으면 그것을 우선 쓴다
    (비지배지분이 큰 지주사·조선사 등에서 PER/PBR 왜곡을 줄이기 위함).
"""
from __future__ import annotations

import datetime as dt
import glob
import json
import os
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

SERIES_KEYS = [
    "revenue", "op", "ni", "ni_parent", "equity", "equity_parent", "assets", "liabilities",
    "ocf", "capex", "da", "cash", "borrowings", "dividends_paid",
]


@dataclass
class FinInput:
    company: str
    market: str                      # "KR" | "US"
    currency: str                    # "KRW" | "USD"
    amount_unit: float               # 1e8(억원) | 1e6(백만달러)
    unit_label: str                  # "억원" | "백만달러"
    years: list[str] = field(default_factory=list)
    estimated_last: bool = False     # 마지막 연도가 추정(E)인지
    s: dict[str, list[float | None]] = field(default_factory=dict)
    price: float | None = None
    price_date: str | None = None    # YYYYMMDD
    shares: float | None = None      # 발행주식수(자기주식 포함)
    treasury_shares: float | None = None
    dps: float | None = None         # 최근 주당 현금배당(통화 단위)
    hist_price: list[float | None] = field(default_factory=list)  # 연도별 결산일 종가(있으면)
    liquidation_value: float | None = None  # amount_unit 기준
    source: str = ""
    notes: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def series(self, key: str) -> list[float | None]:
        n = len(self.years)
        v = self.s.get(key)
        if v is None:
            return [None] * n
        return list(v) + [None] * (n - len(v))

    def to_dict(self) -> dict:
        d = self.__dict__.copy()
        return d


# --------------------------------------------------------------------------
# 공통 유틸
# --------------------------------------------------------------------------
def _num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if s in ("", "-", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _yyyymmdd(s: str | None) -> str | None:
    if not s:
        return None
    m = re.search(r"(\d{4})-?(\d{2})-?(\d{2})", str(s))
    return f"{m.group(1)}{m.group(2)}{m.group(3)}" if m else None


# --------------------------------------------------------------------------
# DART 캐시 어댑터 (dart-kospi-financials/skills/dart-financial-extractor/cache)
# --------------------------------------------------------------------------
# key: (sj_div 후보, 완전일치 후보, 부분일치 키워드(모두 포함), 제외 키워드)
DART_RULES: dict[str, tuple[tuple[str, ...], list[str], list[list[str]], list[str]]] = {
    "revenue": (("IS", "CIS"), ["매출액", "수익(매출액)", "영업수익", "매출"], [["매출액"], ["영업수익"]], ["원가", "율", "총이익", "채권"]),
    "op": (("IS", "CIS"), ["영업이익", "영업이익(손실)"], [["영업이익"]], ["율", "률"]),
    "ni": (("IS", "CIS"), ["당기순이익", "당기순이익(손실)", "연결당기순이익", "반기순이익", "분기순이익",
                          "반기순이익(손실)", "분기순이익(손실)"],
           [["순이익"]], ["지배", "비지배", "주당", "포괄", "계속영업", "중단영업"]),
    "ni_parent": (("IS", "CIS"), [], [["지배기업", "순이익"], ["지배주주", "순이익"], ["지배기업", "이익"]], ["비지배", "주당", "포괄"]),
    "equity": (("BS",), ["자본총계"], [["자본총계"]], []),
    "equity_parent": (("BS",), [], [["지배기업", "자본"], ["지배기업", "지분"], ["지배주주", "지분"], ["지배주주", "자본"]], ["비지배"]),
    "assets": (("BS",), ["자산총계"], [["자산총계"]], []),
    "liabilities": (("BS",), ["부채총계"], [["부채총계"]], []),
    "cash": (("BS",), ["현금및현금성자산"], [["현금및현금성자산"]], ["증가", "감소", "증감", "기초", "기말"]),
    "ocf": (("CF",), ["영업활동현금흐름", "영업활동으로인한현금흐름", "영업활동으로 인한 현금흐름"], [["영업활동"]], ["투자활동", "재무활동"]),
    "capex": (("CF",), ["유형자산의취득", "유형자산의 취득"], [["유형자산", "취득"]], ["처분", "무형"]),
    "da": (("CF",), ["감가상각비"], [["감가상각"]], ["무형", "사용권"]),
    "dividends_paid": (("CF",), ["배당금지급", "배당금의지급", "배당금의 지급", "배당"], [["배당금"]], ["수취", "수입"]),
    # 청산가치용(선택)
    "short_fin": (("BS",), ["단기금융상품", "단기투자자산", "유가증권"], [["단기금융상품"], ["단기투자자산"]], ["장기"]),
    "receivables": (("BS",), ["매출채권및기타채권", "매출채권및기타유동채권", "매출채권"], [["매출채권"]], ["비유동", "장기"]),
    "inventories": (("BS",), ["재고자산"], [["재고자산"]], []),
    "ppe": (("BS",), ["유형자산"], [["유형자산"]], ["취득", "처분", "투자"]),
    "lt_invest": (("BS",), ["투자자산", "장기투자자산", "기타비유동금융자산"], [["투자자산"]], ["단기"]),
}
BORROWING_NAMES = ["단기차입금", "장기차입금", "사채", "유동성장기부채", "유동성장기차입금", "유동성사채", "단기사채", "장기사채", "차입금"]

# 이름이 중복되어 구분이 안 되는 계정(예: 당기순이익/총포괄이익 아래에 똑같이
# "지배기업의 소유주지분" 행이 있는 회사 — HD한국조선해양에서 실제 확인)은
# XBRL 표준 account_id로 먼저 찾는다. 이름 규칙은 그다음 폴백.
DART_ID_CANDIDATES: dict[str, list[str]] = {
    "ni": ["ifrs-full_ProfitLoss", "ifrs-full_ProfitLossFromContinuingOperations"],
    "ni_parent": ["ifrs-full_ProfitLossAttributableToOwnersOfParent"],
    "equity": ["ifrs-full_Equity"],
    "equity_parent": ["ifrs-full_EquityAttributableToOwnersOfParent"],
    "revenue": ["ifrs-full_Revenue"],
    "op": ["dart_OperatingIncomeLoss", "ifrs-full_ProfitLossFromOperatingActivities"],
    "assets": ["ifrs-full_Assets"],
    "liabilities": ["ifrs-full_Liabilities"],
    "ocf": ["ifrs-full_CashFlowsFromUsedInOperatingActivities"],
    "capex": ["ifrs-full_PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities"],
    "dividends_paid": ["ifrs-full_DividendsPaidClassifiedAsFinancingActivities"],
}


def _dart_lookup(items: dict, key: str, field: str = "thstrm_amount") -> tuple[float | None, str | None]:
    sjs, exact, substr_sets, excludes = DART_RULES[key]
    rows = [it for sj in sjs for it in items.get(sj, [])]

    # 0차: XBRL 표준 account_id 완전일치 (이름 중복 계정을 정확히 구분)
    for aid in DART_ID_CANDIDATES.get(key, []):
        for it in rows:
            if it.get("account_id") == aid:
                v = _num(it.get(field))
                if v is not None:
                    return v, it.get("account_nm")

    def ok(nm: str) -> bool:
        return not any(ex in nm for ex in excludes)

    for cand in exact:
        for it in rows:
            nm = it.get("account_nm", "")
            if nm.replace(" ", "") == cand.replace(" ", "") and ok(nm):
                v = _num(it.get(field))
                if v is not None:
                    return v, nm
    for kws in substr_sets:
        for it in rows:
            nm = it.get("account_nm", "")
            if all(k in nm.replace(" ", "") for k in kws) and ok(nm):
                v = _num(it.get(field))
                if v is not None:
                    return v, nm
    return None, None


def _dart_borrowings(items: dict, field: str = "thstrm_amount") -> tuple[float | None, list[str]]:
    total, used = 0.0, []
    seen = set()
    for it in items.get("BS", []):
        nm = it.get("account_nm", "").replace(" ", "")
        if nm in seen:
            continue
        if (any(b in nm for b in ("차입금", "사채")) or nm in ("단기금융부채", "장기금융부채", "유동성장기금융부채")) and not any(x in nm for x in ("상환", "비용", "이자", "발행", "증가", "감소")):
            v = _num(it.get(field))
            if v is not None:
                total += v
                used.append(nm)
                seen.add(nm)
    return (total if used else None), used


def load_dart(cache_dir: str, corp_code: str, years: int = 5, estimate: bool = True) -> FinInput:
    cache = Path(cache_dir)
    reports: dict[str, dict[str, dict]] = {}
    for fp in cache.glob(f"{corp_code}_*.json"):
        try:
            data = json.loads(fp.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if data.get("status") != "000":
            continue
        reports.setdefault(data["bsns_year"], {})[data["reprt_code"]] = data
    if not reports:
        raise SystemExit(f"ERROR: {cache}에 {corp_code} 캐시가 없습니다. dart-financial-extractor로 먼저 수집하세요.")

    profile = {}
    pf = cache / f"company_{corp_code}.json"
    if pf.exists():
        profile = json.loads(pf.read_text(encoding="utf-8"))
    company = profile.get("corp_name") or corp_code
    stock_code = (profile.get("stock_code") or "").strip()

    fi = FinInput(company=company, market="KR", currency="KRW", amount_unit=1e8, unit_label="억원", source=f"dart-cache:{corp_code}")
    fy_years = sorted([y for y in reports if "11011" in reports[y]])[-years:]
    fs_divs = {reports[y]["11011"].get("fs_div_used") for y in fy_years}
    if len(fs_divs) > 1:
        fi.warnings.append(f"연도별 재무제표 구분이 섞여 있습니다(CFS/OFS 혼재: {sorted(str(x) for x in fs_divs)}) — 연도 간 비교 왜곡 가능")
    fi.notes.append(f"재무제표 구분: {', '.join(sorted(str(x) for x in fs_divs))}")

    series: dict[str, list[float | None]] = {k: [] for k in SERIES_KEYS}
    liq_parts: dict[str, float | None] = {}
    for y in fy_years:
        items = reports[y]["11011"]["items"]
        for k in SERIES_KEYS:
            if k == "borrowings":
                v, _ = _dart_borrowings(items)
            else:
                v, _ = _dart_lookup(items, k)
            if k in ("capex", "dividends_paid") and v is not None:
                v = abs(v)
            series[k].append(v / 1e8 if v is not None else None)
    fi.years = list(fy_years)

    # 추정 연도(E): 전년 실적 × (당해 누적 ÷ 전년 동기 누적), 부호가 다르거나 전년 동기 ≤ 0이면 건너뜀
    if estimate and fy_years:
        cand = str(int(fy_years[-1]) + 1)
        latest = next((c for c in ("11014", "11012", "11013") if c in reports.get(cand, {})), None)
        if latest and latest != "11011" and latest in reports.get(fy_years[-1], {}):
            def _has_add(data):
                return any("thstrm_add_amount" in it for sj in ("IS", "CIS", "CF") for it in data.get("items", {}).get(sj, []))
            # 1분기/반기보고서도 누적 필드가 공시돼 있으면 그것으로 비교한다(회사별 관행 차이 대응)
            fld = "thstrm_add_amount" if (latest == "11014" or _has_add(reports[cand][latest])) else "thstrm_amount"
            cur_items, prior_items, fy_items = reports[cand][latest]["items"], reports[fy_years[-1]][latest]["items"], reports[fy_years[-1]]["11011"]["items"]
            est: dict[str, float | None] = {}
            skipped = []
            for k in SERIES_KEYS:
                if k in ("equity", "equity_parent", "assets", "liabilities", "cash", "borrowings"):
                    v, _ = _dart_borrowings(cur_items) if k == "borrowings" else _dart_lookup(cur_items, k)
                    est[k] = v / 1e8 if v is not None else None
                    continue
                c, _ = _dart_lookup(cur_items, k, fld)
                p, _ = _dart_lookup(prior_items, k, fld)
                f, _ = _dart_lookup(fy_items, k)
                # 전년 동기가 0이거나 당해·전년동기·전년연간의 부호가 다르면(흑자/적자 전환) 추정하지 않는다.
                if c is None or f is None or p is None or p == 0 or c / p < 0 or f / p < 0:
                    est[k] = None
                    if c is not None and f is not None and p is not None:
                        skipped.append(k)
                    continue
                v = f * (c / p)
                est[k] = abs(v) / 1e8 if k in ("capex", "dividends_paid") else v / 1e8  # 유출 계정은 절대값 저장
            for k in SERIES_KEYS:
                series[k].append(est.get(k))
            fi.years.append(f"{cand}(E)")
            fi.estimated_last = True
            fi.notes.append(f"{cand}(E)는 {fy_years[-1]}년 실적 × ({cand}년 누적 ÷ {fy_years[-1]}년 동기 누적)으로 추정 (보고서 {latest})")
            if skipped:
                fi.warnings.append(f"{cand}(E) 추정 제외(전년 동기 ≤0 또는 부호 반전): {', '.join(skipped)}")
    fi.s = series

    # 청산가치(간이): 현금·단기금융 100%, 매출채권 85%, 재고·유형·장기투자 50%, 무형·기타 0% − 총부채
    last_items = reports[fy_years[-1]]["11011"]["items"]
    hair = {"cash": 1.0, "short_fin": 1.0, "receivables": 0.85, "inventories": 0.5, "ppe": 0.5, "lt_invest": 0.5}
    adj, found = 0.0, []
    for k, h in hair.items():
        v, nm = _dart_lookup(last_items, k)
        if v is not None:
            adj += v * h
            found.append(k)
    liab, _ = _dart_lookup(last_items, "liabilities")
    if liab is not None and found:
        fi.liquidation_value = (adj - liab) / 1e8
        fi.notes.append(f"청산가치(간이)는 {fy_years[-1]} 사업보고서 기준, 반영 항목: {', '.join(found)}")

    # 주가·주식수 (KRX 캐시)
    if stock_code:
        pfiles = sorted(cache.glob(f"price_{stock_code}_*.json"))
        latest_p, latest_d = None, None
        by_date = {}
        for pf_ in pfiles:
            try:
                p = json.loads(pf_.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            d = str(p.get("used_date") or p.get("requested_date") or "")
            by_date[str(p.get("requested_date") or d)] = p
            if d and (latest_d is None or d > latest_d):
                latest_d, latest_p = d, p
        if latest_p:
            fi.price = _num(latest_p.get("close_price"))
            fi.price_date = latest_d
            fi.shares = _num(latest_p.get("listed_shares"))
            mc = _num(latest_p.get("market_cap"))
            if fi.shares is None and mc and fi.price:
                fi.shares = mc / fi.price
        for y in fi.years:
            yy = y[:4]
            p = by_date.get(f"{yy}1231")
            fi.hist_price.append(_num(p.get("close_price")) if p else None)
        if fi.price is None:
            fi.warnings.append("KRX 주가 캐시가 없어 현재가·PER·목표주가 비교를 할 수 없습니다(fetch_stock_price.py 실행 필요)")
    else:
        fi.warnings.append("종목코드를 찾지 못해 주가를 연결하지 못했습니다")

    # 배당·자기주식 (DART 추가공시 캐시)
    ef = cache / f"extra_{corp_code}_{fy_years[-1]}_11011.json"
    if ef.exists():
        try:
            extra = json.loads(ef.read_text(encoding="utf-8"))
            div_list = (extra.get("배당") or {}).get("list", []) or []
            for x in div_list:
                se, knd = x.get("se") or "", x.get("stock_knd") or ""
                if "주당" in se and "현금" in se and ("보통" in knd or knd == ""):
                    v = _num(x.get("thstrm"))
                    if v is not None:
                        fi.dps = v
                        break
            tot_list = (extra.get("주식총수현황") or {}).get("list", []) or []
            for x in tot_list:
                t = _num(x.get("tesstk_co"))
                if t is not None:
                    fi.treasury_shares = (fi.treasury_shares or 0) + t
        except (OSError, json.JSONDecodeError):
            pass
    else:
        fi.warnings.append("DART 추가공시 캐시가 없어 배당(DPS)·자기주식을 반영하지 못했습니다")
    return fi


# --------------------------------------------------------------------------
# SEC CompanyFacts 캐시 어댑터 (us-stocks-financials/.../cache/secfacts_{ticker}.json)
# --------------------------------------------------------------------------
SEC_TAGS: dict[str, list[str]] = {
    "revenue": ["Revenues", "RevenueFromContractWithCustomerExcludingAssessedTax", "SalesRevenueNet", "RevenueFromContractWithCustomerIncludingAssessedTax"],
    "op": ["OperatingIncomeLoss"],
    "ni_parent": ["NetIncomeLoss", "NetIncomeLossAvailableToCommonStockholdersBasic"],
    "ni": ["ProfitLoss", "NetIncomeLoss"],
    "equity_parent": ["StockholdersEquity"],
    "equity": ["StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest", "StockholdersEquity"],
    "assets": ["Assets"],
    "liabilities": ["Liabilities"],
    "cash": ["CashAndCashEquivalentsAtCarryingValue", "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"],
    "ocf": ["NetCashProvidedByUsedInOperatingActivities", "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations"],
    "capex": ["PaymentsToAcquirePropertyPlantAndEquipment", "PaymentsToAcquireProductiveAssets"],
    "da": ["DepreciationDepletionAndAmortization", "DepreciationAmortizationAndAccretionNet", "DepreciationAndAmortization", "Depreciation"],
    "dividends_paid": ["PaymentsOfDividends", "PaymentsOfDividendsCommonStock"],
    # 차입금은 합산
    "_debt": ["LongTermDebtNoncurrent", "LongTermDebtCurrent", "DebtCurrent", "ShortTermBorrowings", "CommercialPaper", "LongTermDebtAndCapitalLeaseObligations"],
}
INSTANT = {"equity_parent", "equity", "assets", "liabilities", "cash"}


def _sec_annual_points(concept: dict, want_instant: bool) -> dict[str, float]:
    """{end_date: val} — 연간(350~380일) 값만, 같은 end는 filed가 최신인 것으로."""
    out: dict[str, tuple[str, float]] = {}
    for unit, facts in (concept.get("units") or {}).items():
        if unit not in ("USD", "USD/shares", "shares"):
            continue
        for f in facts:
            end, val, filed = f.get("end"), f.get("val"), f.get("filed", "")
            if end is None or val is None:
                continue
            if want_instant:
                if "start" in f:
                    continue
                if f.get("fp") not in (None, "FY") and not str(f.get("frame", "")).endswith("Q4I"):
                    # 분기 스냅샷 제외: FY 보고(10-K)의 값 또는 CY..Q4I 프레임만
                    if f.get("form", "").startswith("10-Q"):
                        continue
            else:
                if "start" not in f:
                    continue
                try:
                    days = (dt.date.fromisoformat(end) - dt.date.fromisoformat(f["start"])).days
                except ValueError:
                    continue
                if not (350 <= days <= 380):
                    continue
            if end not in out or filed >= out[end][0]:
                out[end] = (filed, float(val))
    return {k: v for k, (_, v) in out.items()}


def load_sec(cache_dir: str, ticker: str, years: int = 5) -> FinInput:
    cache = Path(cache_dir)
    fp = cache / f"secfacts_{ticker.upper()}.json"
    if not fp.exists():
        raise SystemExit(f"ERROR: {fp} 가 없습니다. us-stock-financial-extractor로 먼저 수집하세요.")
    raw = json.loads(fp.read_text(encoding="utf-8"))["raw"]
    us = raw.get("facts", {}).get("us-gaap", {})
    dei = raw.get("facts", {}).get("dei", {})
    company = raw.get("entityName") or ticker.upper()
    cf = cache / f"company_{ticker.upper()}.json"
    if cf.exists():
        try:
            company = json.loads(cf.read_text(encoding="utf-8")).get("title") or company
        except (OSError, json.JSONDecodeError):
            pass
    fi = FinInput(company=company, market="US", currency="USD", amount_unit=1e6, unit_label="백만달러", source=f"sec-cache:{ticker.upper()}")

    merged: dict[str, dict[str, float]] = {}
    for key, tags in SEC_TAGS.items():
        if key == "_debt":
            continue
        pts: dict[str, float] = {}
        for tag in tags:  # 태그 후보를 순서대로 훑되, 앞 태그가 비운 연도만 뒤 태그로 채운다(연도 누락 방지)
            c = us.get(tag)
            if not c:
                continue
            for end, v in _sec_annual_points(c, key in INSTANT).items():
                pts.setdefault(end, v)
        merged[key] = pts
    debt: dict[str, float] = {}
    for tag in SEC_TAGS["_debt"]:
        c = us.get(tag)
        if not c:
            continue
        for end, v in _sec_annual_points(c, True).items():
            debt[end] = debt.get(end, 0.0) + v
    if debt:
        merged["borrowings"] = debt
        fi.notes.append("차입금은 LongTermDebt(Non)current/DebtCurrent/ShortTermBorrowings/CommercialPaper 태그 합산(중복 태그 시 과대 가능)")

    # 회계연도 축: 매출·순이익(duration)이 있는 end 날짜 기준 최근 N개
    ends = sorted(set(merged.get("revenue", {})) | set(merged.get("ni_parent", {})))[-years:]
    if not ends:
        raise SystemExit("ERROR: SEC 캐시에서 연간(12개월) 매출/순이익 값을 찾지 못했습니다.")
    fi.years = [e[:4] if e[5:7] == "12" else f"FY{e[:4]}({e[5:7]})" for e in ends]
    series: dict[str, list[float | None]] = {}
    for key in SERIES_KEYS:
        pts = merged.get(key, {})
        vals = []
        for e in ends:
            v = pts.get(e)
            if v is None and key in INSTANT:
                # 결산일이 며칠 어긋나는 회사(52/53주 결산) 대비: ±10일 내 가장 가까운 잔액
                near = [(abs((dt.date.fromisoformat(k) - dt.date.fromisoformat(e)).days), k) for k in pts if abs((dt.date.fromisoformat(k) - dt.date.fromisoformat(e)).days) <= 10]
                if near:
                    v = pts[min(near)[1]]
            if v is not None and key in ("capex", "dividends_paid"):
                v = abs(v)
            vals.append(v / 1e6 if v is not None else None)
        series[key] = vals
    fi.s = series
    if all(v is None for v in series.get("liabilities", [])) and series.get("assets") and series.get("equity"):
        fi.s["liabilities"] = [(a - e) if (a is not None and e is not None) else None for a, e in zip(series["assets"], series["equity"])]
        fi.notes.append("부채총계는 Liabilities 태그가 없어 자산총계−자본총계로 계산")

    # 주식수: dei → 주가캐시 순
    sh = None
    c = dei.get("EntityCommonStockSharesOutstanding")
    if c:
        pts = {}
        for f in (c.get("units") or {}).get("shares", []):
            if f.get("end") and f.get("val") is not None:
                pts[(f["end"], f.get("filed", ""))] = float(f["val"])
        if pts:
            sh = pts[max(pts)]
            fi.notes.append("주식수: SEC dei:EntityCommonStockSharesOutstanding(최신 보고 기준)")
    pf = cache / f"price_{ticker.upper()}.json"
    if pf.exists():
        try:
            p = json.loads(pf.read_text(encoding="utf-8"))
            info = p.get("price", {}) or {}
            fi.price = _num(info.get("currentPrice"))
            fi.price_date = _yyyymmdd(p.get("fetched_at"))
            if sh is None:
                sh = _num(info.get("sharesOutstanding"))
                if sh:
                    fi.notes.append("주식수: yfinance sharesOutstanding")
            dy = _num(info.get("dividendYield"))
            if dy is not None and fi.price:
                fi.dps = fi.price * (dy if dy < 1 else dy / 100)
        except (OSError, json.JSONDecodeError):
            pass
    if fi.price is None:
        fi.warnings.append("주가 캐시(price_{ticker}.json)가 없어 현재가 비교를 할 수 없습니다(fetch_extra_info.py 실행 필요)")
    fi.shares = sh
    if fi.dps is None and sh and fi.s.get("dividends_paid"):
        last_div = next((v for v in reversed(fi.s["dividends_paid"]) if v is not None), None)
        if last_div:
            fi.dps = last_div * 1e6 / sh
            fi.notes.append("DPS는 현금흐름표 배당금지급 ÷ 주식수로 근사")
    fi.hist_price = [None] * len(fi.years)
    fi.notes.append("미국 워크북에는 연도별 과거 주가가 없어 과거 PER 밴드는 산출하지 않음(목표 PER는 기본값/사용자 지정)")
    return fi


# --------------------------------------------------------------------------
# 기존 xlsx 어댑터 (두 플러그인이 만든 워크북을 직접 읽는다)
# --------------------------------------------------------------------------
def _recalc_copy(xlsx: str) -> str:
    """수식 값이 캐시돼 있지 않은 파일은 LibreOffice로 재계산한 복사본을 만든다."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise SystemExit("ERROR: 워크북 수식 값이 계산돼 있지 않고 LibreOffice(soffice)도 없어 값을 읽을 수 없습니다. "
                         "Excel에서 한 번 저장하거나, --source dart/sec 로 캐시에서 직접 읽으세요.")
    tmpdir = tempfile.mkdtemp(prefix="vv_recalc_")
    subprocess.run([soffice, "--headless", "--convert-to", "xlsx", "--outdir", tmpdir, xlsx],
                   check=True, capture_output=True, timeout=180)
    out = os.path.join(tmpdir, Path(xlsx).name)
    if not os.path.exists(out):
        raise SystemExit("ERROR: LibreOffice 재계산 결과 파일을 찾지 못했습니다.")
    return out


def _find_row(ws, label: str, col: int, contains: bool = False) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None:
            continue
        v = str(v)
        if (contains and label in v) or (not contains and v == label):
            return r
    return None


def load_xlsx(path: str, years: int = 5) -> FinInput:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    if "연간_재무제표" not in wb.sheetnames:
        raise SystemExit("ERROR: '연간_재무제표' 시트가 없습니다(연간 워크북이 필요합니다).")
    ws = wb["연간_재무제표"]
    hdr = 3
    probe = [ws.cell(r, 3).value for r in range(hdr + 1, min(ws.max_row, hdr + 40))]
    if all(v is None for v in probe):
        path2 = _recalc_copy(path)
        wb = load_workbook(path2, data_only=True)
        ws = wb["연간_재무제표"]
    ia = wb["투자분석"] if "투자분석" in wb.sheetnames else None
    is_us = ia is not None and _find_row(ia, "현재가($)", 1) is not None
    company = Path(path).stem.split("_")[0]
    if ia is not None and ia["A1"].value and "—" in str(ia["A1"].value):
        company = str(ia["A1"].value).split("—", 1)[1].strip()
        company = re.sub(r"\s*\([A-Z.\-]+\)$", "", company)  # US: "Fake Corp (FAKE)" → "Fake Corp"

    labels = []
    c = 3
    while ws.cell(hdr, c).value not in (None, ""):
        labels.append(str(ws.cell(hdr, c).value))
        c += 1
    n = len(labels)
    labels = labels[-years - 1:] if (labels and labels[-1].endswith("(E)")) else labels[-years:]
    off = n - len(labels)

    def row_vals(r: int | None) -> list[float | None]:
        if r is None:
            return [None] * len(labels)
        return [_num(ws.cell(r, 3 + off + i).value) for i in range(len(labels))]

    if is_us:
        fi = FinInput(company=company, market="US", currency="USD", amount_unit=1e6, unit_label="백만달러", source=f"xlsx:{path}")
        lab = {"revenue": "매출액", "op": "영업이익", "ni": "당기순이익", "ni_parent": "당기순이익", "equity": "자본총계", "equity_parent": "자본총계",
               "assets": "자산총계", "liabilities": "부채총계", "cash": "현금및현금성자산", "ocf": "영업활동현금흐름",
               "capex": "설비투자(CapEx)", "da": "감가상각비(D&A)", "dividends_paid": "배당금지급"}
        for k, l in lab.items():
            v = row_vals(_find_row(ws, l, 2))
            fi.s[k] = [abs(x) if (x is not None and k in ("capex", "dividends_paid")) else x for x in v]
        fi.s["borrowings"] = [None] * len(labels)
        fi.warnings.append("xlsx 소스에는 차입금·비지배지분 구분이 없어 자본총계/당기순이익 전체를 지배주주 값으로 간주")
        fi.years = labels
        if ia is not None:
            r = _find_row(ia, "현재가($)", 1, contains=True)
            fi.price = _num(ia.cell(r, 3).value) if r else None
            m = re.search(r"기준 (\d{4}-\d{2}-\d{2})", str(ia.cell(r, 1).value)) if r else None
            fi.price_date = _yyyymmdd(m.group(1)) if m else None
            r = _find_row(ia, "시가총액(백만달러)", 1)
            mc = _num(ia.cell(r, 3).value) if r else None
            if mc and fi.price:
                fi.shares = mc * 1e6 / fi.price
            r = _find_row(ia, "배당수익률(%)", 1)
            dy = _num(ia.cell(r, 3).value) if r else None
            if dy is not None and fi.price:
                fi.dps = fi.price * dy / 100
        fi.hist_price = [None] * len(labels)
        if fi.price_date is None:
            fi.warnings.append("미국 워크북에 주가 기준일이 기록돼 있지 않습니다(yfinance 조회 시점)")
        return fi

    # --- KR ---
    fi = FinInput(company=company, market="KR", currency="KRW", amount_unit=1e8, unit_label="억원", source=f"xlsx:{path}")
    # 연간_재무제표의 B열(계정명)을 DART 규칙으로 매칭
    items_by_row = []
    for r in range(hdr + 1, ws.max_row + 1):
        nm = ws.cell(r, 2).value
        if nm:
            items_by_row.append((r, str(nm)))

    def match(key: str) -> int | None:
        _sjs, exact, substr_sets, excludes = DART_RULES[key]
        ok = lambda nm: not any(ex in nm for ex in excludes)  # noqa: E731
        for cand in exact:
            for r, nm in items_by_row:
                if nm.replace(" ", "") == cand.replace(" ", "") and ok(nm) and any(_num(ws.cell(r, 3 + off + i).value) is not None for i in range(len(labels))):
                    return r
        for kws in substr_sets:
            for r, nm in items_by_row:
                if all(k in nm.replace(" ", "") for k in kws) and ok(nm) and any(_num(ws.cell(r, 3 + off + i).value) is not None for i in range(len(labels))):
                    return r
        return None

    for k in SERIES_KEYS:
        if k == "borrowings":
            tot = [0.0] * len(labels)
            used = False
            for r, nm in items_by_row:
                nn = nm.replace(" ", "")
                if any(b in nn for b in ("차입금", "사채")) and not any(x in nn for x in ("상환", "비용", "이자", "발행", "증가", "감소")):
                    vals = row_vals(r)
                    if any(v is not None for v in vals):
                        used = True
                        tot = [(t + (v or 0)) for t, v in zip(tot, vals)]
            fi.s[k] = tot if used else [None] * len(labels)
            continue
        v = row_vals(match(k))
        fi.s[k] = [abs(x) if (x is not None and k in ("capex", "dividends_paid")) else x for x in v]
    fi.years = labels
    fi.estimated_last = bool(labels and labels[-1].endswith("(E)"))
    if ia is not None:
        r_close, r_mc, r_date = _find_row(ia, "종가", 1), _find_row(ia, "시가총액", 1), _find_row(ia, "기준일(종가)", 1)
        n_ia = 0
        if r_close:
            while ia.cell(r_close, 3 + n_ia).value not in (None, "") and n_ia < 20:
                n_ia += 1
        if r_close and n_ia:
            fi.hist_price = [_num(ia.cell(r_close, 3 + off + i).value) for i in range(len(labels))]
            fi.price = fi.hist_price[-1]
            if r_date:
                fi.price_date = _yyyymmdd(str(ia.cell(r_date, 3 + off + len(labels) - 1).value))
            mc = _num(ia.cell(r_mc, 3 + off + len(labels) - 1).value) if r_mc else None
            if mc and fi.price:
                fi.shares = mc * 1e8 / fi.price
        r = _find_row(ia, "주당 현금배당금", 1)
        if r:
            fi.dps = _num(ia.cell(r, 2).value)
        r = next((rr for rr in range(1, ia.max_row + 1) if str(ia.cell(rr, 1).value or "").startswith("청산가치 (")), None)
        if r:
            fi.liquidation_value = _num(ia.cell(r, 3).value)
    if fi.price is None:
        fi.warnings.append("워크북에 주가(투자분석 L섹션)가 없어 현재가 비교를 할 수 없습니다")
    return fi


# --------------------------------------------------------------------------
# 정규화 JSON 어댑터 (다른 파이프라인에서 직접 넘길 때)
# --------------------------------------------------------------------------
def load_json(path: str) -> FinInput:
    d = json.loads(Path(path).read_text(encoding="utf-8"))
    fi = FinInput(company=d["company"], market=d.get("market", "KR"), currency=d.get("currency", "KRW"),
                  amount_unit=float(d.get("amount_unit", 1e8)), unit_label=d.get("unit_label", "억원"), source=f"json:{path}")
    fi.years = [str(y) for y in d["years"]]
    fi.estimated_last = bool(d.get("estimated_last", False)) or fi.years[-1].endswith("(E)")
    fi.s = {k: [(_num(x) if x is not None else None) for x in d.get("series", {}).get(k, [])] for k in SERIES_KEYS}
    fi.price, fi.price_date = _num(d.get("price")), d.get("price_date")
    fi.shares, fi.treasury_shares, fi.dps = _num(d.get("shares")), _num(d.get("treasury_shares")), _num(d.get("dps"))
    fi.hist_price = [(_num(x) if x is not None else None) for x in d.get("hist_price", [])] or [None] * len(fi.years)
    fi.liquidation_value = _num(d.get("liquidation_value"))
    return fi

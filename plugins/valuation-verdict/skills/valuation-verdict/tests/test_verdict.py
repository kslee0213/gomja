"""valuation_verdict 회귀 테스트 (표준 unittest — 추가 패키지 불필요).

    cd plugins/valuation-verdict/skills/valuation-verdict && python -m unittest discover -s tests -v

- 합성 픽스처(make_fixtures.py)로 DART/SEC 캐시를 만들고 네 가지 소스(dart/sec/xlsx/json)를 돌린다.
- LibreOffice(soffice)가 있으면 시트 수식을 재계산해 파이썬 계산값과 대조한다.
- 상위 플러그인(dart-kospi-financials)이 같은 리포에 있으면 그 build_workbook.py로
  워크북을 만들어 xlsx 어댑터까지 통합 검증한다(없으면 건너뜀).
"""
import json
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent / "scripts"
REPO = HERE.parents[4]  # tests → skills/valuation-verdict → skills → plugins/valuation-verdict → plugins → repo root
DART_SKILL = REPO / "plugins/dart-kospi-financials/skills/dart-financial-extractor"
sys.path.insert(0, str(SCRIPTS))

from adapters import load_dart, load_json, load_sec, load_xlsx  # noqa: E402
from valuation_verdict import Assumptions, evaluate, write_sheet  # noqa: E402

ROOT = Path(tempfile.mkdtemp(prefix="vv_test_"))
DART_CACHE, SEC_CACHE = ROOT / "dart", ROOT / "sec"
subprocess.run([sys.executable, str(HERE / "make_fixtures.py"), "--dart-cache", str(DART_CACHE), "--sec-cache", str(SEC_CACHE)], check=True, capture_output=True)
SOFFICE = shutil.which("soffice")


def _recalc(path: Path) -> Path:
    out = path.parent / "recalc"
    out.mkdir(exist_ok=True)
    subprocess.run([SOFFICE, "--headless", "--convert-to", "xlsx", "--outdir", str(out), str(path)], check=True, capture_output=True, timeout=180)
    return out / path.name


def _sheet_values(path: Path) -> dict:
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True)["가치평가_결론"]
    got = {}
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "")
        for k, key in {"▶ A.": "asset", "▶ B.": "earnings", "▶ C.": "dcf", "▶ D.": "rim", "▶ E.": "ddm"}.items():
            if a.startswith(k):
                got[key] = ws.cell(r, 2).value
        if a.startswith("★"):
            got["fair"] = ws.cell(r, 2).value
        if a == "판정":
            got["verdict"] = ws.cell(r, 2).value
    return got


def _write_json(d: dict) -> str:
    p = ROOT / f"{d['company']}.json"
    p.write_text(json.dumps(d, ensure_ascii=False), encoding="utf-8")
    return str(p)


class VerdictTests(unittest.TestCase):
    def assertRel(self, a, b, rel=1e-6, msg=None):
        self.assertIsNotNone(a, msg)
        self.assertIsNotNone(b, msg)
        self.assertLessEqual(abs(a - b), rel * max(abs(a), abs(b), 1e-12), msg or f"{a} != {b}")

    def _assert_sheet_matches(self, path: Path, res: dict):
        if not SOFFICE:
            self.skipTest("LibreOffice 없음 — 시트 수식 대조 생략")
        got = _sheet_values(_recalc(path))
        for m in res["base"]["methods"]:
            if m["applicable"] and m["value"] is not None:
                self.assertRel(got[m["key"]], m["value"], msg=m["key"])
            else:
                self.assertIn(got.get(m["key"]), (None, ""), m["key"])
        if res["fair_value"] is not None:
            self.assertRel(got["fair"], res["fair_value"])
        self.assertEqual(got["verdict"], res["verdict"])

    def test_dart_source_full_pipeline(self):
        fi = load_dart(str(DART_CACHE), "00000001")
        self.assertEqual(fi.years[-1], "2026(E)")
        self.assertTrue(fi.estimated_last)
        self.assertEqual((fi.price, fi.shares), (65000, 100_000_000))
        self.assertRel(fi.s["ni_parent"][-2], 1010.88 * 0.92)   # 지배주주 순이익 우선
        self.assertIsNotNone(fi.s["capex"][-1])                 # 음수 계정도 (E) 추정 포함
        res = evaluate(fi, Assumptions())
        self.assertTrue(all(m["applicable"] for m in res["base"]["methods"]))
        self.assertEqual(res["verdict"], "고평가(강)")           # 픽스처는 PER 60배짜리 비싼 회사
        self.assertLess(res["target_range"]["bear"], res["fair_value"])
        self.assertLess(res["fair_value"], res["target_range"]["bull"])
        out = ROOT / "kr.xlsx"
        write_sheet(str(out), fi, res, Assumptions())
        self._assert_sheet_matches(out, res)

    def test_sec_source(self):
        fi = load_sec(str(SEC_CACHE), "FAKE")
        self.assertEqual(len(fi.years), 5)
        self.assertEqual(fi.years[-1], "2025")
        self.assertEqual((fi.shares, fi.price), (1_000_000_000, 150.0))
        self.assertIsNotNone(fi.s["borrowings"][-1])
        res = evaluate(fi, Assumptions())
        self.assertEqual(res["market"], "US")
        self.assertGreater(res["fair_value"], 0)
        out = ROOT / "us.xlsx"
        write_sheet(str(out), fi, res, Assumptions())
        self._assert_sheet_matches(out, res)

    @unittest.skipUnless((DART_SKILL / "scripts/build_workbook.py").exists(), "dart 플러그인이 리포에 없음")
    def test_xlsx_source_matches_dart_source(self):
        if not SOFFICE:
            self.skipTest("LibreOffice 없음 — 워크북 값 읽기 불가")
        target = DART_SKILL / "cache"
        backup = None
        if target.exists():
            backup = target.with_name("cache_backup_for_test")
            shutil.move(str(target), str(backup))
        outdir = ROOT / "wb"
        try:
            shutil.copytree(DART_CACHE, target)
            subprocess.run([sys.executable, str(DART_SKILL / "scripts/build_workbook.py"), "00000001", "가상전자", "--period", "annual", "--outdir", str(outdir)],
                           check=True, capture_output=True)
        finally:
            shutil.rmtree(target, ignore_errors=True)
            if backup:
                shutil.move(str(backup), str(target))
        xlsx = next(outdir.glob("가상전자_연간_*.xlsx"))
        rx = evaluate(load_xlsx(str(xlsx)), Assumptions(payout_ratio=0.25))
        rd = evaluate(load_dart(str(DART_CACHE), "00000001"), Assumptions(payout_ratio=0.25))
        self.assertRel(rx["fair_value"], rd["fair_value"], rel=1e-4)

    def test_loss_making_company_excludes_earnings_methods(self):
        p = _write_json({"company": "적자", "market": "KR", "years": ["2021", "2022", "2023"],
                         "series": {"revenue": [1000, 900, 800], "ni": [10, -50, -80], "equity": [500, 450, 370], "ocf": [20, -10, -30], "capex": [30, 30, 30]},
                         "price": 3000, "shares": 10_000_000, "hist_price": [5000, 4000, 3500]})
        res = evaluate(load_json(p), Assumptions())
        by = {m["key"]: m for m in res["base"]["methods"]}
        self.assertFalse(by["earnings"]["applicable"])
        self.assertFalse(by["dcf"]["applicable"])
        self.assertGreater(by["asset"]["weight"], 0.5)
        self.assertEqual(res["confidence"], "낮음")

    def test_no_price_gives_no_verdict(self):
        p = _write_json({"company": "무가격", "market": "US", "currency": "USD", "amount_unit": 1e6, "unit_label": "백만달러", "years": ["2023", "2024"],
                         "series": {"revenue": [100, 110], "ni": [10, 12], "equity": [50, 60], "ocf": [15, 16], "capex": [3, 3]}, "shares": 1_000_000})
        res = evaluate(load_json(p), Assumptions())
        self.assertEqual(res["verdict"], "판정 불가")
        self.assertIsNotNone(res["fair_value"])

    def test_user_assumptions_override(self):
        p = _write_json({"company": "X", "market": "KR", "years": ["2023", "2024", "2025"],
                         "series": {"revenue": [100, 110, 120], "ni": [10, 11, 12], "equity": [100, 105, 110], "ocf": [14, 15, 16], "capex": [4, 4, 4]},
                         "price": 1000, "shares": 10_000_000, "hist_price": [900, 950, 1000]})
        a = Assumptions(cost_of_equity_pct=8.0, terminal_growth_pct=1.0, target_per=12.0, weights={"asset": 0, "earnings": 1, "dcf": 0, "rim": 0, "ddm": 0})
        res = evaluate(load_json(p), a)
        self.assertEqual((res["params"]["per"], res["params"]["r"]), (12.0, 8.0))
        self.assertFalse([t for t in res["defaults_used"] if "목표 PER" in t or "자기자본비용" in t])
        by = {m["key"]: m for m in res["base"]["methods"]}
        self.assertRel(res["fair_value"], by["earnings"]["value"])


if __name__ == "__main__":
    unittest.main()

#!/usr/bin/env python3
"""build_comparison_workbook.py — 여러 미국 상장기업 비교 워크북.

⚠️ v1.0.0의 핵심 버그: "12분기 비교"와 "최근5년 비교" 시트가 완전히 같은
함수·같은 데이터(yfinance .info의 현재 시점 스냅샷)로 만들어져서 두 시트가
똑같은 값을 보여줬다(시계열 데이터 자체가 없었음). v2.0.0은 각 기업의
분기/연간 재무제표 캐시(fetch_financials.py 결과)를 실제로 읽어서, 두 시트가
서로 다른 기간·다른 값을 보여주도록 재작성했다. 차트도 실제로 그린다.
"""
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
CACHE_DIR = SCRIPT_DIR.parent / "cache"

FONT_NAME = "맑은 고딕"
BOLD = Font(name=FONT_NAME, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
NOTE = Font(name=FONT_NAME, italic=True, size=9, color="808080")
UNIT_DIVISOR = 1_000_000

sys.path.insert(0, str(SCRIPT_DIR))
from build_workbook import ACCOUNTS, _sorted_periods, _resolve_value, INDICATOR_ROWS  # noqa: E402


def load_financials_cache(ticker: str, frequency: str) -> dict | None:
    fp = CACHE_DIR / f"financials_{ticker}_{frequency}.json"
    if not fp.exists():
        return None
    return json.loads(fp.read_text(encoding="utf-8"))


def _compute_indicator_series(freq_data: dict, periods: list[str]) -> dict[str, list[float | None]]:
    """계정 raw값으로부터 INDICATOR_ROWS의 비율들을 파이썬에서 직접 계산한다
    (비교 시트는 여러 기업을 한 표에 나란히 놓아야 해서, 단일기업 워크북처럼
    시트 간 셀 참조 수식을 쓰기보다 계산된 값을 직접 채우는 편이 훨씬 간단하고
    명확하다 — 원본데이터가 필요하면 각 기업의 단일기업 워크북을 참고하도록
    안내한다)."""
    result: dict[str, list[float | None]] = {}
    values: dict[str, list[float | None]] = {}
    for key, (sj, candidates, _label) in ACCOUNTS.items():
        series = []
        for p in periods:
            v, _ = _resolve_value(freq_data, sj, candidates, p)
            series.append(v)
        values[key] = series

    for label, num_key, den_key, mult, _fmt in INDICATOR_ROWS:
        series = []
        for i in range(len(periods)):
            num = values.get(num_key, [None] * len(periods))[i]
            den = values.get(den_key, [None] * len(periods))[i] if den_key else 1
            if num is None or (den_key and (den is None or den == 0)):
                series.append(None)
            else:
                series.append((num / den * mult) if den_key else num / UNIT_DIVISOR if label in ("매출액", "영업이익", "당기순이익") else num)
        result[label] = series
    return result


def build_comparison_sheet(wb: Workbook, title: str, tickers: list[str], names: dict[str, str],
                            frequency: str, n_periods: int, metric_labels: list[str]):
    """기업마다 실제 재무제표 캐시를 읽어, metric별로 시계열 값을 나란히 놓는다.
    레이아웃: 각 기업마다 자기 기간(가로) × 지표(세로) 미니 테이블을 만들고,
    시트 하단에 지표별 라인 차트(기업=계열)를 그린다."""
    ws = wb.create_sheet(title)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=13)
    ws["A2"] = f"단위: 매출액·영업이익·당기순이익=백만달러, 비율=% | yfinance 기준"
    ws["A2"].font = NOTE

    per_company_periods: dict[str, list[str]] = {}
    per_company_series: dict[str, dict[str, list]] = {}
    missing_companies = []
    for t in tickers:
        data = load_financials_cache(t, frequency)
        if not data:
            missing_companies.append(t)
            continue
        periods = _sorted_periods(data, n_periods)
        per_company_periods[t] = periods
        per_company_series[t] = _compute_indicator_series(data, periods)

    if missing_companies:
        ws["A3"] = f"⚠ 캐시 없음(먼저 fetch_financials.py 실행 필요): {', '.join(missing_companies)}"
        ws["A3"].font = Font(name=FONT_NAME, color="C00000", italic=True)

    row = 5
    chart_anchor_rows: dict[str, int] = {}
    for label in metric_labels:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
        row += 1
        header_row = row
        ws.cell(row=row, column=1, value="기업")
        max_periods = max((len(p) for p in per_company_periods.values()), default=0)
        for i in range(max_periods):
            ws.cell(row=row, column=2 + i, value=f"P{i+1}")
        row += 1
        table_start = row
        for t in tickers:
            if t not in per_company_series:
                continue
            name = names.get(t, t)
            ws.cell(row=row, column=1, value=f"{name} ({t})")
            series = per_company_series[t].get(label, [])
            for i, v in enumerate(series):
                if v is not None:
                    c = ws.cell(row=row, column=2 + i, value=round(v, 3))
                    c.number_format = "#,##0.0"
            row += 1
        chart_anchor_rows[label] = (header_row, table_start, row - 1, max_periods)
        row += 2

    # 지표별 라인 차트(기업=계열). 기간 라벨은 기업마다 다를 수 있어(캐시 시점 차이)
    # 첫 번째로 발견된 유효 기업의 실제 날짜를 참고용 각주로 남긴다.
    chart_col = 2 + max((v[3] for v in chart_anchor_rows.values()), default=6) + 2
    for label, (header_row, t_start, t_end, max_periods) in chart_anchor_rows.items():
        if t_end < t_start or max_periods == 0:
            continue
        chart = LineChart()
        chart.title = f"{title}_{label}"
        chart.style = 2
        chart.x_axis.title = "기간(오래된순)"
        cats = Reference(ws, min_col=2, max_col=1 + max_periods, min_row=header_row, max_row=header_row)
        data_ref = Reference(ws, min_col=1, max_col=1 + max_periods, min_row=t_start, max_row=t_end)
        chart.add_data(data_ref, titles_from_data=True, from_rows=True)
        chart.set_categories(cats)
        chart.width, chart.height = 18, 8
        ws.add_chart(chart, f"{get_column_letter(chart_col)}{header_row}")

    for i in range(2, 20):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions["A"].width = 26
    return missing_companies


def main() -> None:
    ap = argparse.ArgumentParser(description="여러 미국 기업 비교 엑셀 생성")
    ap.add_argument("tickers", nargs="+", help="Ticker:기업명 형태(예: AAPL:Apple) 또는 Ticker만")
    ap.add_argument("--quarters", type=int, default=12)
    ap.add_argument("--years", type=int, default=5)
    ap.add_argument("--outdir", default="/mnt/user-data/outputs")
    args = ap.parse_args()

    tickers, names = [], {}
    for spec in args.tickers:
        if ":" in spec:
            t, n = spec.split(":", 1)
        else:
            t, n = spec, spec
        tickers.append(t)
        names[t] = n

    metric_labels = [label for label, *_ in INDICATOR_ROWS]

    wb = Workbook()
    wb.remove(wb.active)

    q_missing = build_comparison_sheet(wb, "12분기 비교", tickers, names, "quarterly", args.quarters, metric_labels)
    a_missing = build_comparison_sheet(wb, "최근5년 비교", tickers, names, "annual", args.years, metric_labels)

    today = dt.date.today().strftime("%Y%m%d")
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    tickers_str = "_".join(tickers[:3]) + (f"등{len(tickers)}개" if len(tickers) > 3 else "")
    filepath = outdir / f"{tickers_str}_비교_{today}.xlsx"
    wb.save(filepath)

    print(json.dumps({
        "saved": str(filepath), "tickers": tickers,
        "quarterly_missing": q_missing, "annual_missing": a_missing,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()

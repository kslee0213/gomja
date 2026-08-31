#!/usr/bin/env python3
"""build_workbook.py — 미국 상장기업 재무 워크북 생성 (sec edgar 기반).

v3.0.0: 데이터 소스를 yfinance(비공식, yahoo 403 차단 위험)에서 sec edgar
공식 api(무료, 키불필요, 차단 위험 낮음)로 전환했다. 주가만 yfinance를
최소한으로 쓴다(fetch_extra_info.py). 감사 가능성 원칙(원본데이터 시트 +
수식 참조)은 v2.0.0과 동일하게 유지한다.
"""
import argparse
import datetime as dt
import json
import sys
from pathlib import path

from openpyxl import workbook
from openpyxl.chart import linechart, reference, series
from openpyxl.styles import alignment, border, font, patternfill, side
from openpyxl.utils import get_column_letter

x-script_dir = path(__file__).resolve().parent
cache_dir = x-script_dir.parent / "cache"
sys.path.insert(0, str(x-script_dir))
from fetch_financials import accounts, instant_keys, build_frequency_payload, load_companyfacts_cache  # noqa: e402

font_name = "맑은 고딕"
title_font = font(name=font_name, bold=true, size=14)
bold = font(name=font_name, bold=true)
green = font(name=font_name, color="006100")
blue = font(name=font_name, color="0000ff")
note = font(name=font_name, italic=true, size=9, color="808080")
header_fill = patternfill("solid", fgcolor="d9d9d9")
thin = side(style="thin", color="000000")
border = border(left=thin, right=thin, top=thin, bottom=thin)
unit_divisor = 1_000_000  # 표시 단위: 백만달러

sj_order = [
    ("income_statement", "손익계산서"),
    ("balance_sheet", "재무상태표"),
    ("cash_flow", "현금흐름표"),
]


def load_price_cache(ticker: str) -> dict | none:
    fp = cache_dir / f"price_{ticker.upper()}.json"
    if not fp.exists():
        return none
    return json.loads(fp.read_text(encoding="utf-8"))


def style_header(ws, row: int, min_col: int, max_col: int) -> none:
    for col in range(min_col, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = bold
        c.fill = header_fill
        c.alignment = alignment(horizontal="center")


def apply_border(ws, r1: int, r2: int, c1: int, c2: int) -> none:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


def write_raw_sheet(wb: workbook, ticker: str, q_payload: dict | none, a_payload: dict | none,
                     q_periods: list[str], a_periods: list[str]) -> dict:
    """원본데이터 시트: sec companyfacts에서 뽑은 값을 그대로 기록한다(달러,
    무환산). 다른 시트는 전부 이 시트를 수식으로 참조한다."""
    ws = wb.create_sheet("원본데이터")
    ws.sheet_state = "hidden"
    ws["a1"] = f"{ticker} — sec edgar xbrl 원자료(달러, 무환산)"
    ws["a1"].font = note

    cell_index: dict[tuple[str, str, str], str] = {}
    row = 3
    for freq, payload, periods in (("quarterly", q_payload, q_periods), ("annual", a_payload, a_periods)):
        if not payload or not periods:
            continue
        ws.cell(row=row, column=1, value=f"[{freq}]").font = bold
        row += 1
        ws.cell(row=row, column=1, value="계정")
        for i, p in enumerate(periods):
            ws.cell(row=row, column=2 + i, value=p)
        row += 1
        for key, (sj, _cand, _label) in accounts.items():
            ws.cell(row=row, column=1, value=f"{key} ({sj})")
            series = payload.get(sj, {}).get(key, {})
            for i, p in enumerate(periods):
                val = series.get(p)
                col = 2 + i
                if val is not none:
                    ws.cell(row=row, column=col, value=val)
                    col_letter = get_column_letter(col)
                    cell_index[(freq, key, p)] = f"'원본데이터'!${col_letter}${row}"
            row += 1
        row += 1

    ws.column_dimensions["a"].width = 32
    return cell_index


per_share_keys = {"eps(희석)"}  # 주당 지표는 백만달러 단위 환산 대상이 아니다(달러 그대로)


def build_statement_sheet(wb: workbook, sheet_name: str, freq: str, periods: list[str], cell_index: dict, hidden: bool = false):
    ws = wb.create_sheet(sheet_name)
    if hidden:
        ws.sheet_state = "hidden"
    ws["a1"] = "단위: 백만달러(usd millions) | sec edgar xbrl 기준, 원본데이터 시트 링크"
    ws["a1"].font = note

    header_row = 3
    ws.cell(row=header_row, column=1, value="구분")
    ws.cell(row=header_row, column=2, value="계정과목")
    labels = [p[:7] for p in periods]
    for i, lab in enumerate(labels):
        ws.cell(row=header_row, column=3 + i, value=lab)
    style_header(ws, header_row, 1, 2 + len(periods))

    row = header_row + 1
    account_row_map: dict[str, int] = {}
    for sj, sj_name in sj_order:
        keys = [k for k, (s, _, _) in accounts.items() if s == sj]
        ws.cell(row=row, column=1, value=sj_name).font = bold
        row += 1
        for key in keys:
            label = accounts[key][2]
            account_row_map[key] = row
            ws.cell(row=row, column=2, value=label)
            for i, p in enumerate(periods):
                ref = cell_index.get((freq, key, p))
                cell = ws.cell(row=row, column=3 + i)
                if ref:
                    if key in per_share_keys:
                        cell.value = f"=({ref})"
                        cell.number_format = "#,##0.00"
                    else:
                        cell.value = f"=({ref})/{unit_divisor}"
                        cell.number_format = "#,##0.0;(#,##0.0);-"
                    cell.font = green
                else:
                    cell.number_format = "#,##0.0;(#,##0.0);-"
            row += 1
        row += 1

    ws.column_dimensions["a"].width = 16
    ws.column_dimensions["b"].width = 24
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    ws.freeze_panes = "c4"
    apply_border(ws, header_row, row - 1, 1, 2 + len(periods))
    return account_row_map, labels


indicator_rows = [
    ("매출액", "매출액", none, 1, "#,##0.0"),
    ("영업이익", "영업이익", none, 1, "#,##0.0"),
    ("당기순이익", "당기순이익", none, 1, "#,##0.0"),
    ("매출총이익률(%)", "매출총이익", "매출액", 100, "0.0"),
    ("영업이익률(%)", "영업이익", "매출액", 100, "0.0"),
    ("순이익률(%)", "당기순이익", "매출액", 100, "0.0"),
    ("roe(%)", "당기순이익", "자본총계", 100, "0.0"),
    ("roa(%)", "당기순이익", "자산총계", 100, "0.0"),
    ("부채비율(%)", "부채총계", "자본총계", 100, "0.0"),
    ("유동비율(%)", "유동자산", "유동부채", 100, "0.0"),
    ("자기자본비율(%)", "자본총계", "자산총계", 100, "0.0"),
]


def build_indicator_sheet(wb: workbook, sheet_name: str, stmt_sheet: str, account_row_map: dict, period_labels: list[str]):
    ws = wb.create_sheet(sheet_name)
    n = len(period_labels)
    ws.cell(row=1, column=1, value="지표")
    for i, lab in enumerate(period_labels):
        ws.cell(row=1, column=3 + i, value=lab)
    style_header(ws, 1, 1, 2 + n)

    row_of: dict[str, int] = {}
    row = 2
    for label, num_key, den_key, mult, fmt in indicator_rows:
        ws.cell(row=row, column=2, value=label)
        row_of[label] = row
        num_row = account_row_map.get(num_key)
        den_row = account_row_map.get(den_key) if den_key else none
        for i in range(n):
            col_letter = get_column_letter(3 + i)
            cell = ws.cell(row=row, column=3 + i)
            if num_row and (den_key is none or den_row):
                num_ref = f"'{stmt_sheet}'!{col_letter}{num_row}"
                if den_key is none:
                    cell.value = f"={num_ref}"
                else:
                    den_ref = f"'{stmt_sheet}'!{col_letter}{den_row}"
                    cell.value = f"=iferror({num_ref}/{den_ref}*{mult},\"\")"
            cell.number_format = fmt
        row += 1
    apply_border(ws, 1, row - 1, 1, 2 + n)
    ws.column_dimensions["b"].width = 20
    return row_of


def x-embed_charts(wb: workbook, ws_ind, row_of: dict, n_periods: int, sheet_label: str):
    chart_specs = [
        ("매출액·영업이익·순이익 추이", ["매출액", "영업이익", "당기순이익"]),
        ("수익성 지표(%) 추이", ["매출총이익률(%)", "영업이익률(%)", "순이익률(%)"]),
        ("roe·roa(%) 추이", ["roe(%)", "roa(%)"]),
        ("건전성 지표(%) 추이", ["부채비율(%)", "유동비율(%)", "자기자본비율(%)"]),
    ]
    anchor_col_idx = 2 + n_periods + 2
    anchor_col = get_column_letter(anchor_col_idx)
    anchor_row = 2
    for title, metric_labels in chart_specs:
        chart = linechart()
        chart.title = f"{sheet_label}_{title}"
        chart.style = 2
        chart.y_axis.title = "값"
        chart.x_axis.title = "기간"
        cats = reference(ws_ind, min_col=3, max_col=2 + n_periods, min_row=1, max_row=1)
        for lab in metric_labels:
            r = row_of.get(lab)
            if not r:
                continue
            data = reference(ws_ind, min_col=2, max_col=2 + n_periods, min_row=r, max_row=r)
            chart.series.append(series(data, title_from_data=true))
        chart.set_categories(cats)
        chart.width, chart.height = 16, 8
        if chart.series:
            ws_ind.add_chart(chart, f"{anchor_col}{anchor_row}")
        anchor_row += 17


def build_investment_analysis_sheet(wb: workbook, ticker: str, company_name: str,
                                     a_row_of: dict, period_labels: list[str],
                                     price_data: dict | none, eps_ref: str | none, bvps_ref: str | none,
                                     rev_per_share_ref: str | none):
    """투자분석 시트. per/pbr/psr은 yfinance의 완제품 값을 그대로 믿지 않고,
    (주가) ÷ (sec 재무제표에서 뽑은 eps/bvps/주당매출)로 우리가 직접 수식
    계산한다 — 감사 가능성 원칙 유지."""
    ws = wb.create_sheet("투자분석")
    n = len(period_labels)
    ws["a1"] = f"투자분석 — {company_name} ({ticker})"
    ws["a1"].font = title_font
    ws["a3"] = "단위: 백만달러(usd millions), 주가는 달러 | 재무제표: sec edgar, 주가: yfinance(최소 사용)"
    ws["a3"].font = note
    row = 5

    def section(title: str):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = bold
        row += 1

    def ratio_header():
        nonlocal row
        ws.cell(row=row, column=1, value="지표")
        for i, lab in enumerate(period_labels):
            ws.cell(row=row, column=3 + i, value=lab)
        style_header(ws, row, 1, 2 + n)
        row += 1

    def copy_row(label: str, ind_row: int | none):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        if ind_row:
            for i in range(n):
                col_letter = get_column_letter(3 + i)
                ws.cell(row=row, column=3 + i, value=f"='지표_연간'!{col_letter}{ind_row}")
        apply_border(ws, row, row, 1, 2 + n)
        row += 1

    section("a. 재무비율 요약")
    ratio_header()
    for label in ["매출총이익률(%)", "영업이익률(%)", "순이익률(%)", "roe(%)", "roa(%)",
                  "부채비율(%)", "유동비율(%)", "자기자본비율(%)"]:
        copy_row(label, a_row_of.get(label))
    row += 1

    section("b. 위험 신호 점검")
    ratio_header()
    debt_row = a_row_of.get("부채비율(%)")
    curr_row = a_row_of.get("유동비율(%)")
    ws.cell(row=row, column=1, value="부채비율 200% 초과")
    for i in range(n):
        col_letter = get_column_letter(3 + i)
        if debt_row:
            ws.cell(row=row, column=3 + i, value=f"=if('지표_연간'!{col_letter}{debt_row}>200,\"⚠ 위험\",\"양호\")")
    apply_border(ws, row, row, 1, 2 + n)
    row += 1
    ws.cell(row=row, column=1, value="유동비율 100% 미만")
    for i in range(n):
        col_letter = get_column_letter(3 + i)
        if curr_row:
            ws.cell(row=row, column=3 + i, value=f"=if('지표_연간'!{col_letter}{curr_row}<100,\"⚠ 위험\",\"양호\")")
    apply_border(ws, row, row, 1, 2 + n)
    row += 2

    # --- c. 주가 연동 지표 (sec 재무제표 값 + 주가로 직접 계산) ---
    section("c. 주가 연동 지표 (최신 연도 기준, per/pbr/psr은 직접 수식 계산)")
    info = (price_data or {}).get("price", {})
    price = info.get("currentprice")
    missing_price = []
    fetched = str((price_data or {}).get("fetched_at", ""))[:10]
    ws.cell(row=row, column=1, value=f"현재가($){' · 기준 ' + fetched if fetched else ''}")
    if price is not none:
        ws.cell(row=row, column=3, value=price).number_format = "#,##0.00"
    else:
        missing_price.append("현재가")
    apply_border(ws, row, row, 1, 3)
    row += 1

    shares = info.get("sharesoutstanding")
    mktcap = info.get("marketcap")
    ws.cell(row=row, column=1, value="시가총액(백만달러)")
    if mktcap is not none:
        ws.cell(row=row, column=3, value=mktcap / unit_divisor).number_format = "#,##0.0"
    else:
        missing_price.append("시가총액")
    apply_border(ws, row, row, 1, 3)
    row += 1

    last_col = get_column_letter(2 + n)
    per_row = row
    ws.cell(row=row, column=1, value="per(배) = 현재가 ÷ eps")
    if price is not none and eps_ref:
        ws.cell(row=row, column=3, value=f"=iferror({price}/{eps_ref},\"\")").number_format = "0.00"
    else:
        missing_price.append("per(eps 없음)")
    apply_border(ws, row, row, 1, 3)
    row += 1

    ws.cell(row=row, column=1, value="pbr(배) = 현재가 ÷ 주당순자산(bvps)")
    if price is not none and bvps_ref:
        ws.cell(row=row, column=3, value=f"=iferror({price}/{bvps_ref},\"\")").number_format = "0.00"
    else:
        missing_price.append("pbr(bvps 없음, 상장주식수 필요)")
    apply_border(ws, row, row, 1, 3)
    row += 1

    ws.cell(row=row, column=1, value="psr(배) = 현재가 ÷ 주당매출")
    if price is not none and rev_per_share_ref:
        ws.cell(row=row, column=3, value=f"=iferror({price}/{rev_per_share_ref},\"\")").number_format = "0.00"
    else:
        missing_price.append("psr(주당매출 없음)")
    apply_border(ws, row, row, 1, 3)
    row += 1

    div_yield = info.get("dividendyield")
    ws.cell(row=row, column=1, value="배당수익률(%)")
    if div_yield is not none:
        ws.cell(row=row, column=3, value=div_yield * 100).number_format = "0.00"
    else:
        missing_price.append("배당수익률")
    apply_border(ws, row, row, 1, 3)
    row += 1

    if missing_price:
        ws.cell(row=row, column=1, value=f"※ 못 가져온 항목(재무제표 sec edgar는 정상, 주가/yfinance 관련만 영향): {', '.join(missing_price)}").font = font(name=font_name, italic=true, size=9, color="c00000")
        row += 1
    row += 1

    section("d. 간이 투자판단 (참고용 — 결정론적 규칙, 투자 조언 아님)")
    ws.cell(row=row, column=1, value="※ 재무비율 3개(부채비율/roe/유동비율)만 보는 간이 등급이며, 정성적 요인은 반영하지 않습니다.").font = note
    row += 1
    if debt_row and a_row_of.get("roe(%)") and curr_row:
        formula = (
            f"=if(and('지표_연간'!{last_col}{debt_row}<150,'지표_연간'!{last_col}{a_row_of['roe(%)']}>15,"
            f"'지표_연간'!{last_col}{curr_row}>120),\"a(우수)\","
            f"if(and('지표_연간'!{last_col}{debt_row}<200,'지표_연간'!{last_col}{a_row_of['roe(%)']}>8),\"b(양호)\",\"c(보통 이하 — 직접 확인 필요)\"))"
        )
        ws.cell(row=row, column=1, value="종합 등급(최신 기간)")
        ws.cell(row=row, column=3, value=formula)
        apply_border(ws, row, row, 1, 3)
    row += 2

    ws.column_dimensions["a"].width = 32
    ws.column_dimensions["b"].width = 4
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    return missing_price


def build_workbook(ticker: str, company_name: str, period: str, quarters: int, years: int, outdir: str) -> dict:
    cached = load_companyfacts_cache(ticker)
    if not cached:
        print(f"warning: {ticker} sec companyfacts 캐시가 없습니다. fetch_financials.py를 먼저 실행하세요.", file=sys.stderr)
        raw_facts = none
    else:
        raw_facts = cached["raw"]

    q_payload = q_periods = a_payload = a_periods = none
    if raw_facts:
        if period in ("quarterly", "both"):
            q_payload, q_periods = build_frequency_payload(raw_facts, "quarterly", quarters)
        if period in ("annual", "both"):
            a_payload, a_periods = build_frequency_payload(raw_facts, "annual", years)

    price_data = load_price_cache(ticker)

    wb = workbook()
    wb.remove(wb.active)

    cell_index = write_raw_sheet(wb, ticker, q_payload, a_payload, q_periods or [], a_periods or [])

    missing_price = []
    if q_periods:
        q_row_map, q_labels = build_statement_sheet(wb, "분기_재무제표", "quarterly", q_periods, cell_index)
        q_ind_row_of = build_indicator_sheet(wb, "지표_분기", "분기_재무제표", q_row_map, q_labels)
        x-embed_charts(wb, wb["지표_분기"], q_ind_row_of, len(q_labels), "분기")

    if a_periods:
        a_row_map, a_labels = build_statement_sheet(wb, "연간_재무제표", "annual", a_periods, cell_index)
        a_ind_row_of = build_indicator_sheet(wb, "지표_연간", "연간_재무제표", a_row_map, a_labels)
        x-embed_charts(wb, wb["지표_연간"], a_ind_row_of, len(a_labels), "연간")

        # per/pbr/psr 계산용 참조: 최신 연도의 eps, bvps(자본÷상장주식수), 주당매출(매출÷상장주식수)
        last_i = len(a_labels) - 1
        last_col = get_column_letter(3 + last_i)
        eps_row = a_row_map.get("eps(희석)")
        eps_ref = f"'연간_재무제표'!{last_col}{eps_row}" if eps_row else none
        shares = (price_data or {}).get("price", {}).get("sharesoutstanding")
        equity_row = a_row_map.get("자본총계")
        revenue_row = a_row_map.get("매출액")
        bvps_ref = rev_per_share_ref = none
        if shares:
            shares_millions = shares / unit_divisor
            if equity_row:
                bvps_ref = f"('연간_재무제표'!{last_col}{equity_row}/{shares_millions})"
            if revenue_row:
                rev_per_share_ref = f"('연간_재무제표'!{last_col}{revenue_row}/{shares_millions})"

        missing_price = build_investment_analysis_sheet(
            wb, ticker, company_name, a_ind_row_of, a_labels, price_data, eps_ref, bvps_ref, rev_per_share_ref,
        )

    desired_order = ["분기_재무제표", "연간_재무제표", "지표_분기", "지표_연간", "투자분석", "원본데이터"]
    wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]

    today = dt.date.today().strftime("%y%m%d")
    outdir_path = path(outdir)
    outdir_path.mkdir(parents=true, exist_ok=true)
    suffix = {"quarterly": "_분기", "annual": "_연간", "both": ""}[period]
    filename = f"{company_name}{suffix}_{today}.xlsx"
    filepath = outdir_path / filename
    wb.save(filepath)

    return {
        "saved": str(filepath),
        "ticker": ticker,
        "period": period,
        "quarters_filled": len(q_periods or []),
        "years_filled": len(a_periods or []),
        "missing_price_fields": missing_price,
    }


def main() -> none:
    ap = argparse.argumentparser(dex-scription="미국 상장기업 재무 엑셀 생성(sec edgar 기반)")
    ap.add_argument("ticker")
    ap.add_argument("company_name")
    ap.add_argument("--period", choices=["quarterly", "annual", "both"], default="both")
    ap.add_argument("--quarters", type=int, default=12)
    ap.add_argument("--years", type=int, default=5)
    ap.add_argument("--outdir", default="/mnt/user-data/outputs")
    args = ap.parse_args()

    result = build_workbook(args.ticker, args.company_name, args.period, args.quarters, args.years, args.outdir)
    print(json.dumps(result, ensure_ascii=false))


if __name__ == "__main__":
    main()

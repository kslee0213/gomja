"""
thesis-auditor: dart-financial-extractor / investment-thesis-writer 산출물 검수기.

이 스크립트는 "결정론적 규칙"만 검사한다 (웹 서치를 다시 하지 않는다).
값을 고치지 않고 발견/표시만 한다. 각 검사는 PASS / WARN / FAIL 로 판정된다.

사용법:
    python audit_workbook.py <xlsx 경로> \
        [--thesis-content thesis_content.json] \
        [--valuation-content valuation_content.json] \
        [--report-md out.md] \
        [--fail-on error|warn]

출력: stdout 에 JSON {"summary": {...}, "checks": [...]}, 그리고 (지정 시) 리포트 md.
종료코드: FAIL 있으면 2, WARN 만 있고 --fail-on warn 이면 1, 그 외 0.
"""
import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl 필요: pip install openpyxl"}), flush=True)
    sys.exit(3)

ERROR_CELL_RE = re.compile(r"#(REF|DIV/0|VALUE|NAME|NUM|NULL|N/A)!?")
POSITIVE_WORDS = ["안정적", "탄탄", "우량", "견고", "매우 건전", "재무구조가 우수"]
HIGH_PROFIT_WORDS = ["고수익", "높은 마진", "탁월한 수익", "매우 수익성"]
URL_RE = re.compile(r"^https?://", re.I)


class Auditor:
    def __init__(self, xlsx_path, thesis=None, valuation=None, missing_indicators=None):
        self.xlsx_path = Path(xlsx_path)
        self.wb = load_workbook(self.xlsx_path, data_only=False)
        self.wb_vals = load_workbook(self.xlsx_path, data_only=True)
        self.thesis = thesis or {}
        self.valuation = valuation or {}
        # build_workbook.py가 stdout JSON으로 남기는 missing_indicators 딕셔너리
        # (예: {"연간": [...], "투자분석(계정 매칭 실패)": [...]}) — 있으면 A4에서 그대로 반영
        self.missing_indicators = missing_indicators or {}
        self.checks = []

    def add(self, cid, group, status, msg, detail=None):
        self.checks.append({
            "id": cid, "group": group, "status": status,
            "message": msg, "detail": detail,
        })

    # ---------- helpers ----------
    def _sheet(self, name):
        return self.wb_vals[name] if name in self.wb_vals.sheetnames else None

    def _iter_cells(self, ws):
        for row in ws.iter_rows():
            for c in row:
                yield c

    def _find_sheets(self, pattern):
        """정규식 패턴에 매칭되는 시트명 목록을 (visible/hidden 무관) 반환한다."""
        rx = re.compile(pattern)
        return [s for s in self.wb_vals.sheetnames if rx.match(s)]

    def _detect_workbook_kind(self):
        """단일기업 워크북('투자분석' 시트 존재) vs 비교 워크북(build_comparison_workbook.py 산출물:
        '지표_{회사명}_연간'/'지표_{회사명}_분기' 같은 회사별 hidden 시트가 여러 개 존재하고
        '투자분석' 시트는 없음)을 구분한다.
        반환: ("single", None) | ("comparison", [회사명, ...]) | ("unknown", None)"""
        if self._sheet("투자분석") is not None:
            return "single", None
        annual_sheets = self._find_sheets(r"^지표_(.+)_연간$")
        quarter_sheets = self._find_sheets(r"^지표_(.+)_분기$")
        companies = set()
        for s in annual_sheets:
            companies.add(re.match(r"^지표_(.+)_연간$", s).group(1))
        for s in quarter_sheets:
            companies.add(re.match(r"^지표_(.+)_분기$", s).group(1))
        if companies:
            return "comparison", sorted(companies)
        return "unknown", None

    def _company_sheet(self, company, kind):
        """비교 워크북에서 회사별 재무/지표 시트명을 우선순위대로 찾아 반환한다.
        kind: "annual" | "quarterly" """
        if kind == "annual":
            candidates = [f"지표_{company}_연간", f"연간_{company}"]
        else:
            candidates = [f"지표_{company}_분기", f"분기_{company}"]
        for name in candidates:
            ws = self._sheet(name)
            if ws is not None:
                return ws, name
        return None, None

    # ---------- Group A: 결정론 무결성 ----------
    def check_A(self):
        # A1 오류 셀
        err_cells = []
        for ws in self.wb.worksheets:
            for c in self._iter_cells(ws):
                if isinstance(c.value, str) and ERROR_CELL_RE.search(c.value):
                    err_cells.append(f"{ws.title}!{c.coordinate}={c.value}")
        if err_cells:
            self.add("A1", "결정론무결성", "FAIL",
                     f"수식 오류 셀 {len(err_cells)}개 발견", err_cells[:20])
        else:
            self.add("A1", "결정론무결성", "PASS", "수식 오류 셀 없음")

        kind, companies = self._detect_workbook_kind()
        self._wb_kind, self._wb_companies = kind, companies

        if kind == "comparison":
            # build_comparison_workbook.py 산출물: "투자분석" 시트가 없고, 대신 회사마다
            # "지표_{회사명}_연간"/"지표_{회사명}_분기"(둘 다 hidden) 시트가 companies 수만큼 존재한다.
            # 단일기업용 A2 로직을 회사별로 반복 적용해 -회사명 접미사를 붙인 개별 체크로 리포트한다.
            for company in companies:
                self._check_balance_identity(company=company)
            # 현금흐름 3단 합계 검사는 "투자분석" 시트 G섹션(현금 순증감 3단 합계)에서만 계산되는데,
            # 비교 워크북에는 이 섹션 자체가 존재하지 않는다 → 검사 불가(N/A)임을 명시한다.
            self.add("A3", "결정론무결성", "WARN",
                     "비교 워크북(build_comparison_workbook.py 산출물)에는 '투자분석' 시트가 없어 "
                     "현금흐름 3단 합계 검사 대상 아님(N/A) — 개별 기업 워크북에서 별도 확인 필요")
        else:
            # A2 회계 항등식 (지표_연간 최신열에서 시도)
            self._check_balance_identity()
            # A3 현금흐름 3단 합계 vs 현금및현금성자산의증가
            self._check_cashflow_sum()

        # A4 계정 매칭 실패(missing_indicators) 반영
        self._check_missing_indicators()

    def _check_cashflow_sum(self):
        """투자분석 시트 G섹션의 '현금 순증감 (3단 합계, 검증용)' 행과
        지표_연간의 '현금및현금성자산의증가' 행을 비교한다."""
        ws_ia = self._sheet("투자분석") or self._sheet("투자분석_분기추정")
        ws_ind = self._sheet("지표_연간") or self._sheet("지표_연환산")
        if ws_ia is None or ws_ind is None:
            self.add("A3", "결정론무결성", "WARN",
                     "투자분석/지표_연간 시트를 찾지 못해 현금흐름 3단 합계 검사 생략")
            return
        r_sum = self._find_row_label(ws_ia, ["현금 순증감"])
        r_incr = self._find_row_label(ws_ind, ["현금및현금성자산의증가"])
        if r_sum is None or r_incr is None:
            self.add("A3", "결정론무결성", "WARN",
                     "'현금 순증감(3단 합계)' 또는 '현금및현금성자산의증가' 행을 못 찾아 검사 생략")
            return
        sum_val = self._last_numeric_in_row(ws_ia, r_sum)
        incr_val = self._last_numeric_in_row(ws_ind, r_incr)
        if sum_val is None or incr_val is None:
            self.add("A3", "결정론무결성", "WARN",
                     "현금흐름 3단 합계 비교에 필요한 값이 비어 있어 검사 생략")
            return
        denom = max(abs(sum_val), abs(incr_val), 1e-9)
        diff = abs(sum_val - incr_val) / denom
        if diff <= 0.05:
            self.add("A3", "결정론무결성", "PASS",
                     f"현금흐름 3단 합계({sum_val:,.1f}) ≈ 현금및현금성자산의증가({incr_val:,.1f}) (오차 {diff:.1%})")
        else:
            self.add("A3", "결정론무결성", "WARN",
                     f"현금흐름 3단 합계({sum_val:,.1f})와 현금및현금성자산의증가({incr_val:,.1f})의 "
                     f"괴리가 큼(오차 {diff:.1%}) — 환율 변동 등이 아니라면 계정 매칭 누락 의심",
                     {"3단합계": sum_val, "현금및현금성자산의증가": incr_val})

    def _check_missing_indicators(self):
        """build_workbook.py / build_comparison_workbook.py가 stdout에 남기는
        missing_indicators 딕셔너리를 (있다면) self.missing_indicators로 전달받아 그대로
        리포트에 반영한다. 두 가지 스키마를 모두 지원한다:
          - 단일기업(build_workbook.py):      {"연간": [...], "분기": [...], ...}
          - 비교워크북(build_comparison_workbook.py): {"회사명": {"연간": [...], "분기": [...]}, ...}
        워크북 자체에서도 지표 시트(단일기업의 "지표_*", 비교워크북의 "지표_{회사명}_연간/분기"
        전부 포함) 하단의 '계정명을 찾지 못해 비어 있습니다' 노트를 보조적으로 스캔해서,
        --missing-indicators 없이 워크북만 넘겨도 회사별로 분류된 결과를 보여준다."""
        # 시트명 -> 발견된 노트 좌표 목록 (스캔은 기존처럼 전체 워크북에 대해 수행)
        found_by_sheet = {}
        for ws in self.wb.worksheets:
            for c in self._iter_cells(ws):
                if isinstance(c.value, str) and "계정명을 찾지 못해" in c.value:
                    found_by_sheet.setdefault(ws.title, []).append(f"{ws.title}!{c.coordinate}")
        found = [coord for lst in found_by_sheet.values() for coord in lst]

        mi = getattr(self, "missing_indicators", None) or {}
        is_nested = isinstance(mi, dict) and mi and all(isinstance(v, dict) for v in mi.values())

        if not mi and not found:
            self.add("A4", "결정론무결성", "PASS", "계정 매칭 실패(missing_indicators) 없음")
            return

        if is_nested:
            # 비교 워크북 스키마: {"회사A": {"연간": [...], "분기": [...]}, "회사B": {...}}
            grand_total = 0
            for company, per_period in mi.items():
                total = sum(len(v) for v in per_period.values())
                grand_total += total
                if total == 0:
                    continue
                detail_parts = [f"{k}: {', '.join(v)}" for k, v in per_period.items() if v]
                msg = f"[{company}] 계정 매칭 실패 지표 {total}개 — " + " / ".join(detail_parts)
                self.add(f"A4-{company}", "결정론무결성", "WARN", msg, per_period)
            if grand_total == 0:
                self.add("A4", "결정론무결성", "PASS", "계정 매칭 실패(missing_indicators) 없음")
            return

        if mi:
            # 단일기업 스키마: {"연간": [...], "분기": [...], ...}
            total = sum(len(v) for v in mi.values()) if isinstance(mi, dict) else 0
            detail_parts = [f"{k}: {', '.join(v)}" for k, v in mi.items() if v]
            msg = f"계정 매칭 실패 지표 {total}개 발견 — " + " / ".join(detail_parts)
            self.add("A4", "결정론무결성", "WARN", msg, mi)
            return

        # --missing-indicators가 없어 워크북 내 노트만으로 판단하는 경우:
        # 비교 워크북이면 "지표_{회사명}_연간/분기" 시트별로 그룹핑해서 회사 단위로 보고한다.
        kind = getattr(self, "_wb_kind", None)
        if kind == "comparison" and found_by_sheet:
            per_company = {}
            for sheet_name, coords in found_by_sheet.items():
                m = re.match(r"^지표_(.+)_(연간|분기)$", sheet_name)
                company = m.group(1) if m else sheet_name
                per_company.setdefault(company, []).extend(coords)
            for company, coords in per_company.items():
                self.add(f"A4-{company}", "결정론무결성", "WARN",
                         f"[{company}] 워크북 내 '계정명을 찾지 못해' 노트 {len(coords)}건 발견 "
                         "— missing_indicators JSON을 --missing-indicators로 함께 넘기면 더 정확히 집계됩니다",
                         coords[:20])
        else:
            self.add("A4", "결정론무결성", "WARN",
                     f"워크북 내 '계정명을 찾지 못해 비어 있습니다' 노트 {len(found)}건 발견 "
                     "— missing_indicators JSON을 --missing-indicators로 함께 넘기면 더 정확히 집계됩니다",
                     found[:20])

    def _find_row_label(self, ws, labels):
        """A/B열 근처에서 라벨 텍스트가 들어간 행을 찾아 (row, first_value_col) 반환."""
        for row in ws.iter_rows():
            for c in row[:3]:
                if isinstance(c.value, str) and any(lb in c.value for lb in labels):
                    return c.row
        return None

    def _find_row_label_exact(self, ws, labels):
        """A/B열 근처에서 라벨과 '정확히' 같은 셀 값을 가진 행을 찾는다.
        '매출액' 같은 짧은 라벨이 'FCF마진(%, FCF/매출액)' 같은 다른 행의
        부분 문자열로 잘못 걸리는 것을 막기 위한 용도(그룹 Q에서 사용)."""
        for row in ws.iter_rows():
            for c in row[:3]:
                if isinstance(c.value, str) and c.value.strip() in labels:
                    return c.row
        return None

    def _find_label_row_with_header_exact(self, ws, labels, search_up=3):
        row = self._find_row_label_exact(ws, labels)
        if row is None:
            return None, None, []
        header_row = self._find_header_row_for_label(ws, row, search_up=search_up)
        headers = self._collect_period_headers(ws, header_row) if header_row else []
        return row, header_row, headers

    def _last_numeric_in_row(self, ws, row):
        vals = []
        for c in ws[row]:
            if isinstance(c.value, (int, float)):
                vals.append(c.value)
        return vals[-1] if vals else None

    def _row_values_numeric(self, ws, row, min_col=1):
        out = []
        for c in ws[row]:
            if c.column >= min_col and isinstance(c.value, (int, float)):
                out.append(c.value)
        return out

    def _find_header_row_for_label(self, ws, label_row, search_up=3):
        start = max(1, label_row - search_up)
        for r in range(label_row - 1, start - 1, -1):
            txts = [str(ws.cell(r, c).value or "") for c in range(1, min(ws.max_column, 12) + 1)]
            joined = " ".join(txts)
            if any(tok in joined for tok in ["Q1", "Q2", "Q3", "Q4", "(E)"]):
                return r
        return None

    def _collect_period_headers(self, ws, header_row, min_col=3):
        vals = []
        for c in range(min_col, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip():
                vals.append((c, v.strip()))
        return vals

    def _parse_q_header(self, label):
        m = re.search(r"(20\d{2})Q([1-4])(?:\(E\))?", str(label))
        if not m:
            return None
        return {"year": int(m.group(1)), "q": int(m.group(2)), "is_est": "(E)" in str(label)}

    def _sheet_has_estimate_headers(self, ws):
        for row in range(1, min(ws.max_row, 8) + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row, c).value
                if isinstance(v, str) and "(E)" in v and re.search(r"20\d{2}Q[1-4]", v):
                    return True
        return False

    def _find_label_row_with_header(self, ws, labels, search_up=3):
        row = self._find_row_label(ws, labels)
        if row is None:
            return None, None, []
        header_row = self._find_header_row_for_label(ws, row, search_up=search_up)
        headers = self._collect_period_headers(ws, header_row) if header_row else []
        return row, header_row, headers

    def _align_row_series_by_headers(self, ws, row, headers):
        vals = []
        for col, _ in headers:
            v = ws.cell(row, col).value
            vals.append(v if isinstance(v, (int, float)) else None)
        return vals

    def _check_balance_identity(self, company=None):
        """회계 항등식(자산=부채+자본) 검사.
        company가 주어지면(비교 워크북) 해당 회사의 "지표_{회사명}_연간"(없으면 "연간_{회사명}")
        시트에서 값을 찾고, 체크 ID를 "A2-{회사명}"으로 구분해 리포트한다.
        company가 없으면(단일기업 워크북) 기존과 동일하게 "A2"로 리포트한다."""
        cid = f"A2-{company}" if company else "A2"
        label = f"[{company}] " if company else ""

        if company:
            ws, sheet_name = self._company_sheet(company, "annual")
        else:
            # 연간 워크북엔 "지표_연간"/"연간_재무제표", 분기 연환산 워크북엔
            # "지표_연환산"/"연환산_재무제표"가 대신 존재한다 — 둘 다 찾아본다.
            ws = (
                self._sheet("지표_연간") or self._sheet("연간_재무제표")
                or self._sheet("지표_연환산") or self._sheet("연환산_재무제표")
            )

        if ws is None:
            self.add(cid, "결정론무결성", "WARN", f"{label}재무 시트를 찾지 못해 회계항등식 검사 생략")
            return
        r_asset = self._find_row_label(ws, ["자산총계"])
        r_liab = self._find_row_label(ws, ["부채총계"])
        r_eq = self._find_row_label(ws, ["자본총계"])
        if not (r_asset and r_liab and r_eq):
            self.add(cid, "결정론무결성", "WARN", f"{label}자산/부채/자본총계 행을 못 찾아 항등식 검사 생략")
            return
        a = self._last_numeric_in_row(ws, r_asset)
        l = self._last_numeric_in_row(ws, r_liab)
        e = self._last_numeric_in_row(ws, r_eq)
        if None in (a, l, e) or a == 0:
            self.add(cid, "결정론무결성", "WARN", f"{label}총계 값이 비어 항등식 검사 생략")
            return
        diff = abs(a - (l + e)) / abs(a)
        if diff <= 0.005:
            self.add(cid, "결정론무결성", "PASS", f"{label}자산≈부채+자본 (오차 {diff:.2%})")
        else:
            self.add(cid, "결정론무결성", "WARN",
                     f"{label}자산≠부채+자본 (오차 {diff:.2%}) — 계정 매칭 누락 의심",
                     {"asset": a, "liab": l, "equity": e})

    # ---------- Group B: 단위 일관성 ----------
    def check_B(self):
        ws = self._sheet("투자분석") or self._sheet("투자분석_분기추정")
        if ws is None:
            kind = getattr(self, "_wb_kind", None)
            if kind == "comparison":
                # build_comparison_workbook.py 산출물은 설계상 "투자분석" 시트(PER/시가총액/내재
                # 기대성장률 등)를 만들지 않는다(비교 시트는 재무지표 위주로만 다룸) — 이는 결함이
                # 아니라 스펙이므로 WARN이 아니라 정보성 PASS로 표시한다.
                self.add("B1", "단위일관성", "PASS",
                         "비교 워크북은 '투자분석' 시트를 만들지 않는 설계 — 단위 검사 해당 없음(N/A)")
                self.add("B2", "단위일관성", "PASS",
                         "비교 워크북은 '투자분석' 시트를 만들지 않는 설계 — 시가총액/주식수 검사 해당 없음(N/A)")
                self.add("B3", "단위일관성", "PASS",
                         "비교 워크북은 '투자분석' 시트를 만들지 않는 설계 — 내재 기대성장률 검사 해당 없음(N/A)")
            else:
                self.add("B1", "단위일관성", "WARN", "투자분석 시트 없음 — 단위 검사 생략")
                self.add("B2", "단위일관성", "WARN", "투자분석 시트 없음 — 시가총액/주식수 검사 생략")
                self.add("B3", "단위일관성", "WARN", "투자분석 시트 없음 — 내재 기대성장률 검사 생략")
            return
        # PER 상식 범위
        per_vals = self._collect_ratio_exact(ws, ["PER(배)"])
        if per_vals:
            bad = [v for v in per_vals if v is not None and (v < 0 or v > 300)]
            if bad:
                self.add("B1", "단위일관성", "WARN",
                         f"PER 상식범위(0~300) 밖 값 {len(bad)}개 — 단위 꼬임 의심", bad[:10])
            else:
                self.add("B1", "단위일관성", "PASS", f"PER 정상범위 ({len(per_vals)}개)")
        else:
            self.add("B1", "단위일관성", "WARN", "PER 값 없음(KRX 데이터 미수집일 수 있음)")

        # B2 시가총액/주가/상장주식수 단위 정합성
        self._check_marketcap_units(ws)
        # B3 시장 내재 기대성장률(%) vs 실제 CAGR(분수) 단위 혼동
        self._check_implied_growth_units(ws)

    def _check_marketcap_units(self, ws):
        """시가총액 ≈ 종가 × 상장주식수(억원 환산) 인지 자릿수 단위로 검사한다.
        build_workbook.py는 시가총액(원)을 UNIT_DIVISOR(1억)로 나눠 억원으로 맞추는데,
        이 변환이 빠지거나 중복 적용되면 시총이 억배/1억배 단위로 어긋난다."""
        # "종가"를 부분일치로 찾으면 "기준일(종가)"나 섹션 제목("...KRX 종가 기준")을
        # 잘못 집는다 — 반드시 정확일치로 찾는다.
        r_close = self._find_row_label_exact(ws, ["종가"])
        r_mktcap = self._find_row_label_exact(ws, ["시가총액"])
        r_per = self._find_row_label_exact(ws, ["PER(배)"])
        r_ni = None
        ws_ind = self._sheet("지표_연간") or self._sheet("지표_연환산")
        if ws_ind is not None:
            r_ni = self._find_row_label_exact(ws_ind, ["당기순이익"])

        if r_close is None or r_mktcap is None:
            self.add("B2", "단위일관성", "WARN", "종가/시가총액 행을 못 찾아 단위 정합성 검사 생략")
            return

        close_v = self._last_numeric_in_row(ws, r_close)
        mktcap_v = self._last_numeric_in_row(ws, r_mktcap)
        if close_v is None or mktcap_v is None or close_v == 0:
            self.add("B2", "단위일관성", "WARN", "종가/시가총액 값이 비어 있어 단위 정합성 검사 생략")
            return

        # 시가총액(억원) / 종가(원) = 상장주식수(억주?) 형태의 자릿수를 역산해 상식 범위인지 확인.
        # 시가총액이 억원 단위이므로 implied_shares(억주) = mktcap_v / close_v
        implied_shares_eok = mktcap_v / close_v
        # 상장주식수는 보통 수백만~수십억주 -> 억주 단위로는 약 0.01(=1백만주)~100(=100억주) 범위가 상식적.
        if implied_shares_eok <= 0 or not (1e-4 <= implied_shares_eok <= 1000):
            self.add("B2", "단위일관성", "FAIL",
                     f"시가총액/종가로 역산한 상장주식수가 상식 범위를 크게 벗어남"
                     f"(역산값 ≈ {implied_shares_eok:,.4f}억주) — UNIT_DIVISOR(억원 환산) 누락/중복 의심",
                     {"종가": close_v, "시가총액(억원)": mktcap_v, "역산_상장주식수(억주)": implied_shares_eok})
            return

        # 추가로 PER × 당기순이익(억원) ≈ 시가총액(억원)인지도 자릿수로 교차검산
        if r_per is not None and r_ni is not None:
            per_v = self._last_numeric_in_row(ws, r_per)
            ni_v = self._last_numeric_in_row(ws_ind, r_ni)
            if per_v is not None and ni_v is not None and ni_v != 0:
                implied_mktcap = per_v * ni_v
                denom = max(abs(implied_mktcap), abs(mktcap_v), 1e-9)
                diff = abs(implied_mktcap - mktcap_v) / denom
                if diff > 0.5:
                    self.add("B2", "단위일관성", "WARN",
                             f"PER×당기순이익({implied_mktcap:,.1f}억원)과 시가총액({mktcap_v:,.1f}억원)의 "
                             f"괴리가 큼(오차 {diff:.1%}) — 단위 꼬임 의심",
                             {"PER×당기순이익": implied_mktcap, "시가총액": mktcap_v})
                    return
        self.add("B2", "단위일관성", "PASS",
                 f"시가총액/종가/PER 간 자릿수 정합 확인됨(역산 상장주식수 ≈ {implied_shares_eok:,.4f}억주)")

    def _check_implied_growth_units(self, ws):
        """시장 내재 기대성장률 = (PER-8.5)/2 는 '%포인트' 단위 관례인 반면,
        thesis content의 revenue_growth_assumptions 등은 '분수'(0.10=10%) 관례다.
        두 값을 그대로 비교/차감하면 자릿수(약 100배)가 어긋난다 — 이를 검사한다."""
        r_per = self._find_row_label_exact(ws, ["PER(배)"])
        per_v = self._last_numeric_in_row(ws, r_per) if r_per is not None else None
        if per_v is None:
            self.add("B3", "단위일관성", "WARN", "PER 값이 없어 시장 내재 기대성장률 검사 생략")
            return
        implied_growth_pct = (per_v - 8.5) / 2  # %포인트 단위

        lst = self.thesis.get("revenue_growth_assumptions")
        if not lst or not isinstance(lst, list):
            self.add("B3", "단위일관성", "WARN",
                     f"내재 기대성장률(계산값 {implied_growth_pct:.1f}%p)과 비교할 "
                     "revenue_growth_assumptions 가 없어 단위 교차검사 생략")
            return
        avg_assump = sum(lst) / len(lst)
        # avg_assump가 분수(0.x) 관례라면 %p 환산은 *100. 만약 실수로 %단위 숫자(예: 10)가
        # 섞여 있으면(=분수 관례를 어김) *100 하지 않은 값 그대로 비교했을 때 자릿수가 어긋난다.
        avg_assump_pct = avg_assump * 100  # 정상 관례라면 이렇게 %p로 변환해서 비교해야 함
        ratio = (abs(avg_assump_pct) / abs(implied_growth_pct)) if implied_growth_pct != 0 else None
        # 분수 관례를 어기고 이미 %단위(예: 10)가 들어간 경우 avg_assump 자체가 %p스케일과 비슷해짐
        # → avg_assump(변환 전)와 implied_growth_pct가 비슷해지는 역설적 상황이 되어 혼동 신호로 삼는다.
        if any(v > 1 for v in lst):
            self.add("B3", "단위일관성", "WARN",
                     f"revenue_growth_assumptions에 1을 초과하는 값 존재({[v for v in lst if v > 1][:5]}) — "
                     "분수(0.10=10%) 관례가 아니라 %단위 숫자가 섞였을 가능성(단위 혼동)",
                     {"revenue_growth_assumptions": lst})
            return
        if ratio is not None and (ratio > 100 or ratio < 0.01):
            self.add("B3", "단위일관성", "WARN",
                     f"시장 내재 기대성장률({implied_growth_pct:.1f}%p)과 가정 평균 성장률"
                     f"({avg_assump_pct:.1f}%p)의 자릿수 차이가 비정상(비율 {ratio:.1f}배) — "
                     "%p vs 분수 단위 혼동 여부 확인 필요",
                     {"implied_growth_pct": implied_growth_pct, "avg_assump_pct": avg_assump_pct})
        else:
            self.add("B3", "단위일관성", "PASS",
                     f"내재 기대성장률({implied_growth_pct:.1f}%p) vs 가정 평균({avg_assump_pct:.1f}%p) "
                     "자릿수 정상 범위")

    def _collect_ratio(self, ws, labels):
        row = self._find_row_label(ws, labels)
        if row is None:
            return []
        return [c.value for c in ws[row] if isinstance(c.value, (int, float))]

    def _collect_ratio_exact(self, ws, labels):
        row = self._find_row_label_exact(ws, labels)
        if row is None:
            return []
        return [c.value for c in ws[row] if isinstance(c.value, (int, float))]

    # ---------- Group C: 정량 vs 정성 모순 ----------
    def check_C(self):
        ws = self._sheet("투자분석")
        grades = self._read_grades(ws) if ws else {}

        fc = " ".join(str(self.thesis.get(k, "")) for k in
                      ["final_conclusion_text", "risk", "competitive_advantage",
                       "profitability_prediction_text"])
        # C1 재무건전성 D/E인데 긍정어
        health = grades.get("재무건전성")
        if health in ("D", "E") and any(w in fc for w in POSITIVE_WORDS):
            self.add("C1", "정량정성모순", "WARN",
                     f"재무건전성 {health}등급인데 텍스트에 긍정 서술 존재 — 모순 확인 필요")
        else:
            self.add("C1", "정량정성모순", "PASS", "재무건전성 등급-서술 방향 모순 없음")

        prof = grades.get("수익성")
        if prof in ("D", "E") and any(w in fc for w in HIGH_PROFIT_WORDS):
            self.add("C1b", "정량정성모순", "WARN",
                     f"수익성 {prof}등급인데 고수익 서술 존재 — 모순 확인 필요")

        # C2 위험신호 여러개 vs 리스크 텍스트 부실
        n_risk = self._count_risk_flags(ws) if ws else 0
        risk_txt = str(self.thesis.get("risk", ""))
        if n_risk >= 2 and len(risk_txt.strip()) < 40:
            self.add("C2", "정량정성모순", "WARN",
                     f"위험신호 {n_risk}건인데 리스크 텍스트가 {len(risk_txt.strip())}자로 부실")
        else:
            self.add("C2", "정량정성모순", "PASS",
                     f"위험신호({n_risk})-리스크서술 정합")

        # C3 기대수익률 과도
        er = self.thesis.get("expected_annual_return_pct")
        if isinstance(er, (int, float)):
            if er > 30:
                self.add("C3", "정량정성모순", "WARN",
                         f"기대연수익률 {er}% 과도 — 근거 확인 필요")
            else:
                self.add("C3", "정량정성모순", "PASS", f"기대연수익률 {er}% 합리적 범위")

    def _read_grades(self, ws):
        """N섹션 등급 읽기 (라벨 옆 셀에서 A~E 단일문자 탐색)."""
        grades = {}
        keymap = {"재무건전성": "재무건전성", "수익성": "수익성", "성장성": "성장성"}
        for row in ws.iter_rows():
            label = None
            for c in row[:4]:
                if isinstance(c.value, str):
                    for k in keymap:
                        if k in c.value:
                            label = keymap[k]
            if label:
                for c in row:
                    if isinstance(c.value, str) and re.fullmatch(r"[A-E]", c.value.strip()):
                        grades[label] = c.value.strip()
                        break
        return grades

    def _count_risk_flags(self, ws):
        n = 0
        for c in self._iter_cells(ws):
            if isinstance(c.value, str) and "위험" in c.value and ("⚠" in c.value or "위험" == c.value.strip()):
                n += 1
        return n

    # ---------- Group D: 가정 경계값 ----------
    def check_D(self):
        v = self.valuation
        if not v:
            self.add("D0", "가정경계값", "WARN", "valuation content 없음 — DCF 가정 검사 생략")
        else:
            dr = v.get("discount_rate_pct")
            tg = v.get("terminal_growth_pct")
            if isinstance(dr, (int, float)) and isinstance(tg, (int, float)):
                spread = dr - tg
                if spread <= 0:
                    self.add("D1", "가정경계값", "FAIL",
                             f"할인율({dr}%)≤영구성장률({tg}%) — DCF 발산")
                elif spread < 3:
                    self.add("D1", "가정경계값", "WARN",
                             f"할인율-영구성장률={spread}%p<3 — TV 발산 위험")
                else:
                    self.add("D1", "가정경계값", "PASS", f"할인율-영구성장률={spread}%p")
                if isinstance(tg, (int, float)) and tg > 4:
                    self.add("D2", "가정경계값", "WARN", f"영구성장률 {tg}%>4% (장기GDP 상단 초과)")
            wacc = v.get("wacc_pct")
            for nm, val in [("wacc", wacc), ("discount_rate", dr)]:
                if isinstance(val, (int, float)) and not (3 <= val <= 20):
                    self.add("D3", "가정경계값", "WARN", f"{nm}={val}% 상식범위(3~20%) 밖")
            self._check_assumption_list(v, "owner_earnings_growth_assumptions",
                                        "projection_years", "valuation")

        t = self.thesis
        if t:
            self._check_assumption_list(t, "revenue_growth_assumptions",
                                        "projection_years", "thesis",
                                        rationale_field="growth_prediction_text")

    def _check_assumption_list(self, content, list_field, years_field, tag, rationale_field=None):
        lst = content.get(list_field)
        yrs = content.get(years_field)
        if lst is None:
            self.add(f"D4-{tag}", "가정경계값", "WARN", f"{list_field} 없음")
            return
        if isinstance(yrs, int) and len(lst) != yrs:
            self.add(f"D4-{tag}", "가정경계값", "FAIL",
                     f"{list_field} 길이({len(lst)}) ≠ {years_field}({yrs})")
        else:
            self.add(f"D4-{tag}", "가정경계값", "PASS",
                     f"{list_field} 길이 {len(lst)} 정합")
        if rationale_field is not None:
            if not str(content.get(rationale_field, "")).strip():
                self.add(f"D4r-{tag}", "가정경계값", "WARN",
                         f"{list_field} 가정에 대한 근거({rationale_field}) 비어 있음")

    # ---------- Group E: 출처/재현성 ----------
    def check_E(self):
        for tag, content in [("thesis", self.thesis), ("valuation", self.valuation)]:
            if not content:
                continue
            srcs = content.get("sources", [])
            if not srcs:
                self.add(f"E1-{tag}", "출처재현성", "WARN",
                         f"{tag} content 에 sources 비어 있음 — 웹 리서치 출처 없음")
            else:
                bad = [s for s in srcs if not (isinstance(s, str) and URL_RE.match(s))]
                if bad:
                    self.add(f"E1-{tag}", "출처재현성", "WARN",
                             f"{tag} sources 에 URL 형식 아닌 항목 {len(bad)}개", bad[:5])
                else:
                    self.add(f"E1-{tag}", "출처재현성", "PASS",
                             f"{tag} sources {len(srcs)}개 URL 정상")
            # E3 회사 일치
            cn = content.get("company_name")
            if cn and cn not in self.xlsx_path.stem:
                self.add(f"E3-{tag}", "출처재현성", "FAIL",
                         f"{tag} company_name('{cn}')가 파일명('{self.xlsx_path.stem}')과 불일치")


    # ---------- Group Q: 분기 연환산 체크 ----------
    def check_Q(self):
        ws_q = self._sheet("투자분석_분기추정")
        ws_ann = self._sheet("연환산_재무제표")
        if ws_q is None and ws_ann is None:
            self.add("Q0", "분기연환산", "PASS", "분기 연환산 산출물이 없는 워크북(연간 전용 또는 비교 워크북) — 그룹 Q 해당 없음(N/A)")
            return
        if ws_q is None or ws_ann is None:
            self.add("Q0", "분기연환산", "FAIL", "투자분석_분기추정/연환산_재무제표 중 하나가 없어 분기 연환산 산출물이 불완전")
            return

        # Q1: FY_E 산식 정합(구조적 검증)
        if self._sheet_has_estimate_headers(ws_ann):
            row, header_row, headers = self._find_label_row_with_header(ws_ann, ["매출액", "영업이익", "당기순이익"], search_up=30)
            parsed = [self._parse_q_header(h) for _, h in headers]
            ok = bool(headers) and all(p is not None and p["is_est"] for p in parsed)
            if ok:
                self.add("Q1", "분기연환산", "PASS", f"연환산_재무제표가 분기별 FY_E 헤더({len(headers)}개, '(E)' 표기)로 구성됨")
            else:
                self.add("Q1", "분기연환산", "WARN", "연환산_재무제표의 FY_E 헤더 구조를 완전히 확인하지 못함")
        else:
            self.add("Q1", "분기연환산", "FAIL", "연환산_재무제표에서 FY_E '(E)' 헤더를 찾지 못함")

        # Q2: TTM vs FY_E 괴리 경고
        # 주의: "투자분석_분기추정"에는 "매출액" 단독 행이 없다(비율만 표시됨,
        # 예: "FCF마진(%, FCF/매출액)"). 부분일치로 그 행을 잘못 집으면 안 되므로
        # 원자료 격인 "지표_연환산"(정확일치로 "매출액" 행이 실제로 존재)을 쓴다.
        ws_ind_ann = self._sheet("지표_연환산") or ws_q
        row_rev_q, hdr_rev_q, headers_q = self._find_label_row_with_header_exact(ws_ind_ann, ["매출액"], search_up=30)
        row_rev_a, hdr_rev_a, headers_a = self._find_label_row_with_header_exact(ws_ann, ["매출액"], search_up=30)
        if row_rev_q and row_rev_a and headers_q and headers_a:
            labels_q = [h for _, h in headers_q]
            labels_a = [h for _, h in headers_a]
            common = [lab for lab in labels_q if lab in labels_a]
            diffs = []
            for lab in common:
                cq = next(c for c, h in headers_q if h == lab)
                ca = next(c for c, h in headers_a if h == lab)
                vq = ws_ind_ann.cell(row_rev_q, cq).value
                va = ws_ann.cell(row_rev_a, ca).value
                if isinstance(vq, (int, float)) and isinstance(va, (int, float)) and max(abs(vq), abs(va)) > 1e-9:
                    diffs.append(abs(vq - va) / max(abs(vq), abs(va)))
            if diffs:
                worst = max(diffs)
                if worst <= 0.25:
                    self.add("Q2", "분기연환산", "PASS", f"투자분석_분기추정 주요 flow 지표가 연환산_재무제표와 대체로 정합(최대 괴리 {worst:.1%})")
                else:
                    self.add("Q2", "분기연환산", "WARN", f"투자분석_분기추정 vs 연환산_재무제표 간 flow 지표 괴리 큼(최대 {worst:.1%}) — TTM/FY_E 혼용 여부 확인 필요")
            else:
                self.add("Q2", "분기연환산", "WARN", "TTM vs FY_E 비교용 수치가 부족해 괴리 검사 생략")
        else:
            self.add("Q2", "분기연환산", "WARN", "매출액 행/헤더를 찾지 못해 TTM vs FY_E 괴리 검사 생략")

        # Q3: 성장률 캡(±300%) 적용 여부 태그 검증(간접)
        headers = []
        for row in range(1, min(ws_ann.max_row, 6) + 1):
            for c in range(3, ws_ann.max_column + 1):
                v = ws_ann.cell(row, c).value
                if isinstance(v, str) and re.search(r"20\d{2}Q[1-4]\(E\)", v):
                    headers.append(v)
        if headers:
            self.add("Q3", "분기연환산", "PASS", f"연환산 추정 컬럼 라벨 {len(headers)}개 확인 — capped/sign-flip 추정 결과가 독립 연환산 시트로 분리됨")
        else:
            self.add("Q3", "분기연환산", "WARN", "연환산 추정 컬럼 라벨을 찾지 못해 성장률 캡 적용 구조 검증 제한")

        # Q4: 적자전환(sign_flip) 폴백 정상 동작(간접)
        title_blob = " ".join(str(ws_q.cell(r, c).value or "") for r in range(1, min(6, ws_q.max_row)+1) for c in range(1, min(6, ws_q.max_column)+1))
        if "추정" in title_blob or "FY_E" in title_blob or "TTM" in title_blob:
            self.add("Q4", "분기연환산", "PASS", "분기추정/FY_E/TTM 안내 문구 존재 — sign_flip 포함 추정 폴백 경로의 결과 시트로 식별 가능")
        else:
            self.add("Q4", "분기연환산", "WARN", "분기추정/FY_E/TTM 안내 문구가 약해 sign_flip 폴백 결과 시트 식별성이 낮음")

        # Q5: PER/PSR이 연환산 분모 기준인지
        per_vals = self._collect_ratio(ws_q, ["PER(", "PER"])
        psr_vals = self._collect_ratio(ws_q, ["PSR(", "PSR"])
        # "투자분석_분기추정"에는 "매출액" 단독 행이 없으므로(Q2와 동일한 이유),
        # "지표_연환산"을 정확일치로 조회한다.
        row_sales_q, _, headers_sales_q = self._find_label_row_with_header_exact(ws_ind_ann, ["매출액"], search_up=30)
        row_sales_a, _, headers_sales_a = self._find_label_row_with_header_exact(ws_ann, ["매출액"], search_up=30)
        aligned = False
        if row_sales_q and row_sales_a and headers_sales_q and headers_sales_a:
            common = [lab for _, lab in headers_sales_q if lab in [x[1] for x in headers_sales_a]]
            aligned = len(common) >= max(1, min(len(headers_sales_q), len(headers_sales_a)) // 2)
        if per_vals or psr_vals:
            bad_per = [v for v in per_vals if v is not None and (v < 0 or v > 300)]
            bad_psr = [v for v in psr_vals if v is not None and (v < 0 or v > 100)]
            if bad_per or bad_psr:
                self.add("Q5", "분기연환산", "WARN", f"분기추정 PER/PSR에 상식범위 밖 값 존재(PER {len(bad_per)}개, PSR {len(bad_psr)}개) — 연환산 분모 기준 여부 확인 필요", {"PER_bad": bad_per[:5], "PSR_bad": bad_psr[:5]})
            elif aligned:
                self.add("Q5", "분기연환산", "PASS", f"분기추정 PER/PSR 정상범위이며 매출액 헤더가 연환산_재무제표와 정렬됨 — 연환산 분모 사용 가능성 높음")
            else:
                self.add("Q5", "분기연환산", "WARN", "분기추정 PER/PSR 정상범위이나 연환산 분모 정렬성은 부분 확인만 됨")
        else:
            self.add("Q5", "분기연환산", "WARN", "분기추정 PER/PSR 값이 없어 연환산 분모 기준 검사 생략")


    # ---------- run ----------
    def run(self):
        for m in (self.check_A, self.check_B, self.check_C, self.check_D, self.check_E, self.check_Q):
            try:
                m()
            except Exception as e:
                self.add(m.__name__, "실행오류", "WARN", f"검사 중 예외: {e}")
        n = {"PASS": 0, "WARN": 0, "FAIL": 0}
        for c in self.checks:
            n[c["status"]] = n.get(c["status"], 0) + 1
        verdict = ("FAIL" if n["FAIL"] else ("WARN" if n["WARN"] else "PASS"))
        return {
            "summary": {
                "file": str(self.xlsx_path.name),
                "verdict": verdict,
                "counts": n,
                "delivery_advice": (
                    "전달 전 수정 필요 (FAIL 존재)" if n["FAIL"]
                    else "전달 가능하나 WARN 항목 사람 확인 권장" if n["WARN"]
                    else "자동 검수 통과"
                ),
                "generated_at": datetime.now().isoformat(timespec="seconds"),
            },
            "checks": self.checks,
        }


def to_markdown(result):
    s = result["summary"]
    lines = [f"# 검수 리포트 — {s['file']}", "",
             f"- 판정: **{s['verdict']}**  (PASS {s['counts']['PASS']} / WARN {s['counts']['WARN']} / FAIL {s['counts']['FAIL']})",
             f"- 권고: {s['delivery_advice']}",
             f"- 생성시각: {s['generated_at']}", "",
             "| ID | 그룹 | 판정 | 내용 |", "|---|---|---|---|"]
    for c in result["checks"]:
        emoji = {"PASS": "✅", "WARN": "⚠️", "FAIL": "❌"}[c["status"]]
        lines.append(f"| {c['id']} | {c['group']} | {emoji} {c['status']} | {c['message']} |")
    fails = [c for c in result["checks"] if c["status"] == "FAIL"]
    warns = [c for c in result["checks"] if c["status"] == "WARN"]
    if fails:
        lines += ["", "## ❌ 반드시 수정 (FAIL)"]
        for c in fails:
            lines.append(f"- **{c['id']}**: {c['message']}" +
                         (f"  \n  세부: `{c['detail']}`" if c.get("detail") else ""))
    if warns:
        lines += ["", "## ⚠️ 사람 확인 권장 (WARN — 비결정론 항목 포함)"]
        for c in warns:
            lines.append(f"- **{c['id']}**: {c['message']}")
    lines += ["", "> 이 검수는 결정론적 규칙만 자동 검사합니다. 웹 서치 사실관계의 진위나 "
              "산업 전망의 타당성은 자동 검수 대상이 아니며 사람 검토가 필요합니다."]
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--thesis-content")
    ap.add_argument("--valuation-content")
    ap.add_argument("--missing-indicators",
                     help="build_workbook.py 실행 시 stdout JSON의 missing_indicators 필드를 "
                          "그대로 파일로 저장해 전달하면 A4 검사가 더 정확해진다.")
    ap.add_argument("--report-md")
    ap.add_argument("--fail-on", choices=["error", "warn"], default="error")
    args = ap.parse_args()

    def load_json(p):
        if not p:
            return None
        return json.loads(Path(p).read_text(encoding="utf-8"))

    auditor = Auditor(args.xlsx, load_json(args.thesis_content),
                      load_json(args.valuation_content),
                      load_json(args.missing_indicators))
    result = auditor.run()
    print(json.dumps(result, ensure_ascii=False, indent=2))

    if args.report_md:
        Path(args.report_md).write_text(to_markdown(result), encoding="utf-8")

    counts = result["summary"]["counts"]
    if counts["FAIL"]:
        sys.exit(2)
    if counts["WARN"] and args.fail_on == "warn":
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()

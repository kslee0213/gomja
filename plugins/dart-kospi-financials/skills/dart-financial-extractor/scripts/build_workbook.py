"""
fetch_financials.py가 cache/ 에 쌓아 둔 원자료를 읽어
'분기_재무제표'(최근 12분기)와 '연간_재무제표'(최근 5개년) 2개 시트 엑셀을 만든다.

핵심 설계:
  - 재무상태표(BS)는 시점 데이터이므로 각 분기 말 보고서 값을 그대로 링크한다.
  - 손익계산서/포괄손익계산서/현금흐름표(IS/CIS/CF)는 흐름 데이터이므로
    2분기·4분기는 반드시 엑셀 수식으로 계산한다(하드코딩 금지):
        2분기 = 반기보고서 thstrm_amount - 1분기보고서 thstrm_amount
        4분기 = 사업보고서 thstrm_amount - 3분기보고서 thstrm_add_amount(9개월 누적)
    1분기·3분기는 원본 값을 그대로 링크한다.
  - 모든 원본 값은 '원본데이터' 시트에 먼저 적재한 뒤, 화면에 보이는 시트의 셀은
    그 원본데이터 시트를 참조하는 수식으로 채운다. → 감사(audit) 가능, 값 변경 시 자동 재계산.

사용법:
    python build_workbook.py <corp_code> <company_name> [--years 5] [--quarters 12] [--outdir /mnt/user-data/outputs]

전제:
    같은 corp_code에 대해 fetch_financials.py를 먼저 필요한 (연도, reprt_code) 조합만큼
    실행해서 cache/ 에 JSON이 쌓여 있어야 한다.
"""
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.drawing.text import CharacterProperties, ParagraphProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
CACHE_DIR = SCRIPT_DIR.parent / "cache"

SJ_ORDER = [("BS", "재무상태표"), ("IS", "손익계산서"), ("CIS", "포괄손익계산서"), ("CF", "현금흐름표")]
REPRT_NAMES = {"11013": "1분기보고서", "11012": "반기보고서", "11014": "3분기보고서", "11011": "사업보고서"}
FLOW_TYPES = {"IS", "CIS", "CF"}  # 누적/차감 로직이 필요한 유형
FONT_NAME = "Arial"

BLUE = Font(name=FONT_NAME, color="0000FF")  # 하드코딩 입력값
BLACK = Font(name=FONT_NAME, color="000000")  # 수식
GREEN = Font(name=FONT_NAME, color="008000")  # 다른 시트 링크
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
BOLD = Font(name=FONT_NAME, bold=True)


def load_cache(corp_code: str) -> dict:
    """cache/{corp_code}_{year}_{reprt}_{fs_div}.json 전부를 로드해
    reports[year][reprt_code] = data 형태로 반환한다."""
    reports: dict[str, dict[str, dict]] = {}
    for fp in CACHE_DIR.glob(f"{corp_code}_*.json"):
        data = json.loads(fp.read_text(encoding="utf-8"))
        if data.get("status") != "000":
            continue
        year = data["bsns_year"]
        reprt = data["reprt_code"]
        reports.setdefault(year, {})[reprt] = data
    return reports


def build_quarter_plan(reports: dict, n_quarters: int) -> list[dict]:
    """사용 가능한 데이터로부터 최근 n_quarters개 분기 계획을 오래된 순으로 만든다.
    각 원소: {year, q, needs: {1Q:[year,'11013'], H1:[year,'11012'], 3Q:[year,'11014'], FY:[year,'11011']}}
    """
    years_desc = sorted(reports.keys(), reverse=True)
    plan = []
    for year in years_desc:
        y_reports = reports[year]
        # 4분기 -> 1분기 역순으로 채운다
        if "11011" in y_reports and "11014" in y_reports:
            plan.append({"year": year, "q": 4, "fy": y_reports["11011"], "q3": y_reports["11014"]})
        if "11014" in y_reports:
            plan.append({"year": year, "q": 3, "q3": y_reports["11014"]})
        if "11012" in y_reports and "11013" in y_reports:
            plan.append({"year": year, "q": 2, "h1": y_reports["11012"], "q1": y_reports["11013"]})
        if "11013" in y_reports:
            plan.append({"year": year, "q": 1, "q1": y_reports["11013"]})
        if len(plan) >= n_quarters:
            break
    plan.sort(key=lambda p: (p["year"], p["q"]))
    return plan[-n_quarters:]


def collect_accounts(periods: list[dict], sj_div: str, period_key_fn) -> list[dict]:
    """여러 기간의 데이터에서 계정과목 유니온을 ord 순으로 만든다.

    ⚠️ 계정명(account_nm) 기준으로 병합한다(account_id 기준이 아님). DART는
    같은 개념의 계정(예: "매출채권")에도 보고서/분기마다 다른 XBRL 계정ID를
    배정하는 경우가 있다. account_id로만 묶으면 같은 계정명이 서로 다른 행으로
    쪼개져서, 어느 분기는 A행에만 값이 있고 다른 분기는 B행에만 값이 있는
    식으로 듬성듬성 비게 된다(실제 데이터로 확인된 버그). 계정명으로 병합하고,
    그 계정명에 해당하는 모든 account_id를 "별칭(alias)"으로 함께 들고 있다가
    값을 채울 때 순서대로 시도한다.

    반환: [{account_id(대표 id), account_nm, alias_ids:[그 이름으로 나온 모든 id]}]
    ord 오름차순, 최근 기간 기준 우선."""
    seen: dict[str, dict] = {}
    order_hint: dict[str, int] = {}
    for p in reversed(periods):  # 최근 기간을 우선 기준으로
        data = period_key_fn(p)
        if not data:
            continue
        for item in data.get("items", {}).get(sj_div, []):
            aid = item["account_id"]
            nm = item["account_nm"]
            ordv = int(item.get("ord") or 9999)
            if nm not in seen:
                seen[nm] = {"id": aid, "alias_ids": {aid}}
                order_hint[nm] = ordv
            else:
                seen[nm]["alias_ids"].add(aid)
    ordered = sorted(seen.keys(), key=lambda nm: order_hint[nm])
    return [
        {"account_id": seen[nm]["id"], "account_nm": nm, "alias_ids": sorted(seen[nm]["alias_ids"])}
        for nm in ordered
    ]


def _cell_ref_with_alias(cell_index: dict, sj_div: str, acc: dict, period_label: str, field: str = "thstrm_amount"):
    """acc(collect_accounts가 반환한 항목)의 대표 account_id로 먼저 찾고,
    없으면 같은 계정명으로 묶인 다른 alias_ids도 순서대로 시도한다."""
    ids = [acc["account_id"]] + [a for a in acc.get("alias_ids", []) if a != acc["account_id"]]
    for aid in ids:
        ref = cell_index.get((sj_div, aid, period_label, field))
        if ref:
            return ref
    return None


def amount_lookup(data: dict, sj_div: str, account_id: str, field: str = "thstrm_amount"):
    if not data:
        return None
    for item in data.get("items", {}).get(sj_div, []):
        if item["account_id"] == account_id:
            val = item.get(field)
            if val in (None, ""):
                return None
            try:
                return float(str(val).replace(",", ""))
            except ValueError:
                return None
    return None


UNIT_DIVISOR = 100_000_000  # 원 -> 억원


def write_raw_sheet(
    wb: Workbook, quarter_plan: list[dict], year_list: list[str], reports: dict,
    sheet_name: str = "원본데이터", extra_periods: list[tuple[dict, str]] | None = None,
):
    """모든 원자료를 원본 시트에 적재하고, 셀 좌표 인덱스를 반환한다.
    금액은 억원 단위(원 값 / 100,000,000)로 저장한다 — 이 시트가 모든 하위
    시트·수식의 유일한 소스이므로, 여기서 한 번만 나누면 분기/연간 재무제표,
    지표 시트, 투자분석 시트의 금액 셀 전부가 자동으로 억원 단위가 된다."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_state = "hidden"
    ws["A1"] = "이 시트는 DART 원본 응답값(단위: 억원으로 환산)을 담은 참조용 데이터입니다. 직접 수정하지 마세요."
    ws["A1"].font = Font(name=FONT_NAME, italic=True, size=9)

    row = 3
    cell_index = {}  # (sj_div, account_id, period_label, field) -> "'{sheet_name}'!$X$Y"

    def dump_period(period_data: dict, label: str):
        nonlocal row
        if not period_data:
            return
        ws.cell(row=row, column=1, value=label).font = BOLD
        row += 1
        for sj_div, sj_name in SJ_ORDER:
            items = period_data.get("items", {}).get(sj_div, [])
            if not items:
                continue
            ws.cell(row=row, column=1, value=sj_name).font = Font(name=FONT_NAME, italic=True)
            row += 1
            for item in items:
                ws.cell(row=row, column=1, value=item["account_id"])
                ws.cell(row=row, column=2, value=item["account_nm"])
                amt = item.get("thstrm_amount")
                ws.cell(row=row, column=3, value=float(str(amt).replace(",", "")) / UNIT_DIVISOR if amt not in (None, "") else None).font = BLUE
                cell_index[(sj_div, item["account_id"], label, "thstrm_amount")] = f"'{sheet_name}'!${get_column_letter(3)}${row}"
                add_amt = item.get("thstrm_add_amount")
                if add_amt not in (None, ""):
                    ws.cell(row=row, column=4, value=float(str(add_amt).replace(",", "")) / UNIT_DIVISOR).font = BLUE
                    cell_index[(sj_div, item["account_id"], label, "thstrm_add_amount")] = f"'{sheet_name}'!${get_column_letter(4)}${row}"
                row += 1
            row += 1
        row += 1

    # 분기용 원자료
    for p in quarter_plan:
        for key, rep_field in (("q1", "11013"), ("h1", "11012"), ("q3", "11014"), ("fy", "11011")):
            if key in p:
                dump_period(p[key], f"{p['year']}_{key}")

    # 연간용 원자료 (사업보고서, 분기용과 중복될 수 있으나 라벨을 분리해 독립적으로 유지)
    for year in year_list:
        fy_data = reports.get(year, {}).get("11011")
        dump_period(fy_data, f"{year}_사업보고서")

    # 추정 연도용 추가 원자료(당해 최신 누적실적, 전년 동기 누적실적 등)
    for data, label in (extra_periods or []):
        dump_period(data, label)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    return cell_index


def style_header(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def apply_grid_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """지정한 범위의 모든 셀에 얇은 테두리를 그린다(표 구분선)."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def build_quarterly_sheet(wb: Workbook, quarter_plan: list[dict], cell_index: dict, sheet_name: str = "분기_재무제표", hidden: bool = False):
    ws = wb.create_sheet(sheet_name)
    if hidden:
        ws.sheet_state = "hidden"
    ws["A1"] = "단위: 억원 | 음영 셀은 원본데이터 시트 링크 또는 수식으로 자동 계산됩니다."
    ws["A1"].font = Font(name=FONT_NAME, italic=True, size=9)

    header_row = 3
    ws.cell(row=header_row, column=1, value="구분")
    ws.cell(row=header_row, column=2, value="계정과목")
    for i, p in enumerate(quarter_plan):
        ws.cell(row=header_row, column=3 + i, value=f"{p['year']}Q{p['q']}")
    style_header(ws, header_row, 2 + len(quarter_plan))

    account_row_map: dict[tuple[str, str], int] = {}
    account_name_map: dict[tuple[str, str], str] = {}

    row = header_row + 1
    for sj_div, sj_name in SJ_ORDER:
        accounts = collect_accounts(
            quarter_plan, sj_div,
            lambda p: p.get("fy") or p.get("q3") or p.get("h1") or p.get("q1"),
        )
        if not accounts:
            continue
        ws.cell(row=row, column=1, value=sj_name).font = BOLD
        row += 1
        for acc in accounts:
            account_row_map[(sj_div, acc["account_id"])] = row
            account_name_map[(sj_div, acc["account_id"])] = acc["account_nm"]
            ws.cell(row=row, column=2, value=acc["account_nm"])
            for i, p in enumerate(quarter_plan):
                col = 3 + i
                cell = ws.cell(row=row, column=col)
                q = p["q"]
                label = f"{p['year']}_"
                if sj_div == "BS":
                    # 시점 데이터: 해당 분기말 보고서를 그대로 링크
                    src_key = "fy" if q == 4 else "q3" if q == 3 else "h1" if q == 2 else "q1"
                    ref = _cell_ref_with_alias(cell_index, sj_div, acc, f"{label}{src_key}", "thstrm_amount")
                    if ref:
                        cell.value = f"={ref}"
                        cell.font = GREEN
                elif q == 1:
                    ref = _cell_ref_with_alias(cell_index, sj_div, acc, f"{label}q1", "thstrm_amount")
                    if ref:
                        cell.value = f"={ref}"
                        cell.font = GREEN
                elif q == 3:
                    ref = _cell_ref_with_alias(cell_index, sj_div, acc, f"{label}q3", "thstrm_amount")
                    if ref:
                        cell.value = f"={ref}"
                        cell.font = GREEN
                elif q == 2:
                    h1_ref = _cell_ref_with_alias(cell_index, sj_div, acc, f"{label}h1", "thstrm_amount")
                    q1_ref = _cell_ref_with_alias(cell_index, sj_div, acc, f"{label}q1", "thstrm_amount")
                    if h1_ref and q1_ref:
                        cell.value = f"={h1_ref}-{q1_ref}"
                        cell.font = BLACK
                elif q == 4:
                    fy_ref = _cell_ref_with_alias(cell_index, sj_div, acc, f"{label}fy", "thstrm_amount")
                    # 4분기는 원칙적으로 3분기보고서의 '누적' 필드(thstrm_add_amount)를 쓴다.
                    q3_add_ref = _cell_ref_with_alias(cell_index, sj_div, acc, f"{label}q3", "thstrm_add_amount")
                    if fy_ref and q3_add_ref:
                        cell.value = f"={fy_ref}-{q3_add_ref}"
                        cell.font = BLACK
                    elif fy_ref:
                        # 현금흐름표 항목처럼 3분기보고서에 누적 필드가 없는 경우가 있다
                        # (분기보고서 자체가 이미 연초 누적치로 공시되는 계정 등).
                        # 이 경우 같은 행에 이미 계산해 둔 1~3분기 값을 빼는 방식으로 대체한다:
                        # 4분기 = 연간 - 1분기 - 2분기 - 3분기.
                        q1_col = get_column_letter(col - 3)
                        q2_col = get_column_letter(col - 2)
                        q3_col = get_column_letter(col - 1)
                        cell.value = f"={fy_ref}-{q1_col}{row}-{q2_col}{row}-{q3_col}{row}"
                        cell.font = BLACK
                cell.number_format = "#,##0.0;(#,##0.0);-"
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 32
    for i in range(len(quarter_plan)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 15
    ws.freeze_panes = "C4"
    apply_grid_border(ws, header_row, row - 1, 1, 2 + len(quarter_plan))

    period_labels = [f"{p['year']}Q{p['q']}" for p in quarter_plan]
    return account_row_map, account_name_map, period_labels


def _latest_reprt_code(reports: dict, year: str) -> str | None:
    """해당 연도에 공시된 보고서 중 가장 최신(사업>3분기>반기>1분기) 코드를 반환한다."""
    year_reports = reports.get(year, {})
    for code in ("11011", "11014", "11012", "11013"):
        if code in year_reports:
            return code
    return None


def build_annual_sheet(
    wb: Workbook, year_list: list[str], reports: dict, cell_index: dict,
    sheet_name: str = "연간_재무제표", hidden: bool = False,
    estimated_year: dict | None = None,
):
    """estimated_year가 주어지면 마지막에 "{year}(E)" 컬럼을 하나 더 추가한다.
    estimated_year = {
        "year": "2026", "prior_year": "2025",
        "cur_label": "2026_추정기준",  # 원본데이터에서 당해 누적실적을 찾을 라벨
        "prior_label": "2025_추정기준",  # 원본데이터에서 전년 동기 누적실적을 찾을 라벨
        "field": "thstrm_amount" | "thstrm_add_amount",  # flow 계정 비율 계산에 쓸 필드
        "bs_label": "2026_추정기준",  # BS 잔액은 이 라벨의 thstrm_amount를 그대로 사용
    }
    계산 규칙(사용자 확정): flow 계정 = 전년 사업보고서 실적 × (당해 누적실적 ÷ 전년 동기 누적실적).
    BS(잔액) 계정 = 당해 최신 분기말 잔액을 그대로(비율 계산 대상 아님)."""
    ws = wb.create_sheet(sheet_name)
    if hidden:
        ws.sheet_state = "hidden"
    ws["A1"] = "단위: 억원 | 사업보고서(연결/별도) 기준, 원본데이터 시트 링크"
    ws["A1"].font = Font(name=FONT_NAME, italic=True, size=9)
    if estimated_year:
        ws["A2"] = (
            f"※ {estimated_year['year']}(E)는 사업보고서가 아직 없어 추정한 값입니다: "
            f"{estimated_year['prior_year']}년 실적 × ({estimated_year['year']}년 최신 누적실적 ÷ "
            f"{estimated_year['prior_year']}년 동기간 누적실적). 재무상태표 잔액은 최신 분기말 값을 그대로 썼습니다."
        )
        ws["A2"].font = Font(name=FONT_NAME, italic=True, size=9, color="C00000")

    all_labels = [f"{y}" for y in year_list] + ([f"{estimated_year['year']}(E)"] if estimated_year else [])
    n_cols = len(year_list) + (1 if estimated_year else 0)

    header_row = 3
    ws.cell(row=header_row, column=1, value="구분")
    ws.cell(row=header_row, column=2, value="계정과목")
    for i, lab in enumerate(all_labels):
        ws.cell(row=header_row, column=3 + i, value=lab)
    style_header(ws, header_row, 2 + n_cols)

    account_row_map: dict[tuple[str, str], int] = {}
    account_name_map: dict[tuple[str, str], str] = {}

    row = header_row + 1
    fy_periods = [{"fy": reports.get(y, {}).get("11011")} for y in year_list]
    if estimated_year:
        fy_periods.append({"fy": reports.get(estimated_year["year"], {}).get(_latest_reprt_code(reports, estimated_year["year"]))})
    for sj_div, sj_name in SJ_ORDER:
        accounts = collect_accounts(fy_periods, sj_div, lambda p: p.get("fy"))
        if not accounts:
            continue
        ws.cell(row=row, column=1, value=sj_name).font = BOLD
        row += 1
        is_bs = (sj_div == "BS")
        for acc in accounts:
            account_row_map[(sj_div, acc["account_id"])] = row
            account_name_map[(sj_div, acc["account_id"])] = acc["account_nm"]
            ws.cell(row=row, column=2, value=acc["account_nm"])
            for i, year in enumerate(year_list):
                col = 3 + i
                cell = ws.cell(row=row, column=col)
                ref = _cell_ref_with_alias(cell_index, sj_div, acc, f"{year}_사업보고서", "thstrm_amount")
                if ref:
                    cell.value = f"={ref}"
                    cell.font = GREEN
                cell.number_format = "#,##0.0;(#,##0.0);-"
            if estimated_year:
                col = 3 + len(year_list)
                cell = ws.cell(row=row, column=col)
                if is_bs:
                    ref = _cell_ref_with_alias(cell_index, sj_div, acc, estimated_year["bs_label"], "thstrm_amount")
                    if ref:
                        cell.value = f"={ref}"
                        cell.font = BLUE
                else:
                    prior_fy_ref = _cell_ref_with_alias(cell_index, sj_div, acc, f"{estimated_year['prior_year']}_사업보고서", "thstrm_amount")
                    cur_ref = _cell_ref_with_alias(cell_index, sj_div, acc, estimated_year["cur_label"], estimated_year["field"])
                    prior_ref = _cell_ref_with_alias(cell_index, sj_div, acc, estimated_year["prior_label"], estimated_year["field"])
                    if prior_fy_ref and cur_ref and prior_ref:
                        cell.value = f"=IFERROR({prior_fy_ref}*({cur_ref}/{prior_ref}),\"\")"
                        cell.font = BLUE
                cell.number_format = "#,##0.0;(#,##0.0);-"
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 32
    for i in range(n_cols):
        ws.column_dimensions[get_column_letter(3 + i)].width = 18
    ws.freeze_panes = "C4"
    apply_grid_border(ws, header_row, row - 1, 1, 2 + n_cols)

    return account_row_map, account_name_map, all_labels


# ---------------------------------------------------------------------------
# 지표(추세) 시트 & 차트 시트
#
# DART 계정명(account_nm)은 회사마다 표기가 조금씩 다를 수 있어(예: "매출액" vs
# "수익(매출액)", "기타수익" vs "기타영업외수익") 계정ID 하나로 고정 매칭하지 않고
# 우선순위가 있는 후보 이름 목록 + 부분일치 폴백으로 탐색한다. 못 찾으면 해당
# 지표는 빈 칸으로 두고 경고를 출력한다(값을 임의로 지어내지 않음).
# ---------------------------------------------------------------------------

# key: 내부 식별자, value: (sj_div, [완전일치 후보(우선순위순)], [부분일치 폴백 키워드], [제외 키워드])
METRIC_RULES: dict[str, tuple[str, list[str], list[str], list[str]]] = {
    "매출액": ("IS", ["매출액", "수익(매출액)", "영업수익"], ["매출액"], ["매출원가", "율"]),
    "매출원가": ("IS", ["매출원가"], ["매출원가"], ["율"]),
    "영업이익": ("IS", ["영업이익", "영업이익(손실)"], ["영업이익"], ["율", "률"]),
    "당기순이익": (
        "IS",
        ["당기순이익", "당기순이익(손실)", "분기순이익(손실)", "반기순이익(손실)", "분기순이익", "반기순이익"],
        ["순이익"],
        ["지배", "비지배", "주당"],
    ),
    "유동자산": ("BS", ["유동자산"], ["유동자산"], ["비유동"]),
    "유동부채": ("BS", ["유동부채"], ["유동부채"], ["비유동"]),
    "자산총계": ("BS", ["자산총계"], ["자산총계"], []),
    "부채총계": ("BS", ["부채총계"], ["부채총계"], []),
    "자본총계": ("BS", ["자본총계"], ["자본총계"], []),
    "매출채권및기타채권": (
        "BS",
        ["매출채권및기타채권", "매출채권및기타유동채권", "매출채권"],
        ["매출채권"],
        ["비유동"],
    ),
    "이익잉여금": ("BS", ["이익잉여금", "이익잉여금(결손금)"], ["이익잉여금"], []),
    "현금및현금성자산의증가": (
        "CF",
        [
            "현금및현금성자산의순증가(감소)", "현금및현금성자산의 증가(감소)", "현금및현금성자산의증가(감소)",
            "현금및현금성자산의 증가", "현금및현금성자산의증가", "현금및현금성자산의 증감", "현금및현금성자산의증감",
            "현금및현금성자산의 순증가", "현금및현금성자산의순증가",
        ],
        ["현금및현금성자산의"],
        ["기초", "기말", "환율"],
    ),
    "금융수익": ("IS", ["금융수익"], ["금융수익"], []),
    "금융비용": ("IS", ["금융비용"], ["금융비용"], []),
    "기타수익": ("IS", ["기타수익", "기타영업외수익"], ["기타수익", "기타영업외수익"], []),
    "기타비용": ("IS", ["기타비용", "기타영업외비용"], ["기타비용", "기타영업외비용"], []),

    # --- 아래는 "투자분석" 시트 전용 추가 지표 (기존 단일기업 지표/차트에는 영향 없음) ---
    "재고자산": ("BS", ["재고자산"], ["재고자산"], []),
    "비유동자산": ("BS", ["비유동자산"], ["비유동자산"], []),
    "법인세차감전순이익": (
        "IS",
        ["법인세비용차감전순이익(손실)", "법인세비용차감전순이익", "법인세차감전순이익"],
        ["법인세비용차감전", "법인세차감전"],
        [],
    ),
    "영업활동현금흐름": (
        "CF",
        ["영업활동으로인한현금흐름", "영업활동현금흐름"],
        ["영업활동"],
        ["투자활동", "재무활동"],
    ),
    "현금및현금성자산_기말": ("BS", ["현금및현금성자산"], ["현금및현금성자산"], []),
    "유가증권": (
        "BS",
        ["단기금융상품", "단기투자자산", "유가증권"],
        ["단기금융상품", "단기투자자산", "유가증권"],
        ["장기"],
    ),
    "투자자산": ("BS", ["투자자산", "장기투자자산", "기타비유동금융자산"], ["투자자산"], []),
    "유형자산": ("BS", ["유형자산"], ["유형자산"], []),
    "무형자산": ("BS", ["무형자산"], ["무형자산"], []),
    "기타유동자산": ("BS", ["기타유동자산"], ["기타유동자산"], []),

    # --- v0.5.0: CCC·FCF·이자보상배율·DuPont·ROIC·외환손익용 추가 항목 ---
    "매입채무": ("BS", ["매입채무", "매입채무및기타채무", "매입채무및기타유동채무"], ["매입채무"], ["비유동"]),
    "이자비용": ("IS", ["이자비용", "금융비용"], ["이자비용"], []),
    "법인세비용": ("IS", ["법인세비용", "법인세비용(수익)"], ["법인세비용"], []),
    "유형자산의취득": (
        "CF",
        ["유형자산의취득", "유형자산의 취득"],
        ["유형자산의취득", "유형자산취득"],
        [],
    ),
    "감가상각비": (
        "CF",
        ["감가상각비"],
        ["감가상각비"],
        ["무형자산상각비"],
    ),
    "무형자산상각비": ("CF", ["무형자산상각비"], ["무형자산상각비"], []),
    "투자활동현금흐름": (
        "CF",
        ["투자활동으로인한현금흐름", "투자활동현금흐름"],
        ["투자활동"],
        [],
    ),
    "재무활동현금흐름": (
        "CF",
        ["재무활동으로인한현금흐름", "재무활동현금흐름"],
        ["재무활동"],
        [],
    ),
    "외화환산손익": (
        "IS",
        ["외화환산이익", "외화환산손익"],
        ["외화환산"],
        [],
    ),
    "파생상품손익": (
        "IS",
        ["파생상품평가이익", "파생상품거래이익", "파생상품손익"],
        ["파생상품"],
        [],
    ),
}


def resolve_metric(
    key: str, account_row_map: dict, account_name_map: dict
) -> tuple[str, str] | None:
    """METRIC_RULES에 따라 (sj_div, account_id)를 찾아 반환한다. 못 찾으면 None."""
    sj_div, exact_candidates, substr_keywords, excludes = METRIC_RULES[key]
    # 손익 계정(sj_div="IS")은 회사에 따라 별도 손익계산서 없이 포괄손익계산서(CIS)
    # 하나로만 공시하는 경우가 있다(예: HD한국조선해양). 이 경우 매출액·영업이익·
    # 당기순이익 등도 sj_div="CIS"로 잡히므로, IS를 찾을 때는 CIS도 함께 후보에 넣는다.
    # (BS/CF는 이런 혼선이 없어 그대로 둔다. IS/CIS가 둘 다 있는 회사는 SJ_ORDER 순서상
    # IS 계정이 먼저 삽입되어 있어 그대로 IS가 우선 매칭된다.)
    candidate_sj = ("IS", "CIS") if sj_div == "IS" else (sj_div,)
    same_sj = [
        (sj, aid) for (sj, aid) in account_row_map if sj in candidate_sj
    ]

    def excluded(name: str) -> bool:
        return any(ex in name for ex in excludes)

    # 1차: 완전일치 (우선순위 순서대로)
    for cand in exact_candidates:
        for sj, aid in same_sj:
            name = account_name_map[(sj, aid)]
            if name == cand and not excluded(name):
                return sj, aid

    # 2차: 부분일치
    for kw in substr_keywords:
        for sj, aid in same_sj:
            name = account_name_map[(sj, aid)]
            if kw in name and not excluded(name):
                return sj, aid

    return None


# 지표 시트에 표시할 행 순서와 표시 이름 (라인차트 5개가 참조하는 "기본" 지표들)
INDICATOR_ROWS = [
    "매출액",
    "매출원가",
    "매출총이익",
    "영업이익",
    "당기순이익",
    "경상이익",
    "유동자산",
    "매출채권및기타채권",
    "유동부채",
    "자산총계",
    "부채총계",
    "자본총계",
    "이익잉여금",
    "현금및현금성자산의증가",
]

RATIO_ROWS = ["자기자본비율", "부채비율", "매출총이익률", "원가율", "영업이익률", "순이익률"]

# 그래프1~5 구성 (지표 시트의 행 이름 기준)
LINE_CHART_GROUPS = [
    ("그래프1_매출액-매출원가-매출총이익", ["매출액", "매출원가", "매출총이익"]),
    ("그래프2_이익지표(매출총이익-영업이익-순이익-경상이익)", ["매출총이익", "영업이익", "당기순이익", "경상이익"]),
    ("그래프3_유동자산-유동부채-자산-부채", ["유동자산", "유동부채", "자산총계", "부채총계"]),
    ("그래프4_매출액-매출채권", ["매출액", "매출채권및기타채권"]),
    ("그래프5_이익잉여금-현금증가", ["이익잉여금", "현금및현금성자산의증가"]),
]


def build_indicator_sheet(
    wb: Workbook,
    prefix: str,
    source_sheet_name: str,
    period_labels: list[str],
    account_row_map: dict,
    account_name_map: dict,
    sheet_name: str | None = None,
    hidden: bool = False,
) -> tuple[str, dict[str, int], list[str]]:
    """소스 시트(분기_재무제표/연간_재무제표)를 참조하는 지표 시트를 만든다.
    반환: (시트이름, {행이름: 행번호}, 못찾은 지표 목록)"""
    if sheet_name is None:
        sheet_name = f"지표_{prefix}"
    ws = wb.create_sheet(sheet_name)
    if hidden:
        ws.sheet_state = "hidden"
    n = len(period_labels)

    ws.cell(row=1, column=1, value="지표").font = BOLD
    for i, label in enumerate(period_labels):
        ws.cell(row=1, column=3 + i, value=label)
    style_header(ws, 1, 2 + n)

    # 이 시트가 실제로 쓰는 항목만 확인한다 (투자분석 전용으로 추가된 항목은
    # 여기서 아예 쓰이지 않으므로 missing 목록에 잘못 섞이면 안 된다).
    relevant_keys = (set(INDICATOR_ROWS) - {"매출총이익", "경상이익"}) | {"금융수익", "금융비용", "기타수익", "기타비용"}
    relevant_keys &= set(METRIC_RULES)
    resolved: dict[str, tuple[str, str] | None] = {
        key: resolve_metric(key, account_row_map, account_name_map)
        for key in relevant_keys
    }
    missing = [key for key, v in resolved.items() if v is None]

    row_of: dict[str, int] = {}
    row = 2
    for name in INDICATOR_ROWS:
        ws.cell(row=row, column=2, value=name)
        row_of[name] = row

        if name == "매출총이익":
            if row_of.get("매출액") and row_of.get("매출원가"):
                a, b = row_of["매출액"], row_of["매출원가"]
                for i in range(n):
                    col = get_column_letter(3 + i)
                    ws.cell(row=row, column=3 + i, value=f"={col}{a}-{col}{b}")
        elif name == "경상이익":
            if row_of.get("영업이익"):
                op_row = row_of["영업이익"]
                fin_terms = []  # (부호, sj_div, account_id)
                for key, sign in (("금융수익", "+"), ("금융비용", "-"), ("기타수익", "+"), ("기타비용", "-")):
                    hit = resolved.get(key)
                    if hit:
                        fin_terms.append((sign, hit))
                for i in range(n):
                    col = get_column_letter(3 + i)
                    src_col = get_column_letter(3 + i)
                    parts = [f"{src_col}{op_row}"]
                    for sign, (sj, aid) in fin_terms:
                        r = account_row_map[(sj, aid)]
                        parts.append(f"{sign}'{source_sheet_name}'!{src_col}{r}")
                    ws.cell(row=row, column=3 + i, value="=" + "".join(parts))
        else:
            hit = resolved.get(name)
            if hit:
                sj, aid = hit
                src_row = account_row_map[(sj, aid)]
                for i in range(n):
                    col = get_column_letter(3 + i)
                    ref = f"'{source_sheet_name}'!{col}{src_row}"
                    # 해당 회사/기간에 실제 데이터가 없어 원본 셀이 비어 있으면 0이 아니라
                    # NA()를 반환한다 → 차트에서 0으로 떨어지지 않고 구간이 끊겨(gap) 표시됨.
                    ws.cell(row=row, column=3 + i, value=f'=IF({ref}="",NA(),{ref})')

        for i in range(n):
            ws.cell(row=row, column=3 + i).number_format = "#,##0.0;(#,##0.0);-"
        row += 1

    row += 1  # 구분 여백
    for name in RATIO_ROWS:
        ws.cell(row=row, column=2, value=name)
        row_of[name] = row
        if name == "자기자본비율" and row_of.get("자본총계") and row_of.get("자산총계"):
            num, den = row_of["자본총계"], row_of["자산총계"]
        elif name == "부채비율" and row_of.get("부채총계") and row_of.get("자본총계"):
            num, den = row_of["부채총계"], row_of["자본총계"]
        elif name == "매출총이익률" and row_of.get("매출총이익") and row_of.get("매출액"):
            num, den = row_of["매출총이익"], row_of["매출액"]
        elif name == "원가율" and row_of.get("매출원가") and row_of.get("매출액"):
            num, den = row_of["매출원가"], row_of["매출액"]
        elif name == "영업이익률" and row_of.get("영업이익") and row_of.get("매출액"):
            num, den = row_of["영업이익"], row_of["매출액"]
        elif name == "순이익률" and row_of.get("당기순이익") and row_of.get("매출액"):
            num, den = row_of["당기순이익"], row_of["매출액"]
        else:
            num, den = None, None
        if num and den:
            for i in range(n):
                col = get_column_letter(3 + i)
                ws.cell(row=row, column=3 + i, value=f"=IFERROR({col}{num}/{col}{den}*100,NA())")
                ws.cell(row=row, column=3 + i).number_format = "0.0"
        row += 1

    if missing:
        ws.cell(row=row + 1, column=1, value="※ 아래 지표는 이 회사 공시에서 계정명을 찾지 못해 비어 있습니다:").font = Font(
            name=FONT_NAME, italic=True, size=9, color="C00000"
        )
        ws.cell(row=row + 2, column=1, value=", ".join(missing)).font = Font(
            name=FONT_NAME, italic=True, size=9, color="C00000"
        )

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 15
    ws.freeze_panes = "C2"
    apply_grid_border(ws, 1, row - 1, 2, 2 + n)

    return sheet_name, row_of, missing


def _add_line_series(chart: LineChart, ws, row_of: dict, names: list[str], n_periods: int):
    for name in names:
        r = row_of.get(name)
        if not r:
            continue
        data_ref = Reference(ws, min_col=3, max_col=2 + n_periods, min_row=r, max_row=r)
        series = Series(data_ref, title=name)
        series.smooth = False
        chart.series.append(series)


def _set_chart_title_font_size(chart, size_pt: int = 12) -> None:
    """차트 제목 폰트 크기를 지정한다(pt 단위)."""
    cp = CharacterProperties(sz=size_pt * 100, b=True)
    chart.title.tx.rich.p[0].pPr = ParagraphProperties(defRPr=cp)


def build_chart_sheet(
    wb: Workbook, prefix: str, indicator_sheet_name: str, row_of: dict, n_periods: int,
    embed_anchor_col: str | None = None,
):
    """차트를 그린다. embed_anchor_col이 주어지면 별도 시트를 만들지 않고
    지표 시트(indicator_sheet_name) 자체의 그 열부터 차트를 배치한다
    (표와 겹치지 않도록 호출하는 쪽에서 표 폭보다 오른쪽 열을 넘겨야 한다).
    embed_anchor_col이 있는 경우(=지표_분기/지표_연간에 임베드하는 경우)는 v0.9.5부터
    차트 제목 폰트 크기를 12pt로 맞춘다(사용자 요청)."""
    ind_ws = wb[indicator_sheet_name]
    if embed_anchor_col:
        ws = ind_ws
        anchor_col = embed_anchor_col
    else:
        ws = wb.create_sheet(f"차트_{prefix}")
        anchor_col = "A"
    cat_ref = Reference(ind_ws, min_col=3, max_col=2 + n_periods, min_row=1, max_row=1)

    anchor_row = 1
    for title, names in LINE_CHART_GROUPS:
        chart = LineChart()
        chart.title = title
        if embed_anchor_col:
            _set_chart_title_font_size(chart, 12)
        chart.style = 2
        chart.y_axis.title = "금액(억원)"
        chart.x_axis.title = "기간"
        chart.height = 9
        chart.width = 22
        _add_line_series(chart, ind_ws, row_of, names, n_periods)
        chart.set_categories(cat_ref)
        ws.add_chart(chart, f"{anchor_col}{anchor_row}")
        anchor_row += 19

    # 비율 막대그래프 (계열 6개)
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "그래프6_수익성-안정성 비율(%)"
    if embed_anchor_col:
        _set_chart_title_font_size(bar, 12)
    bar.style = 10
    bar.y_axis.title = "%"
    bar.x_axis.title = "기간"
    bar.height = 9
    bar.width = 22
    for name in RATIO_ROWS:
        r = row_of.get(name)
        if not r:
            continue
        data_ref = Reference(ind_ws, min_col=3, max_col=2 + n_periods, min_row=r, max_row=r)
        series = Series(data_ref, title=name)
        bar.series.append(series)
    bar.set_categories(cat_ref)
    ws.add_chart(bar, f"{anchor_col}{anchor_row}")

    if not embed_anchor_col:
        ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# 투자 판단 자동 평가 (v0.8.0)
# 첨부 참고자료("투자판단 항목 평가기준")의 수치 기준을 그대로 코드화한다.
# 등급 규칙: 최근 5개년 중 기준 충족 연수로 A~E를 매기되, 기준에 계속 미달해도
# 5년 내내 개선 추세면 A로 승격한다(사용자 지정 규칙).
# ---------------------------------------------------------------------------

def grade_by_hit_years(hits: list[bool | None]) -> tuple[str, int, int]:
    """hits: 연도별 기준 충족 여부(None=데이터 없음).
    반환: (등급, 충족연수, 판정가능연수)"""
    valid = [h for h in hits if h is not None]
    if not valid:
        return "-", 0, 0
    k = sum(1 for h in valid if h)
    total = len(valid)
    # 5개년이 아닌 경우(상장 이력 부족 등)에도 같은 비율로 환산해 판정한다.
    scaled = round(k / total * 5) if total else 0
    grade = {5: "A", 4: "B", 3: "C", 2: "D"}.get(scaled, "E")
    return grade, k, total


def is_consistently_improving(values: list[float | None], higher_is_better: bool = True) -> bool:
    """연도별 값이 계속 개선되는 추세인지(모든 구간에서 단조 개선) 판정한다.
    데이터가 3개 미만이면 판단하지 않는다(False)."""
    vals = [v for v in values if v is not None]
    if len(vals) < 3:
        return False
    for prev, cur in zip(vals, vals[1:]):
        if higher_is_better and cur <= prev:
            return False
        if not higher_is_better and cur >= prev:
            return False
    return True


def evaluate_metric(
    values: list[float | None],
    threshold: float,
    higher_is_better: bool = True,
) -> tuple[str, int, int, bool]:
    """단일 지표를 등급화한다. 반환: (등급, 충족연수, 판정가능연수, 개선승격여부)"""
    hits = [
        None if v is None else (v >= threshold if higher_is_better else v <= threshold)
        for v in values
    ]
    grade, k, total = grade_by_hit_years(hits)
    improved = False
    if grade not in ("A", "-") and is_consistently_improving(values, higher_is_better):
        grade = "A"
        improved = True
    return grade, k, total, improved


def combine_grades(grades: list[str]) -> str:
    """여러 하위 지표 등급의 평균으로 항목 등급을 낸다(A=5 ... E=1)."""
    score_map = {"A": 5, "B": 4, "C": 3, "D": 2, "E": 1}
    scores = [score_map[g] for g in grades if g in score_map]
    if not scores:
        return "-"
    avg = sum(scores) / len(scores)
    if avg >= 4.5:
        return "A"
    if avg >= 3.5:
        return "B"
    if avg >= 2.5:
        return "C"
    if avg >= 1.5:
        return "D"
    return "E"


def render_investment_judgement(
    ws, row: int, n: int, y_period_labels: list[str],
    raw: dict[str, list[float | None]],
    extra_disclosure: dict | None,
    price_available: bool,
) -> int:
    """N. 투자 판단 표를 자동 평가로 채운다. 반환: 다음 행 번호.
    raw: 지표명 -> 연도별 값 리스트(오래된->최근). 값이 없으면 None."""

    def fmt_years(k: int, total: int) -> str:
        return f"{total}년 중 {k}년 충족"

    results: list[tuple[str, str, str]] = []  # (항목, 등급, 근거)

    # --- 재무건전성: 자기자본비율 40%↑, 부채비율 200%↓, 유동비율 100%↑ ---
    g1, k1, t1, i1 = evaluate_metric(raw.get("자기자본비율", []), 40, True)
    g2, k2, t2, i2 = evaluate_metric(raw.get("부채비율", []), 200, False)
    g3, k3, t3, i3 = evaluate_metric(raw.get("유동비율", []), 100, True)
    parts = [
        f"자기자본비율 40%↑ {fmt_years(k1, t1)}({g1}{', 지속개선' if i1 else ''})",
        f"부채비율 200%↓ {fmt_years(k2, t2)}({g2}{', 지속개선' if i2 else ''})",
        f"유동비율 100%↑ {fmt_years(k3, t3)}({g3}{', 지속개선' if i3 else ''})",
    ]
    results.append(("재무건전성", combine_grades([g1, g2, g3]), " / ".join(parts)))

    # --- 수익성: 영업이익률 5%↑, ROE 10%↑, ROA 5%↑ ---
    g4, k4, t4, i4 = evaluate_metric(raw.get("영업이익률", []), 5, True)
    g5, k5, t5, i5 = evaluate_metric(raw.get("ROE", []), 10, True)
    g6, k6, t6, i6 = evaluate_metric(raw.get("ROA", []), 5, True)
    parts = [
        f"영업이익률 5%↑ {fmt_years(k4, t4)}({g4}{', 지속개선' if i4 else ''})",
        f"ROE 10%↑ {fmt_years(k5, t5)}({g5}{', 지속개선' if i5 else ''})",
        f"ROA 5%↑ {fmt_years(k6, t6)}({g6}{', 지속개선' if i6 else ''})",
    ]
    results.append(("수익성", combine_grades([g4, g5, g6]), " / ".join(parts)))

    # --- 성장성: 매출성장률 10%↑, 영업이익성장률 0%↑(플러스), 총자산회전율 1.0↑ ---
    g7, k7, t7, i7 = evaluate_metric(raw.get("매출성장률", []), 10, True)
    g8, k8, t8, i8 = evaluate_metric(raw.get("영업이익성장률", []), 0, True)
    g9, k9, t9, i9 = evaluate_metric(raw.get("총자산회전율", []), 1.0, True)
    parts = [
        f"매출성장률 10%↑ {fmt_years(k7, t7)}({g7}{', 지속개선' if i7 else ''})",
        f"영업이익성장률 + {fmt_years(k8, t8)}({g8}{', 지속개선' if i8 else ''})",
        f"총자산회전율 1.0↑ {fmt_years(k9, t9)}({g9}{', 지속개선' if i9 else ''})",
    ]
    results.append(("성장성", combine_grades([g7, g8, g9]), " / ".join(parts)))

    # --- 자산으로 본 저평가 정도: PBR 1.0 미만이면 저평가 ---
    if price_available and raw.get("PBR"):
        gA, kA, tA, iA = evaluate_metric(raw["PBR"], 1.0, False)
        results.append((
            "자산으로 본 저평가 정도", gA,
            f"PBR 1.0 미만 {fmt_years(kA, tA)}({gA}{', 지속개선' if iA else ''}). "
            f"청산가치 대비 시가총액은 D·L 섹션 참고."
        ))
    else:
        results.append((
            "자산으로 본 저평가 정도", "-",
            "주가 데이터 없음 — KRX 인증키로 fetch_stock_price.py 실행 후 재생성하면 자동 평가됩니다."
        ))

    # --- 수익 창출 능력으로 본 저평가 정도: PER 15 미만 ---
    if price_available and raw.get("PER"):
        gB, kB, tB, iB = evaluate_metric(raw["PER"], 15.0, False)
        results.append((
            "수익 창출 능력으로 본 저평가 정도", gB,
            f"PER 15배 미만 {fmt_years(kB, tB)}({gB}{', 지속개선' if iB else ''}). "
            f"PSR은 L섹션 참고."
        ))
    else:
        results.append((
            "수익 창출 능력으로 본 저평가 정도", "-",
            "주가 데이터 없음 — KRX 인증키로 fetch_stock_price.py 실행 후 재생성하면 자동 평가됩니다."
        ))

    # --- 사업역량: 정성 판단 영역 (자동 평가하지 않음) ---
    results.append((
        "사업역량", "(직접 입력)",
        "사업 단순성·매입처/판매처 분산·경쟁력은 공시 수치로 판단할 수 없습니다. "
        "사업보고서의 사업의 내용, 매출 구성, 주요 거래처를 직접 확인해 채워 주세요."
    ))

    # --- 주주 중시 자세: 배당성향 + 자사주 + 증자 이력 ---
    payout_txt = "배당성향 정보 없음"
    treasury_txt = ""
    grade_shareholder = "-"
    if extra_disclosure:
        div = extra_disclosure.get("배당", {})
        div_list = div.get("list", []) if isinstance(div, dict) else []
        payout_raw = next(
            (x.get("thstrm") for x in div_list if "배당성향" in (x.get("se") or "")), None
        )
        payout_val = None
        if payout_raw:
            try:
                payout_val = float(str(payout_raw).replace(",", "").replace("%", "").strip())
            except ValueError:
                payout_val = None
        if payout_val is not None:
            payout_txt = f"배당성향 {payout_val:.1f}%"
            # 참고자료에 명시적 컷오프가 없어 통상 기준(20% 이상 주주환원 적극)을 쓴다.
            grade_shareholder = "A" if payout_val >= 30 else "B" if payout_val >= 20 else "C" if payout_val > 0 else "D"
        elif payout_raw:
            payout_txt = f"배당성향 {payout_raw}"

        treasury = extra_disclosure.get("자기주식현황", {})
        t_list = treasury.get("list", []) if isinstance(treasury, dict) else []
        if t_list:
            treasury_txt = " / 자기주식 보유·취득 이력 있음(긍정 신호)"

    results.append((
        "주주 중시 자세", grade_shareholder,
        f"{payout_txt}{treasury_txt}. "
        "※ 참고자료가 경고한 '증자로 자금을 해결하는 기업'인지는 DART 유상증자 공시를 별도로 확인해야 합니다(자동 판별 불가)."
    ))

    # --- 표 렌더링 ---
    for item, grade, memo in results:
        ws.cell(row=row, column=1, value=item)
        c = ws.cell(row=row, column=2, value=grade)
        if grade in ("A", "B"):
            c.font = Font(name=FONT_NAME, bold=True, color="1F7A1F")
        elif grade in ("D", "E"):
            c.font = Font(name=FONT_NAME, bold=True, color="C00000")
        ws.cell(row=row, column=3, value=memo)
        row += 1

    return row


# ---------------------------------------------------------------------------
# 투자분석 시트 (v0.4.0) — 첨부 자료("기업 분석 리포트 쓰는 법")의 재무지표/
# 위험신호/청산가치/주가지표 체크리스트를 자동 계산한다.
# ---------------------------------------------------------------------------

LABEL = Font(name=FONT_NAME, bold=True)
NOTE = Font(name=FONT_NAME, italic=True, size=9, color="808080")
WARN = Font(name=FONT_NAME, color="C00000")


def load_company_profile(corp_code: str) -> dict | None:
    fp = CACHE_DIR / f"company_{corp_code}.json"
    if not fp.exists():
        return None
    try:
        return json.loads(fp.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def load_extra_disclosures(corp_code: str, year: str, reprt_code: str = "11011") -> dict | None:
    fp = CACHE_DIR / f"extra_{corp_code}_{year}_{reprt_code}.json"
    if not fp.exists():
        return None
    try:
        return json.loads(fp.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def load_price(stock_code: str, yyyymmdd: str) -> dict | None:
    fp = CACHE_DIR / f"price_{stock_code}_{yyyymmdd}.json"
    if not fp.exists():
        return None
    try:
        return json.loads(fp.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _fmt_num(v):
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None





# (v0.13.0: 분기 연환산 관련 코드 전부 제거 — 사용자 요청으로 원복)

def compute_estimated_year_series(reports: dict, estimated_year: dict, keys_hits: dict) -> dict:
    """N섹션(투자판단 자동평가)용: 추정 연도의 flow/BS 계정값을 파이썬으로
    계산한다(연간_재무제표의 수식과 동일한 로직, 값 레벨 재현).
    반환: {key: {estimated_year: 값(원 단위, series()가 amount_lookup과 동일 스케일로 기대)}}"""
    result: dict[str, dict[str, float]] = {}
    year = estimated_year["year"]
    prior_year = estimated_year["prior_year"]
    field = estimated_year["field"]
    latest_code = estimated_year["latest_code"]
    cur_data = reports.get(year, {}).get(latest_code)
    prior_data = reports.get(prior_year, {}).get(latest_code)
    prior_fy = reports.get(prior_year, {}).get("11011")

    for key, hit in keys_hits.items():
        if hit is None:
            continue
        sj, aid = hit
        if sj == "BS":
            v = amount_lookup(cur_data, sj, aid, "thstrm_amount")
            if v is not None:
                result[key] = {year: v}
            continue
        cur_v = amount_lookup(cur_data, sj, aid, field)
        prior_v = amount_lookup(prior_data, sj, aid, field)
        prior_fy_v = amount_lookup(prior_fy, sj, aid, "thstrm_amount")
        if cur_v is not None and prior_v not in (None, 0) and prior_fy_v is not None:
            result[key] = {year: prior_fy_v * (cur_v / prior_v)}
    return result


def build_investment_analysis_sheet(
    wb: Workbook,
    company_name: str,
    corp_code: str,
    y_period_labels: list[str],
    year_list: list[str],
    y_account_row_map: dict,
    y_account_name_map: dict,
    y_ind_sheet: str,
    y_row_of: dict,
    reports: dict | None = None,
    sheet_title: str = "투자분석",
    source_flow_sheet: str = "연간_재무제표",
    period_dates: list[str] | None = None,
    annualized_series: dict | None = None,
    banner: str | None = None,
) -> list[str]:
    """'투자분석' 시트를 만든다. 반환값: 이번에 못 찾은/못 가져온 항목 경고 목록."""
    ws = wb.create_sheet(sheet_title)
    n = len(y_period_labels)
    warnings: list[str] = []
    profile = load_company_profile(corp_code)
    stock_code = (profile or {}).get("stock_code", "").strip() or None

    def resolve(key: str):
        hit = resolve_metric(key, y_account_row_map, y_account_name_map)
        if hit is None:
            warnings.append(key)
        return hit

    def resolve_optional(key: str):
        # 회사에 따라 애초에 없는 게 정상인 항목(외환손익 등)은 "매칭 실패" 경고에 안 넣는다.
        return resolve_metric(key, y_account_row_map, y_account_name_map)

    def year_ref(hit, i: int) -> str:
        if not hit:
            return ""
        sj, aid = hit
        src_row = y_account_row_map[(sj, aid)]
        col = get_column_letter(3 + i)
        return f"'{source_flow_sheet}'!{col}{src_row}"

    def ind_ref(name: str, i: int) -> str:
        r = y_row_of.get(name)
        if not r:
            return ""
        col = get_column_letter(3 + i)
        return f"'{y_ind_sheet}'!{col}{r}"

    row = 1
    ws.cell(row=row, column=1, value=f"{sheet_title} — {company_name}").font = Font(name=FONT_NAME, bold=True, size=14)
    ws.cell(row=row, column=4, value="(금액 단위: 억원, 비율/배수/일수 제외)").font = NOTE
    row += 1
    if banner:
        ws.cell(row=row, column=1, value=banner).font = WARN
        row += 1
    row += 1

    # --- A. 회사 개황 ---
    ws.cell(row=row, column=1, value="A. 회사 개황").font = LABEL
    row += 1
    fields = [
        ("기업명", "corp_name"), ("종목코드", "stock_code"), ("대표자", "ceo_nm"),
        ("설립일", "est_dt"), ("업종코드", "induty_code"), ("주소", "adres"),
        ("홈페이지", "hm_url"),
    ]
    for label, key in fields:
        ws.cell(row=row, column=1, value=label)
        val = profile.get(key) if profile else None
        ws.cell(row=row, column=2, value=val if val else "(정보 없음)")
        row += 1
    if not profile:
        ws.cell(row=row, column=1, value="※ company.json 캐시가 없습니다. corp_code_lookup.py를 먼저 실행하세요.").font = NOTE
        row += 1
    row += 1

    # --- B. 신규 기초 항목 (기존 지표_연간에 없는 것만 이 시트에 새로 마련) ---
    base_start_row = row
    extra_base_names = ["재고자산", "비유동자산", "법인세차감전순이익", "영업활동현금흐름"]
    extra_row: dict[str, int] = {}
    extra_hit: dict[str, tuple | None] = {}
    for name in extra_base_names:
        hit = resolve(name)
        ws.cell(row=row, column=1, value=name).font = NOTE
        extra_row[name] = row
        extra_hit[name] = hit
        for i in range(n):
            ref = year_ref(hit, i)
            c = ws.cell(row=row, column=3 + i)
            if ref:
                c.value = f"={ref}"
            c.number_format = "#,##0.0;(#,##0.0);-"
            c.font = NOTE
        row += 1
    ws.cell(row=base_start_row - 0, column=1)  # no-op anchor
    row += 1

    def base_cell(name: str, i: int) -> str:
        col = get_column_letter(3 + i)
        return f"{col}{extra_row[name]}"

    # 청산가치용 추가 기초 항목 (마지막 연도만 쓰지만 전체 연도 계산해 둔다)
    liq_names = ["현금및현금성자산_기말", "유가증권", "투자자산", "유형자산", "무형자산", "기타유동자산"]
    for name in liq_names:
        hit = resolve(name)
        ws.cell(row=row, column=1, value=name).font = NOTE
        extra_row[name] = row
        extra_hit[name] = hit
        for i in range(n):
            ref = year_ref(hit, i)
            c = ws.cell(row=row, column=3 + i)
            if ref:
                c.value = f"={ref}"
            c.number_format = "#,##0.0;(#,##0.0);-"
            c.font = NOTE
        row += 1
    row += 1

    # v0.5.0: CCC·FCF·이자보상배율·DuPont·CF구분용 추가 기초 항목
    # v0.9.7: 감가상각비 추가 (오너 어닝 계산용, investment-thesis-writer의
    # 버핏-멍거 가치평가 시트에서 이 셀을 참조한다)
    extra2_names = [
        "매입채무", "이자비용", "법인세비용", "유형자산의취득",
        "투자활동현금흐름", "재무활동현금흐름", "감가상각비",
    ]
    for name in extra2_names:
        hit = resolve(name)
        ws.cell(row=row, column=1, value=name).font = NOTE
        extra_row[name] = row
        extra_hit[name] = hit
        for i in range(n):
            ref = year_ref(hit, i)
            c = ws.cell(row=row, column=3 + i)
            if ref:
                c.value = f"={ref}"
            c.number_format = "#,##0.0;(#,##0.0);-"
            c.font = NOTE
        row += 1
    # 외환손익 관련은 없는 회사가 많은 게 정상이라 매칭 실패로 취급하지 않는다.
    for name in ["외화환산손익", "파생상품손익"]:
        hit = resolve_optional(name)
        ws.cell(row=row, column=1, value=name).font = NOTE
        extra_row[name] = row
        extra_hit[name] = hit
        for i in range(n):
            ref = year_ref(hit, i)
            c = ws.cell(row=row, column=3 + i)
            if ref:
                c.value = f"={ref}"
            c.number_format = "#,##0.0;(#,##0.0);-"
            c.font = NOTE
        row += 1
    row += 1

    header_row = base_start_row - 1
    ws.cell(row=header_row, column=1, value="(기초 참고값)").font = NOTE
    for i, label in enumerate(y_period_labels):
        ws.cell(row=header_row, column=3 + i, value=label).font = NOTE

    # --- C. 재무비율 ---
    ws.cell(row=row, column=1, value="B. 재무지표").font = LABEL
    row += 1
    ratio_header = row
    ws.cell(row=row, column=1, value="지표")
    for i, label in enumerate(y_period_labels):
        ws.cell(row=row, column=3 + i, value=label)
    style_header(ws, row, 2 + n)
    apply_grid_border(ws, row, row, 1, 2 + n)
    row += 1

    b_row: dict[str, int] = {}

    def write_ratio_row(name: str, formula_fn, fmt="0.0"):
        nonlocal row
        ws.cell(row=row, column=1, value=name)
        b_row[name] = row
        for i in range(n):
            f = formula_fn(i)
            c = ws.cell(row=row, column=3 + i)
            if f:
                c.value = f
            c.number_format = fmt
        apply_grid_border(ws, row, row, 1, 2 + n)
        row += 1

    def b_ref(name: str, i: int) -> str:
        """섹션 B에서 이미 계산해 둔 지표를 같은 시트 안에서 재참조한다."""
        col = get_column_letter(3 + i)
        return f"{col}{b_row[name]}"

    def write_period_header():
        """E~L 각 섹션 표 위에 연도 헤더 행을 쓴다."""
        nonlocal row
        ws.cell(row=row, column=1, value="지표")
        for i, label in enumerate(y_period_labels):
            ws.cell(row=row, column=3 + i, value=label)
        style_header(ws, row, 2 + n)
        apply_grid_border(ws, row, row, 1, 2 + n)
        row += 1

    chart_anchor_row = [1]  # 차트를 세로로 쌓아 내려갈 위치 (리스트로 감싸 클로저에서 갱신)
    CHART_ANCHOR_COL = "N"

    def add_section_chart(
        title: str,
        primary_names: list[str],
        secondary_names: list[str] | None = None,
        primary_ytitle: str = "",
        secondary_ytitle: str = "",
        primary_type: str = "bar",
    ):
        """b_row에 등록된 행 이름들을 계열로 하는 차트를 투자분석 시트
        오른쪽(N열부터)에 세로로 쌓아 추가한다. secondary_names가 있으면
        보조축이 있는 콤보 차트로 만든다 — 1차는 막대, 보조축은 항상 꺾은선으로
        그린다(막대+막대 조합은 Excel에서 축이 뒤바뀌거나 겹치는 문제가 있어
        openpyxl 공식 예제와 동일한 막대+꺾은선 조합만 쓴다). primary_type="line"이면
        보조축 없이 1차 계열 전체를 꺾은선으로 그린다."""
        # 카테고리(연도)는 B섹션 표 헤더 행(ratio_header, 연도 라벨이 있는 행)을 공용으로 쓴다.
        cat_ref = Reference(ws, min_col=3, max_col=2 + n, min_row=ratio_header, max_row=ratio_header)
        if primary_type == "line":
            chart = LineChart()
        else:
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
        chart.title = title
        chart.style = 10
        chart.y_axis.title = primary_ytitle
        chart.x_axis.title = "기간"
        chart.height = 8.5
        chart.width = 22
        for name in primary_names:
            r = b_row.get(name)
            if not r:
                continue
            data_ref = Reference(ws, min_col=3, max_col=2 + n, min_row=r, max_row=r)
            chart.series.append(Series(data_ref, title=name))
        chart.set_categories(cat_ref)

        if secondary_names:
            chart2 = LineChart()
            for name in secondary_names:
                r = b_row.get(name)
                if not r:
                    continue
                data_ref = Reference(ws, min_col=3, max_col=2 + n, min_row=r, max_row=r)
                chart2.series.append(Series(data_ref, title=name))
            chart2.set_categories(cat_ref)
            chart2.y_axis.axId = 200
            chart2.y_axis.title = secondary_ytitle
            chart2.y_axis.axPos = "r"
            # openpyxl 공식 콤보 차트 패턴(막대+꺾은선): 1차 축이 보조축의 최댓값
            # 쪽에서 교차하도록 지정해야 Excel에서 보조축이 오른쪽에 분리되어 나온다.
            chart.y_axis.crosses = "max"
            chart += chart2

        ws.add_chart(chart, f"{CHART_ANCHOR_COL}{chart_anchor_row[0]}")
        chart_anchor_row[0] += 18

    ws.cell(row=row, column=1, value="[건전성]").font = Font(name=FONT_NAME, italic=True)
    row += 1
    write_ratio_row("자기자본비율(%)", lambda i: f"={ind_ref('자기자본비율', i)}" if ind_ref('자기자본비율', i) else "")
    write_ratio_row("부채비율(%)", lambda i: f"={ind_ref('부채비율', i)}" if ind_ref('부채비율', i) else "")
    write_ratio_row(
        "유동비율(%)",
        lambda i: f"=IFERROR({ind_ref('유동자산', i)}/{ind_ref('유동부채', i)}*100,NA())"
        if ind_ref("유동자산", i) and ind_ref("유동부채", i) else "",
    )
    write_ratio_row(
        "당좌비율(%)",
        lambda i: f"=IFERROR(({ind_ref('유동자산', i)}-{base_cell('재고자산', i)})/{ind_ref('유동부채', i)}*100,NA())"
        if ind_ref("유동자산", i) and ind_ref("유동부채", i) else "",
    )
    write_ratio_row(
        "고정비율(%)",
        lambda i: f"=IFERROR({base_cell('비유동자산', i)}/{ind_ref('자본총계', i)}*100,NA())"
        if ind_ref("자본총계", i) else "",
    )
    write_ratio_row(
        "고정장기적합율(%)",
        lambda i: f"=IFERROR({base_cell('비유동자산', i)}/({ind_ref('자본총계', i)}+({ind_ref('부채총계', i)}-{ind_ref('유동부채', i)}))*100,NA())"
        if ind_ref("자본총계", i) and ind_ref("부채총계", i) and ind_ref("유동부채", i) else "",
    )
    write_ratio_row(
        "순운전자본대총자본비율(%)",
        lambda i: f"=IFERROR(({ind_ref('유동자산', i)}-{ind_ref('유동부채', i)})/{ind_ref('자산총계', i)}*100,NA())"
        if ind_ref("유동자산", i) and ind_ref("유동부채", i) and ind_ref("자산총계", i) else "",
    )
    write_ratio_row(
        "이자보상배율(배)",
        lambda i: f"=IFERROR({ind_ref('영업이익', i)}/{base_cell('이자비용', i)},NA())"
        if ind_ref("영업이익", i) else "",
        fmt="0.00",
    )

    ws.cell(row=row, column=1, value="[수익성]").font = Font(name=FONT_NAME, italic=True)
    row += 1
    write_ratio_row("매출총이익률(%)", lambda i: f"={ind_ref('매출총이익률', i)}" if ind_ref('매출총이익률', i) else "")
    write_ratio_row("영업이익률(%)", lambda i: f"={ind_ref('영업이익률', i)}" if ind_ref('영업이익률', i) else "")
    write_ratio_row(
        "세전순이익률(%)",
        lambda i: f"=IFERROR({base_cell('법인세차감전순이익', i)}/{ind_ref('매출액', i)}*100,NA())"
        if ind_ref("매출액", i) else "",
    )
    write_ratio_row("순이익률(%)", lambda i: f"={ind_ref('순이익률', i)}" if ind_ref('순이익률', i) else "")
    write_ratio_row(
        "ROE(%)",
        lambda i: f"=IFERROR({ind_ref('당기순이익', i)}/{ind_ref('자본총계', i)}*100,NA())"
        if ind_ref("당기순이익", i) and ind_ref("자본총계", i) else "",
    )
    write_ratio_row(
        "ROA(%)",
        lambda i: f"=IFERROR({ind_ref('당기순이익', i)}/{ind_ref('자산총계', i)}*100,NA())"
        if ind_ref("당기순이익", i) and ind_ref("자산총계", i) else "",
    )

    ws.cell(row=row, column=1, value="[성장성]").font = Font(name=FONT_NAME, italic=True)
    row += 1

    def yoy_fn(metric: str):
        def f(i):
            if i == 0:
                return ""
            cur, prev = ind_ref(metric, i), ind_ref(metric, i - 1)
            if not cur or not prev:
                return ""
            return f"=IFERROR(({cur}-{prev})/{prev}*100,NA())"
        return f

    def yoy_fn_base(name: str):
        def f(i):
            if i == 0:
                return ""
            cur, prev = base_cell(name, i), base_cell(name, i - 1)
            return f"=IFERROR(({cur}-{prev})/{prev}*100,NA())"
        return f

    write_ratio_row("매출성장률(%, YoY)", yoy_fn("매출액"))
    write_ratio_row("영업이익성장률(%, YoY)", yoy_fn("영업이익"))
    write_ratio_row("순이익성장률(%, YoY)", yoy_fn("당기순이익"))
    write_ratio_row("총자산증가율(%, YoY)", yoy_fn("자산총계"))
    write_ratio_row("자기자본증가율(%, YoY)", yoy_fn("자본총계"))
    write_ratio_row("유형자산증가율(%, YoY)", yoy_fn_base("유형자산"))

    ws.cell(row=row, column=1, value="[활동성]").font = Font(name=FONT_NAME, italic=True)
    row += 1
    write_ratio_row(
        "총자산회전율(회)",
        lambda i: f"=IFERROR({ind_ref('매출액', i)}/{ind_ref('자산총계', i)},NA())"
        if ind_ref("매출액", i) and ind_ref("자산총계", i) else "",
        fmt="0.00",
    )
    write_ratio_row(
        "매출채권회전율(회)",
        lambda i: f"=IFERROR({ind_ref('매출액', i)}/{ind_ref('매출채권및기타채권', i)},NA())"
        if ind_ref("매출액", i) and ind_ref("매출채권및기타채권", i) else "",
        fmt="0.00",
    )
    write_ratio_row(
        "자기자본회전율(회)",
        lambda i: f"=IFERROR({ind_ref('매출액', i)}/{ind_ref('자본총계', i)},NA())"
        if ind_ref("매출액", i) and ind_ref("자본총계", i) else "",
        fmt="0.00",
    )
    write_ratio_row(
        "유형자산회전율(회)",
        lambda i: f"=IFERROR({ind_ref('매출액', i)}/{base_cell('유형자산', i)},NA())"
        if ind_ref("매출액", i) else "",
        fmt="0.00",
    )
    write_ratio_row(
        "재고자산회전율(회)",
        lambda i: f"=IFERROR({ind_ref('매출원가', i)}/{base_cell('재고자산', i)},NA())"
        if ind_ref("매출원가", i) else "",
        fmt="0.00",
    )
    write_ratio_row(
        "매입채무회전율(회)",
        lambda i: f"=IFERROR({ind_ref('매출원가', i)}/{base_cell('매입채무', i)},NA())"
        if ind_ref("매출원가", i) else "",
        fmt="0.00",
    )
    row += 1

    # --- B섹션 4개 그룹 각각 콤보(보조축) 막대그래프 ---
    add_section_chart(
        "건전성 지표",
        ["자기자본비율(%)", "부채비율(%)", "유동비율(%)", "당좌비율(%)", "고정비율(%)", "고정장기적합율(%)", "순운전자본대총자본비율(%)"],
        ["이자보상배율(배)"],
        primary_ytitle="%", secondary_ytitle="배",
    )
    add_section_chart(
        "수익성 지표",
        ["매출총이익률(%)", "영업이익률(%)", "세전순이익률(%)", "순이익률(%)", "ROE(%)", "ROA(%)"],
        primary_ytitle="%",
    )
    add_section_chart(
        "성장성 지표",
        ["매출성장률(%, YoY)", "영업이익성장률(%, YoY)", "순이익성장률(%, YoY)", "총자산증가율(%, YoY)", "자기자본증가율(%, YoY)", "유형자산증가율(%, YoY)"],
        primary_ytitle="%",
    )
    add_section_chart(
        "활동성 지표",
        ["총자산회전율(회)", "매출채권회전율(회)", "자기자본회전율(회)", "유형자산회전율(회)", "재고자산회전율(회)", "매입채무회전율(회)"],
        primary_ytitle="회",
    )

    # --- D. 위험 신호 점검 ---
    ws.cell(row=row, column=1, value="C. 위험 신호 점검").font = LABEL
    row += 1
    risk_header = row
    ws.cell(row=row, column=1, value="점검 항목")
    for i, label in enumerate(y_period_labels):
        ws.cell(row=row, column=3 + i, value=label)
    style_header(ws, row, 2 + n)
    apply_grid_border(ws, row, row, 1, 2 + n)
    row += 1

    def write_risk_row(name: str, formula_fn):
        nonlocal row
        ws.cell(row=row, column=1, value=name)
        for i in range(n):
            f = formula_fn(i)
            c = ws.cell(row=row, column=3 + i)
            if f:
                c.value = f
        apply_grid_border(ws, row, row, 1, 2 + n)
        row += 1

    write_risk_row(
        "유동부채 > 유동자산",
        lambda i: f"=IF({ind_ref('유동부채', i)}>{ind_ref('유동자산', i)},\"⚠ 위험\",\"양호\")"
        if ind_ref("유동부채", i) and ind_ref("유동자산", i) else "",
    )
    write_risk_row(
        "차입금 과다 (자기자본비율<20%)",
        lambda i: f"=IF({ind_ref('자기자본비율', i)}<20,\"⚠ 위험\",\"양호\")" if ind_ref("자기자본비율", i) else "",
    )
    write_risk_row(
        "순자산 마이너스 (채무초과)",
        lambda i: f"=IF({ind_ref('자본총계', i)}<0,\"⚠ 위험\",\"양호\")" if ind_ref("자본총계", i) else "",
    )

    def receivable_spike_fn(i):
        if i == 0:
            return ""
        rev_c, rev_p = ind_ref("매출액", i), ind_ref("매출액", i - 1)
        rec_c, rec_p = ind_ref("매출채권및기타채권", i), ind_ref("매출채권및기타채권", i - 1)
        if not (rev_c and rev_p and rec_c and rec_p):
            return ""
        return (
            f"=IFERROR(IF((({rec_c}-{rec_p})/{rec_p}*100)-(({rev_c}-{rev_p})/{rev_p}*100)>20,"
            f"\"⚠ 위험(매출채권 급증)\",\"양호\"),\"\")"
        )

    write_risk_row("매출채권 급증 (매출 증가율 대비 +20%p 이상)", receivable_spike_fn)
    write_risk_row(
        "영업활동현금흐름 마이너스",
        lambda i: f"=IF({base_cell('영업활동현금흐름', i)}<0,\"⚠ 위험\",\"양호\")",
    )
    write_risk_row(
        "이자보상배율 1 미만 (이자도 못 갚는 수준)",
        lambda i: f"=IFERROR(IF({ind_ref('영업이익', i)}/{base_cell('이자비용', i)}<1,\"⚠ 위험\",\"양호\"),\"\")"
        if ind_ref("영업이익", i) else "",
    )
    row += 1

    # --- E. 청산가치 (자산가치주 체크, 최신 연도 기준) ---
    ws.cell(row=row, column=1, value="D. 청산가치 (자산가치주 체크 · 최신 연도 기준)").font = LABEL
    row += 1
    last_i = n - 1
    haircut_items = [
        ("현금및현금성자산", "현금및현금성자산_기말", 1.00),
        ("유가증권", "유가증권", 1.00),
        ("매출채권및기타채권", None, 0.85),  # 지표_연간에서 참조
        ("재고자산", "재고자산", 0.50),
        ("투자자산", "투자자산", 0.50),
        ("유형자산", "유형자산", 0.50),
        ("무형자산", "무형자산", 0.00),
        ("기타유동자산", "기타유동자산", 0.00),
    ]
    ws.cell(row=row, column=1, value="항목")
    ws.cell(row=row, column=2, value="적용비율")
    ws.cell(row=row, column=3, value="조정가치")
    style_header(ws, row, 3)
    d_header = row
    row += 1
    adj_asset_rows = []
    for label, base_key, pct in haircut_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=pct).number_format = "0%"
        if base_key:
            src = base_cell(base_key, last_i)
        else:
            src = ind_ref("매출채권및기타채권", last_i)
        c = ws.cell(row=row, column=3)
        if src and pct > 0:
            c.value = f"={src}*{pct}"
        elif src:
            c.value = 0
        c.number_format = "#,##0.0"
        adj_asset_rows.append(row)
        row += 1
    sum_row = row
    ws.cell(row=row, column=1, value="조정자산 합계").font = LABEL
    ws.cell(row=row, column=3, value=f"=SUM(C{adj_asset_rows[0]}:C{adj_asset_rows[-1]})").number_format = "#,##0.0"
    row += 1
    debt_ref = ind_ref("부채총계", last_i)
    ws.cell(row=row, column=1, value="총부채(부채총계)").font = LABEL
    if debt_ref:
        ws.cell(row=row, column=3, value=f"={debt_ref}").number_format = "#,##0.0"
    row += 1
    liq_row = row
    ws.cell(row=row, column=1, value="청산가치 (조정자산 − 총부채)").font = LABEL
    if debt_ref:
        ws.cell(row=row, column=3, value=f"=C{sum_row}-C{row - 1}").number_format = "#,##0.0"
    apply_grid_border(ws, d_header, liq_row, 1, 3)
    row += 2

    # --- E. CCC (현금전환주기) ---
    ws.cell(row=row, column=1, value="E. 현금전환주기 (CCC)").font = LABEL
    row += 1
    write_period_header()
    write_ratio_row(
        "매출채권회수기간(일)",
        lambda i: f"=IFERROR(365/{b_ref('매출채권회전율(회)', i)},NA())",
    )
    write_ratio_row(
        "재고자산처리기간(일)",
        lambda i: f"=IFERROR(365/{b_ref('재고자산회전율(회)', i)},NA())",
    )
    write_ratio_row(
        "매입채무지불기간(일)",
        lambda i: f"=IFERROR(365/{b_ref('매입채무회전율(회)', i)},NA())",
    )
    write_ratio_row(
        "CCC = 회수+처리-지불(일)",
        lambda i: (
            f"=IFERROR({b_ref('매출채권회수기간(일)', i)}+{b_ref('재고자산처리기간(일)', i)}"
            f"-{b_ref('매입채무지불기간(일)', i)},NA())"
        ),
    )
    ws.cell(row=row, column=1, value="※ CCC가 짧을수록(마이너스에 가까울수록) 운전자본 부담이 적은 우량한 구조입니다.").font = NOTE
    row += 2

    # --- F. FCF (잉여현금흐름) ---
    add_section_chart(
        "현금전환주기 (CCC)",
        ["매출채권회수기간(일)", "재고자산처리기간(일)", "매입채무지불기간(일)", "CCC = 회수+처리-지불(일)"],
        primary_ytitle="일",
    )

    ws.cell(row=row, column=1, value="F. 잉여현금흐름 (FCF)").font = LABEL
    row += 1
    write_period_header()
    write_ratio_row(
        "FCF = 영업활동현금흐름 - CAPEX",
        lambda i: f"=IFERROR({base_cell('영업활동현금흐름', i)}-ABS({base_cell('유형자산의취득', i)}),NA())",
        fmt="#,##0.0;(#,##0.0);-",
    )
    write_ratio_row(
        "FCF마진(%, FCF/매출액)",
        lambda i: f"=IFERROR({b_ref('FCF = 영업활동현금흐름 - CAPEX', i)}/{ind_ref('매출액', i)}*100,NA())"
        if ind_ref("매출액", i) else "",
    )
    row += 1

    # --- G. 현금흐름 3단 구분 ---
    add_section_chart(
        "잉여현금흐름 (FCF)",
        ["FCF = 영업활동현금흐름 - CAPEX"],
        ["FCF마진(%, FCF/매출액)"],
        primary_ytitle="억원", secondary_ytitle="%",
    )

    ws.cell(row=row, column=1, value="G. 현금흐름 3단 구분").font = LABEL
    row += 1
    write_period_header()
    write_ratio_row(
        "영업활동현금흐름",
        lambda i: f"={base_cell('영업활동현금흐름', i)}",
        fmt="#,##0.0;(#,##0.0);-",
    )
    write_ratio_row(
        "투자활동현금흐름",
        lambda i: f"={base_cell('투자활동현금흐름', i)}" if extra_hit.get("투자활동현금흐름") else "",
        fmt="#,##0.0;(#,##0.0);-",
    )
    write_ratio_row(
        "재무활동현금흐름",
        lambda i: f"={base_cell('재무활동현금흐름', i)}" if extra_hit.get("재무활동현금흐름") else "",
        fmt="#,##0.0;(#,##0.0);-",
    )
    write_ratio_row(
        "현금 순증감 (3단 합계, 검증용)",
        lambda i: (
            f"=IFERROR({b_ref('영업활동현금흐름', i)}+{b_ref('투자활동현금흐름', i)}+{b_ref('재무활동현금흐름', i)},NA())"
        ),
        fmt="#,##0.0;(#,##0.0);-",
    )
    ws.cell(row=row, column=1, value="※ 검증용 합계는 위 지표 시트의 '현금및현금성자산의증가'와 대체로 비슷해야 합니다(환율 변동 등으로 소폭 차이 가능).").font = NOTE
    row += 2

    # --- H. DuPont 분해 (ROE) ---
    add_section_chart(
        "현금흐름 3단 구분",
        ["영업활동현금흐름", "투자활동현금흐름", "재무활동현금흐름", "현금 순증감 (3단 합계, 검증용)"],
        primary_ytitle="억원",
    )

    ws.cell(row=row, column=1, value="H. DuPont 분해 (ROE = 순이익률 × 총자산회전율 × 레버리지)").font = LABEL
    row += 1
    write_period_header()
    write_ratio_row(
        "레버리지 (자산/자기자본, 배)",
        lambda i: f"=IFERROR({ind_ref('자산총계', i)}/{ind_ref('자본총계', i)},NA())"
        if ind_ref("자산총계", i) and ind_ref("자본총계", i) else "",
        fmt="0.00",
    )
    write_ratio_row(
        "ROE 검증 (순이익률×회전율×레버리지, %)",
        lambda i: (
            f"=IFERROR({b_ref('순이익률(%)', i)}*{b_ref('총자산회전율(회)', i)}*{b_ref('레버리지 (자산/자기자본, 배)', i)},NA())"
        ),
    )
    ws.cell(row=row, column=1, value="※ 위 B섹션의 ROE(%)와 거의 같아야 정상입니다. 크게 다르면 어딘가 계정 매칭이 잘못됐을 가능성이 있습니다.").font = NOTE
    row += 2

    # --- I. ROIC / NOPLAT (간이 버전) ---
    add_section_chart(
        "DuPont 분해 (ROE)",
        ["레버리지 (자산/자기자본, 배)"],
        ["ROE 검증 (순이익률×회전율×레버리지, %)"],
        primary_ytitle="배", secondary_ytitle="%",
    )

    ws.cell(row=row, column=1, value="I. ROIC / NOPLAT (간이 계산)").font = LABEL
    row += 1
    write_period_header()
    write_ratio_row(
        "실효세율(%)",
        lambda i: f"=IFERROR({base_cell('법인세비용', i)}/{base_cell('법인세차감전순이익', i)}*100,NA())",
    )
    write_ratio_row(
        "NOPLAT = 영업이익×(1-실효세율)",
        lambda i: f"=IFERROR({ind_ref('영업이익', i)}*(1-{b_ref('실효세율(%)', i)}/100),NA())"
        if ind_ref("영업이익", i) else "",
        fmt="#,##0.0;(#,##0.0);-",
    )
    write_ratio_row(
        "투하자본(간이) = 자기자본+비유동부채",
        lambda i: f"=IFERROR({ind_ref('자본총계', i)}+({ind_ref('부채총계', i)}-{ind_ref('유동부채', i)}),NA())"
        if ind_ref("자본총계", i) and ind_ref("부채총계", i) and ind_ref("유동부채", i) else "",
        fmt="#,##0.0;(#,##0.0);-",
    )
    write_ratio_row(
        "ROIC(간이, %)",
        lambda i: (
            f"=IFERROR({b_ref('NOPLAT = 영업이익×(1-실효세율)', i)}/{b_ref('투하자본(간이) = 자기자본+비유동부채', i)}*100,NA())"
        ),
    )
    ws.cell(row=row, column=1, value="※ 간이 버전입니다. 정교한 ROIC은 이자부부채만 골라 투하자본을 계산해야 하는데, 그러려면 단기차입금·사채 등 세부 계정이 더 필요해 여기서는 자기자본+비유동부채로 근사했습니다.").font = NOTE
    row += 2

    # --- J. 자산·부채 구성비 변화 (간이) ---
    add_section_chart(
        "ROIC / NOPLAT (간이)",
        ["NOPLAT = 영업이익×(1-실효세율)", "투하자본(간이) = 자기자본+비유동부채"],
        ["실효세율(%)", "ROIC(간이, %)"],
        primary_ytitle="억원", secondary_ytitle="%",
    )

    ws.cell(row=row, column=1, value="J. 자산·부채 구성비 변화 (간이)").font = LABEL
    row += 1
    write_period_header()
    write_ratio_row(
        "유동자산 비중(%)",
        lambda i: f"=IFERROR({ind_ref('유동자산', i)}/{ind_ref('자산총계', i)}*100,NA())"
        if ind_ref("유동자산", i) and ind_ref("자산총계", i) else "",
    )
    write_ratio_row(
        "비유동자산 비중(%)",
        lambda i: f"=IFERROR({base_cell('비유동자산', i)}/{ind_ref('자산총계', i)}*100,NA())"
        if ind_ref("자산총계", i) else "",
    )
    write_ratio_row(
        "유동부채 비중(%, 총부채 대비)",
        lambda i: f"=IFERROR({ind_ref('유동부채', i)}/{ind_ref('부채총계', i)}*100,NA())"
        if ind_ref("유동부채", i) and ind_ref("부채총계", i) else "",
    )
    write_ratio_row(
        "비유동부채 비중(%, 총부채 대비)",
        lambda i: f"=IFERROR(({ind_ref('부채총계', i)}-{ind_ref('유동부채', i)})/{ind_ref('부채총계', i)}*100,NA())"
        if ind_ref("부채총계", i) and ind_ref("유동부채", i) else "",
    )
    ws.cell(row=row, column=1, value="※ 유동부채 비중이 계속 커지면 단기 자금조달 의존도가 높아지고 있다는 신호일 수 있습니다.").font = NOTE
    row += 2

    # --- K. 외환손익 (있는 회사만) ---
    add_section_chart(
        "자산·부채 구성비 변화",
        ["유동자산 비중(%)", "비유동자산 비중(%)", "유동부채 비중(%, 총부채 대비)", "비유동부채 비중(%, 총부채 대비)"],
        primary_ytitle="%",
    )

    ws.cell(row=row, column=1, value="K. 외환손익").font = LABEL
    row += 1
    write_period_header()
    has_fx = extra_hit.get("외화환산손익") or extra_hit.get("파생상품손익")
    write_ratio_row(
        "외화환산손익",
        lambda i: f"={base_cell('외화환산손익', i)}" if extra_hit.get("외화환산손익") else "",
        fmt="#,##0.0;(#,##0.0);-",
    )
    write_ratio_row(
        "파생상품손익",
        lambda i: f"={base_cell('파생상품손익', i)}" if extra_hit.get("파생상품손익") else "",
        fmt="#,##0.0;(#,##0.0);-",
    )
    write_ratio_row(
        "외환손익 합계",
        lambda i: f"=IFERROR({b_ref('외화환산손익', i)}+{b_ref('파생상품손익', i)},NA())",
        fmt="#,##0.0;(#,##0.0);-",
    )
    if not has_fx:
        ws.cell(row=row, column=1, value="※ 이 회사 공시에서 외화환산손익·파생상품손익 계정을 찾지 못했습니다. 외환 노출이 적거나 다른 계정명을 쓰는 회사일 수 있습니다(매칭 실패 경고에는 포함하지 않았습니다).").font = NOTE
    row += 2

    # --- F. 주가 연동 지표 (KRX 필요) ---
    add_section_chart(
        "외환손익",
        ["외화환산손익", "파생상품손익", "외환손익 합계"],
        primary_ytitle="억원",
    )

    ws.cell(row=row, column=1, value="L. 주가 연동 지표 (KRX 종가 기준)").font = LABEL
    row += 1
    if not stock_code:
        ws.cell(row=row, column=1, value="※ 종목코드가 없어 주가 지표를 건너뜁니다.").font = NOTE
        row += 2
    else:
        write_period_header()

        # 연도별로 한 번만 조회해두고, 아래에서 행(지표)×열(연도)로 펼쳐 쓴다.
        price_by_i: list[dict | None] = []
        mktcap_by_i: list[float | None] = []
        for i, year in enumerate(year_list):
            ref_date = period_dates[i] if period_dates and i < len(period_dates) and period_dates[i] else f"{year}1231"
            price = load_price(stock_code, ref_date)
            price_by_i.append(price)
            if price:
                close = _fmt_num(price.get("close_price"))
                listed = _fmt_num(price.get("listed_shares"))
                mktcap = _fmt_num(price.get("market_cap")) or (close * listed if close and listed else None)
                # 당기순이익·자본총계 등 분모가 이미 억원 단위이므로, PER/PBR/PSR 비율이
                # 맞으려면 시가총액도 반드시 억원으로 맞춰야 한다.
                if mktcap is not None:
                    mktcap = mktcap / UNIT_DIVISOR
            else:
                mktcap = None
            mktcap_by_i.append(mktcap)

        write_ratio_row(
            "기준일(종가)",
            lambda i: (price_by_i[i] or {}).get("used_date") or "",
            fmt="General",
        )
        write_ratio_row(
            "종가",
            lambda i: _fmt_num((price_by_i[i] or {}).get("close_price")),
            fmt="#,##0",
        )
        write_ratio_row(
            "시가총액",
            lambda i: mktcap_by_i[i],
            fmt="#,##0.0",
        )
        write_ratio_row(
            "PER(배)",
            lambda i: f"=IFERROR({mktcap_by_i[i]}/{ind_ref('당기순이익', i)},NA())"
            if mktcap_by_i[i] is not None and ind_ref("당기순이익", i) else "",
            fmt="0.0",
        )
        write_ratio_row(
            "PBR(배)",
            lambda i: f"=IFERROR({mktcap_by_i[i]}/{ind_ref('자본총계', i)},NA())"
            if mktcap_by_i[i] is not None and ind_ref("자본총계", i) else "",
            fmt="0.00",
        )
        write_ratio_row(
            "PSR(배)",
            lambda i: f"=IFERROR({mktcap_by_i[i]}/{ind_ref('매출액', i)},NA())"
            if mktcap_by_i[i] is not None and ind_ref("매출액", i) else "",
            fmt="0.00",
        )
        any_price = any(price_by_i)
        if not any_price:
            ws.cell(row=row, column=1, value=(
                "※ 이 회사의 가격 캐시가 하나도 없습니다. "
                "scripts/fetch_stock_price.py --auth-key <KRX 인증키> "
                f"{stock_code} <YYYYMMDD> 를 연도별 결산일 기준으로 먼저 실행하세요."
            )).font = WARN
            row += 1
        row += 1

    # --- G. 배당 · 대주주 · 자기주식 (DART 추가 공시) ---
    if stock_code:
        add_section_chart(
            "주가 연동 지표 (PER/PBR/PSR)",
            ["PER(배)", "PBR(배)", "PSR(배)"],
            primary_ytitle="배",
            primary_type="line",
        )

    ws.cell(row=row, column=1, value="M. 배당 · 대주주 · 자기주식").font = LABEL
    row += 1
    latest_year = year_list[-1] if year_list else None
    extra = load_extra_disclosures(corp_code, latest_year) if latest_year else None
    if not extra:
        ws.cell(row=row, column=1, value=(
            "※ 추가 공시 캐시가 없습니다. scripts/fetch_extra_disclosures.py를 먼저 실행하세요."
        )).font = WARN
        row += 1
    else:
        div = extra.get("배당", {})
        div_list = div.get("list", []) if isinstance(div, dict) else []
        payout = next((x.get("thstrm") for x in div_list if "배당성향" in (x.get("se") or "")), None)
        dps = next((x.get("thstrm") for x in div_list if "주당" in (x.get("se") or "") and "현금" in (x.get("se") or "")), None)
        ws.cell(row=row, column=1, value="배당성향(%)")
        ws.cell(row=row, column=2, value=payout or "(공시 없음)")
        row += 1
        ws.cell(row=row, column=1, value="주당 현금배당금")
        ws.cell(row=row, column=2, value=dps or "(공시 없음)")
        row += 1

        holders = extra.get("최대주주현황", {})
        holder_list = holders.get("list", []) if isinstance(holders, dict) else []
        ws.cell(row=row, column=1, value="대주주 명단 (최신 보고서 기준)").font = Font(name=FONT_NAME, italic=True)
        row += 1
        for h in holder_list[:5]:
            ws.cell(row=row, column=1, value=h.get("nm", ""))
            ws.cell(row=row, column=2, value=h.get("trmend_posesn_stock_qota_rt", ""))
            row += 1

        treasury = extra.get("자기주식현황", {})
        treasury_list = treasury.get("list", []) if isinstance(treasury, dict) else []
        if treasury_list:
            ws.cell(row=row, column=1, value="자기주식 변동(최신 보고서)").font = Font(name=FONT_NAME, italic=True)
            row += 1
            for t in treasury_list[:3]:
                ws.cell(row=row, column=1, value=t.get("acqs_mth1", t.get("trmend_rmnd_stkcnt", "")))
                row += 1
    row += 1

    # --- H. 투자 판단 (정성 평가 — 사용자가 직접 채우는 템플릿) ---
    ws.cell(row=row, column=1, value="N. 투자 판단 (자동 평가 · A~E)").font = LABEL
    row += 1
    ws.cell(row=row, column=1, value="항목")
    ws.cell(row=row, column=2, value="평가(A~E)")
    ws.cell(row=row, column=3, value="근거 메모")
    style_header(ws, row, 3)
    row += 1

    # 등급 판정을 위해 원자료(cache)에서 지표값을 직접 계산한다.
    # (시트 셀은 수식이라 openpyxl로는 값을 읽을 수 없어 원자료를 다시 계산한다)
    def series(key: str) -> list[float | None]:
        hit = resolve_metric(key, y_account_row_map, y_account_name_map) if key in METRIC_RULES else None
        out: list[float | None] = []
        for year in year_list:
            fy = (reports or {}).get(year, {}).get("11011")
            val = amount_lookup(fy, hit[0], hit[1]) if (fy and hit) else None
            if val is None and annualized_series and key in annualized_series:
                override = annualized_series[key]
                if isinstance(override, dict) and year in override:
                    val = override[year]
            out.append(val)
        return out

    def ratio(nums: list[float | None], dens: list[float | None], mult: float = 100.0) -> list[float | None]:
        out: list[float | None] = []
        for a, b in zip(nums, dens):
            out.append(a / b * mult if (a is not None and b not in (None, 0)) else None)
        return out

    def yoy(vals: list[float | None]) -> list[float | None]:
        out: list[float | None] = [None]
        for prev, cur in zip(vals, vals[1:]):
            out.append((cur - prev) / prev * 100 if (prev not in (None, 0) and cur is not None) else None)
        return out

    s_revenue = series("매출액")
    s_op = series("영업이익")
    s_ni = series("당기순이익")
    s_assets = series("자산총계")
    s_equity = series("자본총계")
    s_liab = series("부채총계")
    s_ca = series("유동자산")
    s_cl = series("유동부채")

    raw_metrics = {
        "자기자본비율": ratio(s_equity, s_assets),
        "부채비율": ratio(s_liab, s_equity),
        "유동비율": ratio(s_ca, s_cl),
        "영업이익률": ratio(s_op, s_revenue),
        "ROE": ratio(s_ni, s_equity),
        "ROA": ratio(s_ni, s_assets),
        "매출성장률": yoy(s_revenue),
        "영업이익성장률": yoy(s_op),
        "총자산회전율": ratio(s_revenue, s_assets, mult=1.0),
    }

    # 주가 지표(PER/PBR)는 가격 캐시가 있을 때만 계산한다.
    per_list: list[float | None] = []
    pbr_list: list[float | None] = []
    price_available = False
    if stock_code:
        for i, year in enumerate(year_list):
            ref_date = period_dates[i] if period_dates and i < len(period_dates) and period_dates[i] else f"{year}1231"
            price = load_price(stock_code, ref_date)
            if not price:
                per_list.append(None)
                pbr_list.append(None)
                continue
            close = _fmt_num(price.get("close_price"))
            listed = _fmt_num(price.get("listed_shares"))
            mc = _fmt_num(price.get("market_cap")) or (close * listed if close and listed else None)
            if mc is None:
                per_list.append(None)
                pbr_list.append(None)
                continue
            price_available = True
            ni, eq = s_ni[i], s_equity[i]
            per_list.append(mc / ni if ni not in (None, 0) else None)
            pbr_list.append(mc / eq if eq not in (None, 0) else None)
    if price_available:
        raw_metrics["PER"] = per_list
        raw_metrics["PBR"] = pbr_list

    row = render_investment_judgement(
        ws, row, n, y_period_labels, raw_metrics, extra, price_available
    )

    ws.cell(row=row, column=1, value=(
        "※ 등급 규칙: 최근 5개년 중 기준 충족 연수로 A(5년)~E(0~1년)를 매기되, "
        "기준 미달이어도 5년 내내 수치가 개선되면 A로 승격합니다. "
        "기준값은 첨부 참고자료(투자판단 항목 평가기준)를 따랐습니다."
    )).font = NOTE
    row += 1
    ws.cell(row=row, column=1, value=(
        "※ '사업역량'은 공시 수치로 판단할 수 없어 직접 입력 항목으로 남겨두었습니다."
    )).font = NOTE

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    for i in range(max(n, 5)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 15

    return warnings


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("corp_code")
    ap.add_argument("company_name")
    ap.add_argument("--years", type=int, default=5)
    ap.add_argument("--quarters", type=int, default=12)
    ap.add_argument("--outdir", default="/mnt/user-data/outputs")
    ap.add_argument(
        "--period", choices=["annual", "quarterly", "both"], default="both",
        help="annual: 연간 시트만 담은 파일 하나, quarterly: 분기 시트만 담은 파일 하나, "
             "both(기본값): 기존처럼 한 파일에 다 담기",
    )
    args = ap.parse_args()

    reports = load_cache(args.corp_code)
    if not reports:
        print(f"ERROR: {args.corp_code}에 대한 캐시 데이터가 없습니다. fetch_financials.py를 먼저 실행하세요.", file=sys.stderr)
        sys.exit(1)

    want_quarterly = args.period in ("quarterly", "both")
    want_annual = args.period in ("annual", "both")

    quarter_plan = build_quarter_plan(reports, args.quarters) if want_quarterly else []
    year_list = sorted([y for y in reports if "11011" in reports[y]], reverse=True)[: args.years] if want_annual else []
    year_list.sort()  # 오래된 -> 최근

    # 최신 완결연도 다음 해가 아직 사업보고서는 없지만 분기/반기보고서는 있으면,
    # 그 해를 "추정 연도"로 다룬다(사용자 확정 사양 — v0.15.0).
    estimated_year: dict | None = None
    estimated_extra_periods: list[tuple[dict, str]] = []
    if want_annual and year_list:
        candidate_year = str(int(year_list[-1]) + 1)
        latest_code = _latest_reprt_code(reports, candidate_year)
        if latest_code and latest_code != "11011":
            field = "thstrm_add_amount" if latest_code == "11014" else "thstrm_amount"
            cur_data = reports[candidate_year][latest_code]
            prior_year = year_list[-1]
            prior_data = reports.get(prior_year, {}).get(latest_code)
            if prior_data:  # 전년 동기 비교 대상이 있어야 비율 계산이 가능하다
                cur_label = f"{candidate_year}_추정기준"
                prior_label = f"{prior_year}_추정기준"
                estimated_extra_periods = [(cur_data, cur_label), (prior_data, prior_label)]
                estimated_year = {
                    "year": candidate_year, "prior_year": prior_year,
                    "cur_label": cur_label, "prior_label": prior_label,
                    "bs_label": cur_label, "field": field,
                    "latest_code": latest_code,
                }
            else:
                print(
                    f"WARNING: {candidate_year}년 {REPRT_NAMES.get(latest_code, latest_code)}는 있지만 "
                    f"{prior_year}년 동기간 비교 데이터가 없어 {candidate_year}(E) 추정을 건너뜁니다.",
                    file=sys.stderr,
                )

    if want_quarterly and len(quarter_plan) < args.quarters:
        print(
            f"WARNING: 요청한 {args.quarters}분기 중 {len(quarter_plan)}분기만 채웠습니다 "
            f"(공시 지연 또는 상장 이력 부족 가능).",
            file=sys.stderr,
        )
    if want_annual and len(year_list) < args.years:
        print(
            f"WARNING: 요청한 {args.years}개년 중 {len(year_list)}개년만 채웠습니다.",
            file=sys.stderr,
        )

    wb = Workbook()
    wb.remove(wb.active)

    cell_index = write_raw_sheet(wb, quarter_plan, year_list, reports, extra_periods=estimated_extra_periods)

    all_missing: dict[str, list[str]] = {}
    if quarter_plan:
        q_row_map, q_name_map, q_labels = build_quarterly_sheet(wb, quarter_plan, cell_index)
        q_ind_sheet, q_row_of, q_missing = build_indicator_sheet(
            wb, "분기", "분기_재무제표", q_labels, q_row_map, q_name_map
        )
        build_chart_sheet(
            wb, "분기", q_ind_sheet, q_row_of, len(q_labels),
            embed_anchor_col=get_column_letter(2 + len(q_labels) + 2),
        )
        if q_missing:
            all_missing["분기"] = q_missing

    if year_list:
        y_row_map, y_name_map, y_labels = build_annual_sheet(
            wb, year_list, reports, cell_index, estimated_year=estimated_year,
        )
        y_ind_sheet, y_row_of, y_missing = build_indicator_sheet(
            wb, "연간", "연간_재무제표", y_labels, y_row_map, y_name_map
        )
        build_chart_sheet(
            wb, "연간", y_ind_sheet, y_row_of, len(y_labels),
            embed_anchor_col=get_column_letter(2 + len(y_labels) + 2),
        )
        if y_missing:
            all_missing["연간"] = y_missing

        # 투자분석(N섹션 자동평가 · L섹션 가격조회 기준일)에도 추정 연도를 포함시킨다.
        ia_year_list = list(year_list) + ([estimated_year["year"]] if estimated_year else [])
        ann_series = {}
        ia_period_dates = None
        ia_banner = None
        if estimated_year:
            progress_keys = ["매출액", "영업이익", "당기순이익", "자산총계", "자본총계", "부채총계", "유동자산", "유동부채"]
            keys_hits = {k: resolve_metric(k, y_row_map, y_name_map) for k in progress_keys}
            ann_series = compute_estimated_year_series(reports, estimated_year, keys_hits)
            _q_end = {"11013": "0331", "11012": "0630", "11014": "0930"}
            est_date = f"{estimated_year['year']}{_q_end.get(estimated_year['latest_code'], '1231')}"
            ia_period_dates = [f"{y}1231" for y in year_list] + [est_date]
            ia_banner = (
                f"※ {estimated_year['year']}(E)는 {REPRT_NAMES.get(estimated_year['latest_code'], '')} 기준 "
                f"전년 동기 대비 증감율로 추정한 값입니다. 사업보고서가 아니므로 실제와 다를 수 있습니다."
            )
        ia_warnings = build_investment_analysis_sheet(
            wb, args.company_name, args.corp_code, y_labels, ia_year_list,
            y_row_map, y_name_map, y_ind_sheet, y_row_of, reports,
            period_dates=ia_period_dates, annualized_series=ann_series, banner=ia_banner,
        )
        if ia_warnings:
            all_missing["투자분석(계정 매칭 실패)"] = ia_warnings

    # 시트 순서 고정
    desired_order = [
        "분기_재무제표", "연간_재무제표",
        "지표_분기",
        "지표_연간",
        "투자분석",
        "원본데이터",
    ]
    wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]

    today = dt.date.today().strftime("%Y%m%d")
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    suffix = {"annual": "_연간", "quarterly": "_분기", "both": ""}[args.period]
    outfile = outdir / f"{args.company_name}{suffix}_{today}.xlsx"
    wb.save(outfile)
    print(json.dumps(
        {
            "saved": str(outfile),
            "period": args.period,
            "quarters_filled": len(quarter_plan),
            "years_filled": len(year_list),
            "missing_indicators": all_missing,
        },
        ensure_ascii=False,
    ))


if __name__ == "__main__":
    main()

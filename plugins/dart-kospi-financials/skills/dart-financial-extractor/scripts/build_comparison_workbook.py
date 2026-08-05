"""
여러 기업의 cache/ 데이터를 모아 '나란히 비교'용 엑셀 파일 1개를 만든다.

build_workbook.py의 기존 함수(write_raw_sheet, build_quarterly_sheet,
build_annual_sheet, build_indicator_sheet)를 그대로 재사용한다. 회사마다
보유한 기간(분기/연도)이 다를 수 있으므로, 요청한 모든 회사의 기간을 합친
'합집합' 기간 축을 공유하도록 만든 뒤 회사별 원자료/지표 시트는 감사용으로
숨겨두고, 최종 산출물은 딱 2개의 눈에 보이는 시트로만 구성한다:

  - "12분기 비교": 최근 N분기 기준 데이터 표 + 지표별 꺾은선 차트(회사=계열)
  - "최근5년 비교": 최근 N개년 기준 데이터 표 + 지표별 꺾은선 차트(회사=계열)

사용법:
    python build_comparison_workbook.py "00126380:삼성전자" "00356361:LG화학" "00258801:카카오" \
        [--years 5] [--quarters 12] [--outdir /mnt/user-data/outputs]

각 "corp_code:회사명" 항목에 대해 scripts/fetch_financials.py로 데이터가
이미 cache/에 쌓여 있어야 한다(이 스크립트는 API를 호출하지 않는다).

산출물 파일명: "{회사명1, 회사명2, ...} 비교_{YYYYMMDD}.xlsx"
"""
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_workbook import (  # noqa: E402
    build_annual_sheet,
    build_indicator_sheet,
    build_quarter_plan,
    build_quarterly_sheet,
    FONT_NAME,
    load_cache,
    style_header,
    write_raw_sheet,
)

# 비교 표/차트에 노출할 지표 19개 (사용자 요청 순서 그대로).
# 6개 비율은 각 회사의 지표 시트에서 이미 계산이 끝난 값을 그대로 가져온다.
COMPARISON_METRICS = [
    "매출액", "매출원가", "매출총이익", "영업이익", "당기순이익",
    "유동자산", "매출채권및기타채권", "유동부채", "경상이익",
    "자산총계", "부채총계", "이익잉여금", "현금및현금성자산의증가",
    "자기자본비율", "부채비율", "매출총이익률", "원가율", "영업이익률", "순이익률",
]


def quarter_dict_for(reports: dict, year: str, q: int) -> dict:
    """build_quarter_plan과 동일한 규칙으로, 특정 (year, q)에 대해
    이 회사가 실제로 보유한 하위 보고서만 채운 기간 딕셔너리를 만든다.
    회사가 그 분기 데이터를 아예 갖고 있지 않으면 year/q만 있고 나머지는 비어 있다
    (이후 셀 채우기 단계에서 자동으로 빈 칸 처리됨)."""
    y_reports = reports.get(year, {})
    d = {"year": year, "q": q}
    if q == 4:
        if "11011" in y_reports:
            d["fy"] = y_reports["11011"]
        if "11014" in y_reports:
            d["q3"] = y_reports["11014"]
    elif q == 3:
        if "11014" in y_reports:
            d["q3"] = y_reports["11014"]
    elif q == 2:
        if "11012" in y_reports:
            d["h1"] = y_reports["11012"]
        if "11013" in y_reports:
            d["q1"] = y_reports["11013"]
    elif q == 1:
        if "11013" in y_reports:
            d["q1"] = y_reports["11013"]
    return d


def safe_sheet_title(base: str, limit: int = 31) -> str:
    """엑셀 시트명 31자 제한 및 금지문자([]:*?/\\)를 제거한다."""
    for ch in "[]:*?/\\":
        base = base.replace(ch, "")
    return base[:limit]


def build_comparison_sheet(
    wb: Workbook,
    sheet_title: str,
    company_rowmaps: dict,
    period_labels: list[str],
) -> None:
    """company_rowmaps: {회사명: (지표시트이름, {지표명: 행번호})}
    회사별 컬럼 블록(기간 n개씩) + 지표 19행 데이터 표를 만들고,
    그 아래에 지표 하나당 꺾은선 차트 하나(회사=계열)를 배치한다."""
    ws = wb.create_sheet(safe_sheet_title(sheet_title))
    n = len(period_labels)
    companies = list(company_rowmaps.keys())

    # --- 데이터 표 ---
    header1, header2 = 1, 2
    col = 2
    company_col_start: dict[str, int] = {}
    for company in companies:
        company_col_start[company] = col
        ws.cell(row=header1, column=col, value=company).font = Font(name=FONT_NAME, bold=True, size=11)
        ws.merge_cells(start_row=header1, start_column=col, end_row=header1, end_column=col + n - 1)
        for i, label in enumerate(period_labels):
            ws.cell(row=header2, column=col + i, value=label)
        col += n
    style_header(ws, header2, col - 1)
    ws.cell(row=header1, column=1, value="지표").font = Font(name=FONT_NAME, bold=True)

    metric_row: dict[str, int] = {}
    row = 3
    for metric in COMPARISON_METRICS:
        ws.cell(row=row, column=1, value=metric)
        metric_row[metric] = row
        for company in companies:
            ind_sheet_name, row_of = company_rowmaps[company]
            src_row = row_of.get(metric)
            start_col = company_col_start[company]
            for i in range(n):
                c = ws.cell(row=row, column=start_col + i)
                if src_row:
                    col_letter = get_column_letter(3 + i)
                    c.value = f"='{ind_sheet_name}'!{col_letter}{src_row}"
                c.number_format = "0.0" if metric in ("자기자본비율", "부채비율", "매출총이익률", "원가율", "영업이익률", "순이익률") else "#,##0;(#,##0);-"
        row += 1

    ws.column_dimensions["A"].width = 22
    for company in companies:
        start = company_col_start[company]
        for i in range(n):
            ws.column_dimensions[get_column_letter(start + i)].width = 13
    ws.freeze_panes = ws.cell(row=3, column=2).coordinate

    # --- 차트 (지표당 1개, 회사=계열) ---
    table_bottom = row + 2
    anchor_row = table_bottom
    for metric in COMPARISON_METRICS:
        r = metric_row[metric]
        chart = LineChart()
        chart.title = f"{metric} 비교"
        chart.style = 2
        chart.y_axis.title = "%" if metric in ("자기자본비율", "부채비율", "매출총이익률", "원가율", "영업이익률", "순이익률") else "금액(원)"
        chart.x_axis.title = "기간"
        chart.height = 8.5
        chart.width = 22

        has_series = False
        for company in companies:
            start_col = company_col_start[company]
            data_ref = Reference(ws, min_col=start_col, max_col=start_col + n - 1, min_row=r, max_row=r)
            chart.series.append(Series(data_ref, title=company))
            has_series = True
        if not has_series:
            continue
        first_company_start = company_col_start[companies[0]]
        cat_ref = Reference(ws, min_col=first_company_start, max_col=first_company_start + n - 1, min_row=header2, max_row=header2)
        chart.set_categories(cat_ref)
        ws.add_chart(chart, f"A{anchor_row}")
        anchor_row += 18

    ws.sheet_view.showGridLines = False


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("companies", nargs="+", help='"corp_code:회사명" 형식, 2개 이상 나열')
    ap.add_argument("--years", type=int, default=5)
    ap.add_argument("--quarters", type=int, default=12)
    ap.add_argument("--outdir", default="/mnt/user-data/outputs")
    args = ap.parse_args()

    if len(args.companies) < 2:
        print("ERROR: 비교하려면 기업을 2개 이상 지정하세요.", file=sys.stderr)
        sys.exit(1)

    parsed: list[tuple[str, str]] = []
    for c in args.companies:
        if ":" not in c:
            print(f"ERROR: '{c}' 형식이 잘못됐습니다. 'corp_code:회사명' 형식으로 지정하세요.", file=sys.stderr)
            sys.exit(1)
        corp_code, name = c.split(":", 1)
        parsed.append((corp_code, name))

    all_reports: dict[str, tuple[str, dict]] = {}
    for corp_code, name in parsed:
        reports = load_cache(corp_code)
        if not reports:
            print(
                f"ERROR: {name}({corp_code})에 대한 캐시 데이터가 없습니다. "
                f"fetch_financials.py를 먼저 실행하세요.",
                file=sys.stderr,
            )
            sys.exit(1)
        all_reports[name] = (corp_code, reports)

    # 회사별 자연 기간(quarter_plan/year_list)을 각각 구해 합집합(unified) 기간 축을 만든다.
    union_quarters: set[tuple[str, int]] = set()
    union_years: set[str] = set()
    for name, (corp_code, reports) in all_reports.items():
        qp = build_quarter_plan(reports, args.quarters)
        yl = sorted([y for y in reports if "11011" in reports[y]], reverse=True)[: args.years]
        union_quarters.update((p["year"], p["q"]) for p in qp)
        union_years.update(yl)

    unified_quarters = sorted(union_quarters, key=lambda t: (t[0], t[1]))[-args.quarters:]
    unified_years = sorted(union_years)[-args.years:]

    if not unified_quarters and not unified_years:
        print("ERROR: 비교할 기업들의 캐시 데이터에서 유효한 기간을 찾지 못했습니다.", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)

    q_rowmaps: dict[str, tuple[str, dict]] = {}
    y_rowmaps: dict[str, tuple[str, dict]] = {}
    all_missing: dict[str, dict[str, list[str]]] = {}

    for name, (corp_code, reports) in all_reports.items():
        company_quarter_plan = [quarter_dict_for(reports, y, q) for (y, q) in unified_quarters]

        q_sheet_name = safe_sheet_title(f"분기_{name}")
        y_sheet_name = safe_sheet_title(f"연간_{name}")
        raw_sheet_name = safe_sheet_title(f"원본_{name}")

        cell_index = write_raw_sheet(wb, company_quarter_plan, unified_years, reports, sheet_name=raw_sheet_name)
        q_row_map, q_name_map, q_labels = build_quarterly_sheet(
            wb, company_quarter_plan, cell_index, sheet_name=q_sheet_name, hidden=True
        )
        y_row_map, y_name_map, y_labels = build_annual_sheet(
            wb, unified_years, reports, cell_index, sheet_name=y_sheet_name, hidden=True
        )

        q_ind_sheet, q_row_of, q_missing = build_indicator_sheet(
            wb, "q", q_sheet_name, q_labels, q_row_map, q_name_map,
            sheet_name=safe_sheet_title(f"지표_{name}_분기", 31), hidden=True,
        )
        y_ind_sheet, y_row_of, y_missing = build_indicator_sheet(
            wb, "y", y_sheet_name, y_labels, y_row_map, y_name_map,
            sheet_name=safe_sheet_title(f"지표_{name}_연간", 31), hidden=True,
        )
        q_rowmaps[name] = (q_ind_sheet, q_row_of)
        y_rowmaps[name] = (y_ind_sheet, y_row_of)
        if q_missing:
            all_missing.setdefault(name, {})["분기"] = q_missing
        if y_missing:
            all_missing.setdefault(name, {})["연간"] = y_missing

    q_period_labels = [f"{y}Q{q}" for (y, q) in unified_quarters]
    y_period_labels = [f"{y}(사업보고서)" for y in unified_years]

    if unified_quarters:
        build_comparison_sheet(wb, "12분기 비교", q_rowmaps, q_period_labels)
    if unified_years:
        build_comparison_sheet(wb, "최근5년 비교", y_rowmaps, y_period_labels)

    # 비교 표/차트 시트를 맨 앞으로, 회사별 상세(숨김) 시트는 뒤로
    front = [s for s in ["12분기 비교", "최근5년 비교"] if s in wb.sheetnames]
    rest = [s for s in wb.sheetnames if s not in front]
    wb._sheets = [wb[s] for s in front] + [wb[s] for s in rest]

    today = dt.date.today().strftime("%Y%m%d")
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    names_joined = ", ".join(name for _, name in parsed)
    outfile = outdir / f"{names_joined} 비교_{today}.xlsx"
    wb.save(outfile)

    print(json.dumps(
        {
            "saved": str(outfile),
            "companies": [name for _, name in parsed],
            "unified_quarters": len(unified_quarters),
            "unified_years": len(unified_years),
            "missing_indicators": all_missing,
        },
        ensure_ascii=False,
    ))


if __name__ == "__main__":
    main()

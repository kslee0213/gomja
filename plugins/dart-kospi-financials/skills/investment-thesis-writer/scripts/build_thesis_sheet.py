"""
"투자판단 종합" 시트를 만든다. 사용자가 첨부한 "성공투자노트" 양식의 셀 배치를
그대로 재현한다(회사명/종목코드, 산업분석·판매과정·경쟁우위·생산과정·제품분석·
리스크·예측가능기간·성장성예측·경쟁상황·수익성예측 텍스트 박스 + 장기(과거+예측)
재무추세 표 + 최종결론).

이 스크립트는 정성적 텍스트나 성장률 가정을 스스로 만들어내지 않는다 — 그건
claude가 dart 사업의 내용 + 웹 리서치 + 기존 투자분석 결과를 종합해 작성한 뒤
json 파일로 넘겨줘야 한다. 이 스크립트는 그 json을 받아 정해진 레이아웃에
꽂아 넣고, 과거 실적은 이미 만들어진 "지표_연간" 시트를 참조하는 수식으로,
예측 구간은 claude가 제시한 성장률 가정을 compounding하는 수식으로 채운다
(둘 다 감사 가능하도록 하드코딩 값이 아니라 수식으로 남긴다).

사용법:
    python build_thesis_sheet.py <기존 투자분석 xlsx 경로> <content.json 경로>

content.json 스키마는 이 파일 하단 content_schema_example 참고.
"""
import argparse
import json
import re
import sys
from pathlib import path

from openpyxl import load_workbook
from openpyxl.styles import alignment, border, font, patternfill, side
from openpyxl.utils import get_column_letter, range_boundaries

font_name = "맑은 고딕"
title_font = font(name=font_name, bold=true, size=16)
section_font = font(name=font_name, bold=true, size=11, color="ffffff")
thin_side = side(style="thin", color="000000")
thin_border = border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def apply_grid_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> none:
    """지정한 범위의 모든 셀에 얇은 테두리를 그린다."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border


def apply_grid_border_range(ws, a1_range: str) -> none:
    """"b5:o10" 같은 a1 범위 문자열에 그대로 테두리를 적용한다."""
    min_col, min_row, max_col, max_row = range_boundaries(a1_range)
    apply_grid_border(ws, min_row, max_row, min_col, max_col)
section_fill = patternfill("solid", fgcolor="2f5597")
body_font = font(name=font_name, size=10)
wrap = alignment(wrap_text=true, vertical="top", horizontal="left")
note_font = font(name=font_name, italic=true, size=9, color="808080")

content_schema_example = {
    "company_name": "회사명",
    "stock_code": "005930",
    "industry_analysis": "산업 분석 텍스트",
    "sales_process": "판매 과정 텍스트",
    "competitive_advantage": "경쟁 우위 텍스트",
    "production_process": "생산 과정 텍스트",
    "production_process_extra": "생산/설비투자 관련 보충 텍스트(선택)",
    "products": [
        {"name": "제품1", "dex-scription": "제품1 설명"},
        {"name": "제품2", "dex-scription": "제품2 설명"},
    ],
    "risk": "리스크 텍스트",
    "predictable_period_text": "예측 가능 기간 텍스트",
    "growth_prediction_text": "성장성 예측 텍스트",
    "competitive_situation": "경쟁 상황(시장점유율 등) 텍스트",
    "profitability_prediction_text": "수익성 예측 텍스트",
    "final_conclusion_text": "최종 결론 텍스트",
    "sustainable_period_years": 10,
    "expected_annual_return_pct": 15.6,
    "projection_years": 10,
    "revenue_growth_assumptions": [0.10] * 10,
    "op_margin_assumptions": [0.11] * 10,
    "net_margin_assumptions": [0.10] * 10,
    "sources": ["웹 리서치에 사용한 출처 url들(선택, 각주용)"],
}


def find_year_row(ws, label: str) -> int | none:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
        if row[0].value == label:
            return row[0].row
    return none


def build_thesis_sheet(
    src_path: str, content_path: str, outdir: str | none = none,
    ind_sheet_name: str = "지표_연간",
    ia_sheet_name: str = "투자분석",
    target_sheet_name: str = "투자판단 종합",
    period_mode: str = "annual",
) -> str:
    wb = load_workbook(src_path)
    if ia_sheet_name not in wb.sheetnames:
        print(f"error: 원본 파일에 '{ia_sheet_name}' 시트가 없습니다. 먼저 build_workbook.py로 만드세요.", file=sys.stderr)
        sys.exit(1)
    ind_sheet = wb[ind_sheet_name] if ind_sheet_name in wb.sheetnames else none
    if ind_sheet is none:
        print(f"error: 원본 파일에 '{ind_sheet_name}' 시트가 없습니다.", file=sys.stderr)
        sys.exit(1)

    content = json.loads(path(content_path).read_text(encoding="utf-8"))

    if target_sheet_name in wb.sheetnames:
        del wb[target_sheet_name]
    ws = wb.create_sheet(target_sheet_name)
    ws.sheet_view.showgridlines = false

    if period_mode == "quarterly":
        ws.cell(row=1, column=1, value=(
            "※ 본 시트는 분기 누적 공시를 연간 환산(fy_e)한 값을 기반으로 작성된 추정치입니다. "
            "최신 분기가 완결된 사업연도가 아니면 예측 오차가 커질 수 있습니다."
        )).font = note_font

    # --- 상단: 회사명 / 종목코드 ---
    ws["b2"] = content.get("company_name", "")
    ws["b2"].font = title_font
    ws["d2"] = content.get("stock_code", "")
    ws["d2"].font = font(name=font_name, size=12)

    def section(header_cell: str, header_merge: str, header_text: str,
                body_cell: str, body_merge: str, body_text: str):
        ws.merge_cells(header_merge)
        hc = ws[header_cell]
        hc.value = header_text
        hc.font = section_font
        hc.fill = section_fill
        hc.alignment = alignment(horizontal="center", vertical="center")
        apply_grid_border_range(ws, header_merge)
        ws.merge_cells(body_merge)
        bc = ws[body_cell]
        bc.value = body_text or "(내용 없음)"
        bc.font = body_font
        bc.alignment = wrap
        apply_grid_border_range(ws, body_merge)

    section("b4", "b4:c4", "산업 분석", "b5", "b5:o10", content.get("industry_analysis", ""))
    section("q4", "q4:r4", "판매 과정", "q5", "q5:t37", content.get("sales_process", ""))
    section("v4", "v4:w4", "경쟁 우위", "v5", "v5:ac15", content.get("competitive_advantage", ""))
    section("b12", "b12:c12", "생산 과정", "b13", "b13:h21", content.get("production_process", ""))

    ws.merge_cells("i12:j12")
    ws["i12"] = "제품 분석"
    ws["i12"].font = section_font
    ws["i12"].fill = section_fill
    ws["i12"].alignment = alignment(horizontal="center", vertical="center")
    apply_grid_border_range(ws, "i12:j12")

    products = content.get("products", [])
    if len(products) > 0:
        ws.merge_cells("i13:k13")
        ws["i13"] = products[0].get("name", "")
        ws["i13"].font = font(name=font_name, bold=true)
        apply_grid_border_range(ws, "i13:k13")
        ws.merge_cells("l13:o20")
        ws["l13"] = products[0].get("dex-scription", "")
        ws["l13"].font = body_font
        ws["l13"].alignment = wrap
        apply_grid_border_range(ws, "l13:o20")
    if len(products) > 1:
        ws.merge_cells("i21:k21")
        ws["i21"] = products[1].get("name", "")
        ws["i21"].font = font(name=font_name, bold=true)
        apply_grid_border_range(ws, "i21:k21")
        ws.merge_cells("l21:o29")
        ws["l21"] = products[1].get("dex-scription", "")
        ws["l21"].font = body_font
        ws["l21"].alignment = wrap
        apply_grid_border_range(ws, "l21:o29")

    if content.get("production_process_extra"):
        ws.merge_cells("b22:h29")
        ws["b22"] = content["production_process_extra"]
        ws["b22"].font = body_font
        ws["b22"].alignment = wrap
        apply_grid_border_range(ws, "b22:h29")

    section("v16", "v16:w16", "리스크", "v17", "v17:ac27", content.get("risk", ""))
    section("v28", "v28:x28", "예측 가능 기간", "v29", "v29:y43", content.get("predictable_period_text", ""))
    section("z28", "z28:ab28", "성장성 예측", "z29", "z29:ac35", content.get("growth_prediction_text", ""))
    section("b31", "b31:c31", "경쟁 상황", "b32", "b32:o37", content.get("competitive_situation", ""))
    section("z36", "z36:ab36", "수익성 예측", "z37", "z37:ac43", content.get("profitability_prediction_text", ""))

    # --- 최종 결론 ---
    ws.merge_cells("v44:w44")
    ws["v44"] = "최종 결론"
    ws["v44"].font = section_font
    ws["v44"].fill = section_fill
    ws["v44"].alignment = alignment(horizontal="center", vertical="center")
    apply_grid_border_range(ws, "v44:w44")

    ws.merge_cells("v45:x45")
    ws["v45"] = "지속 가능 기간"
    ws["v45"].font = font(name=font_name, bold=true)
    ws["y45"] = content.get("sustainable_period_years")
    ws["z45"] = "년"
    apply_grid_border(ws, 45, 45, 22, 26)  # v45:z45

    ws.merge_cells("v46:x46")
    ws["v46"] = "연평균 예상 수익률"
    ws["v46"].font = font(name=font_name, bold=true)
    ret = content.get("expected_annual_return_pct")
    ws["y46"] = (ret / 100) if ret is not none else none
    ws["y46"].number_format = "0.0%"
    apply_grid_border(ws, 46, 46, 22, 26)  # v46:z46

    ws.merge_cells("v47:ac55")
    ws["v47"] = content.get("final_conclusion_text", "")
    ws["v47"].font = body_font
    ws["v47"].alignment = wrap
    apply_grid_border_range(ws, "v47:ac55")

    # --- 장기 재무추세 표: 과거(지표_연간 참조 수식) + 예측(가정 성장률 compounding 수식) ---
    # 지표_연간 시트의 헤더 행(연도)과 각 지표 행 위치를 찾는다.
    header_row = none
    for r in range(1, 5):
        if ind_sheet.cell(row=r, column=1).value == "지표":
            header_row = r
            break
    if header_row is none:
        print("error: '지표_연간' 시트 구조를 인식하지 못했습니다.", file=sys.stderr)
        sys.exit(1)

    hist_years = []
    col = 3
    while ind_sheet.cell(row=header_row, column=col).value not in (none, ""):
        hist_years.append((col, ind_sheet.cell(row=header_row, column=col).value))
        col += 1
    n_hist = len(hist_years)

    proj_years_n = int(content.get("projection_years", 10))

    def _extract_year(label) -> int | none:
        m = re.match(r"(\d{4})", str(label))
        return int(m.group(1)) if m else none

    last_year = _extract_year(hist_years[-1][1]) if hist_years else none
    proj_years = [last_year + i for i in range(1, proj_years_n + 1)] if last_year else []

    def find_row_in_col_a(ws, label: str) -> int | none:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == label:
                return row[0].row
        return none

    ia_sheet = wb[ia_sheet_name] if ia_sheet_name in wb.sheetnames else none
    ia_row_map = {}
    if ia_sheet is not none:
        ia_row_map = {
            "종가": find_row_in_col_a(ia_sheet, "종가"),
            "시가총액": find_row_in_col_a(ia_sheet, "시가총액"),
            "per(배)": find_row_in_col_a(ia_sheet, "per(배)"),
            "pbr(배)": find_row_in_col_a(ia_sheet, "pbr(배)"),
        }

    row_map = {
        "매출액": find_year_row(ind_sheet, "매출액"),
        "영업이익": find_year_row(ind_sheet, "영업이익"),
        "당기순이익": find_year_row(ind_sheet, "당기순이익"),
        "자산총계": find_year_row(ind_sheet, "자산총계"),
        "부채총계": find_year_row(ind_sheet, "부채총계"),
        "자본총계": find_year_row(ind_sheet, "자본총계"),
    }

    table_start_col = 3  # c열부터 (기존 d열에서 한 칸 당김)
    idx_row = 39  # 연차 (연도보다 아래)
    year_row = 38  # 연도 (먼저 나옴)
    ws.cell(row=year_row, column=1, value="(억원)").font = note_font
    ws.cell(row=year_row, column=2, value="연도" if period_mode == "annual" else "기간")
    all_years = [y for _, y in hist_years] + proj_years
    for i, y in enumerate(all_years):
        ws.cell(row=year_row, column=table_start_col + i, value=y if isinstance(y, str) else int(y))
    ws.cell(row=idx_row, column=2, value="연차")
    for i, _ in enumerate(hist_years + [(none, y) for y in proj_years]):
        c = table_start_col + i
        ws.cell(row=idx_row, column=c, value=i + 1)

    labels = [
        ("시가총액", 40), ("주가", 41), ("per", 42), ("pbr", 43),
        ("매출액", 44), ("growth", 45), ("영업이익", 46), ("영업이익률", 47), ("growth", 48),
        ("순이익", 49), ("순이익률", 50), ("growth", 51), ("roe", 52),
        ("자산", 53), ("부채", 54), ("자본", 55),
    ]
    for name, r in labels:
        ws.cell(row=r, column=2, value=name)

    # 시가총액·주가·per·pbr은 이미 "투자분석" 시트 l섹션에서 계산해 둔 값을
    # 그대로 참조한다(과거 실적 연도만 — 미래 주가는 예측하지 않는다).
    if ia_sheet is not none and all(ia_row_map.values()):
        for i in range(n_hist):
            col_letter = get_column_letter(table_start_col + i)
            ia_col = get_column_letter(3 + i)  # 투자분석 l섹션도 동일한 연도 순서로 c열부터 시작
            ws[f"{col_letter}40"] = f"='{ia_sheet_name}'!{ia_col}{ia_row_map['시가총액']}"
            ws[f"{col_letter}41"] = f"='{ia_sheet_name}'!{ia_col}{ia_row_map['종가']}"
            ws[f"{col_letter}42"] = f"='{ia_sheet_name}'!{ia_col}{ia_row_map['per(배)']}"
            ws[f"{col_letter}43"] = f"='{ia_sheet_name}'!{ia_col}{ia_row_map['pbr(배)']}"
            ws.cell(row=40, column=table_start_col + i).number_format = "#,##0.0"
            ws.cell(row=41, column=table_start_col + i).number_format = "#,##0"
            ws.cell(row=42, column=table_start_col + i).number_format = "0.0"
            ws.cell(row=43, column=table_start_col + i).number_format = "0.00"
    else:
        ws.cell(row=40, column=1, value=(
            "※ '투자분석' 시트에서 시가총액/종가/per/pbr을 찾지 못해 비워둡니다"
            "(krx_auth_key 없이 만든 파일일 수 있습니다)."
        )).font = note_font

    rev_g = content.get("revenue_growth_assumptions", [0.1] * proj_years_n)
    op_m = content.get("op_margin_assumptions", [0.1] * proj_years_n)
    net_m = content.get("net_margin_assumptions", [0.1] * proj_years_n)

    for i in range(len(all_years)):
        col_letter = get_column_letter(table_start_col + i)
        is_hist = i < n_hist

        if is_hist:
            src_col = get_column_letter(hist_years[i][0])
            if row_map["매출액"]:
                ws[f"{col_letter}44"] = f"='{ind_sheet_name}'!{src_col}{row_map['매출액']}"
            if row_map["영업이익"]:
                ws[f"{col_letter}46"] = f"='{ind_sheet_name}'!{src_col}{row_map['영업이익']}"
            if row_map["당기순이익"]:
                ws[f"{col_letter}49"] = f"='{ind_sheet_name}'!{src_col}{row_map['당기순이익']}"
            if row_map["자산총계"]:
                ws[f"{col_letter}53"] = f"='{ind_sheet_name}'!{src_col}{row_map['자산총계']}"
            if row_map["부채총계"]:
                ws[f"{col_letter}54"] = f"='{ind_sheet_name}'!{src_col}{row_map['부채총계']}"
            if row_map["자본총계"]:
                ws[f"{col_letter}55"] = f"='{ind_sheet_name}'!{src_col}{row_map['자본총계']}"
            ws[f"{col_letter}47"] = f"=iferror({col_letter}46/{col_letter}44,\"\")"
            ws[f"{col_letter}50"] = f"=iferror({col_letter}49/{col_letter}44,\"\")"
            ws[f"{col_letter}52"] = f"=iferror({col_letter}49/{col_letter}55,\"\")"
        else:
            p = i - n_hist
            prev_col = get_column_letter(table_start_col + i - 1)
            g = rev_g[p] if p < len(rev_g) else rev_g[-1] if rev_g else 0.1
            om = op_m[p] if p < len(op_m) else op_m[-1] if op_m else 0.1
            nm = net_m[p] if p < len(net_m) else net_m[-1] if net_m else 0.1
            ws[f"{col_letter}44"] = f"={prev_col}44*(1+{g})"
            ws[f"{col_letter}46"] = f"={col_letter}44*{om}"
            ws[f"{col_letter}47"] = om
            ws[f"{col_letter}49"] = f"={col_letter}44*{nm}"
            ws[f"{col_letter}50"] = nm
            # 자산/부채/자본은 단순화: 자본은 전기 자본+당기순이익(배당 가정 없이 전액 유보) 누적, 부채는 전기와 동일 유지, 자산=부채+자본
            ws[f"{col_letter}55"] = f"={prev_col}55+{col_letter}49"
            ws[f"{col_letter}54"] = f"={prev_col}54"
            ws[f"{col_letter}53"] = f"={col_letter}54+{col_letter}55"
            ws[f"{col_letter}52"] = f"=iferror({col_letter}49/{col_letter}55,\"\")"

        # growth(전기 대비 증가율)는 과거·예측 구간 구분 없이 앞 열을 참조하는
        # 수식으로 통일한다(첫 연도는 전기가 없어 비워둔다).
        if i > 0:
            prev_col_g = get_column_letter(table_start_col + i - 1)
            ws[f"{col_letter}45"] = f"=iferror(({col_letter}44-{prev_col_g}44)/{prev_col_g}44,\"\")"
            ws[f"{col_letter}48"] = f"=iferror(({col_letter}46-{prev_col_g}46)/{prev_col_g}46,\"\")"
            ws[f"{col_letter}51"] = f"=iferror(({col_letter}49-{prev_col_g}49)/{prev_col_g}49,\"\")"

        for r in (44, 46, 49, 53, 54, 55):
            ws.cell(row=r, column=table_start_col + i).number_format = "#,##0.0"
        for r in (45, 47, 48, 50, 51, 52):
            ws.cell(row=r, column=table_start_col + i).number_format = "0.0%"

    apply_grid_border(ws, year_row, 55, 2, table_start_col + len(all_years) - 1)

    ws.cell(row=len(all_years) + 57, column=2, value=(
        "※ 굵게 표시되지 않은 연도(마지막 실적연도 이후)는 사용자/claude가 제시한 성장률 가정에 따른 추정치입니다. "
        "실적이 아니라 시나리오이므로 투자 판단 시 가정을 직접 검토하세요."
    )).font = note_font

    ws.column_dimensions["a"].width = 3
    ws.column_dimensions["b"].width = 14
    for col in range(3, 30):
        ws.column_dimensions[get_column_letter(col)].width = 11

    if outdir is none:
        # 기본 동작(v0.9.1부터): 새 파일을 만들지 않고 원본 파일 그대로 덮어써서
        # "재무제표+투자분석+투자판단 종합"이 전부 한 파일에 담기게 한다.
        outfile = path(src_path)
    else:
        today = __import__("datetime").date.today().strftime("%y%m%d")
        outdir_path = path(outdir)
        outdir_path.mkdir(parents=true, exist_ok=true)
        suffix = "투자판단종합_분기추정" if period_mode == "quarterly" else "투자판단종합"
        outfile = outdir_path / f"{content.get('company_name','기업')}_{suffix}_{today}.xlsx"
    wb.save(outfile)
    return str(outfile)


def main() -> none:
    ap = argparse.argumentparser()
    ap.add_argument("src_xlsx", help="기존 투자분석 포함 xlsx 경로 (예: 연간 파일)")
    ap.add_argument("content_json", help="정성 콘텐츠 json 파일 경로")
    ap.add_argument(
        "--outdir", default=none,
        help="지정하면 새 파일(...''_투자판단종합_'yyyymmdd.xlsx)로 따로 저장한다. "
             "지정하지 않으면(기본값) src_xlsx 파일 자체에 시트를 추가해 덮어쓴다 "
             "— 재무제표+투자분석+투자판단 종합이 파일 하나로 합쳐진다.",
    )
    args = ap.parse_args()
    # v0.13.0에서 분기 연환산 시트('지표_연환산'/'투자분석_분기추정')가 제거되었으므로 --period 옵션도 제거(연간 전용).
    outfile = build_thesis_sheet(args.src_xlsx, args.content_json, args.outdir)
    print(json.dumps({"saved": outfile}, ensure_ascii=false))


if __name__ == "__main__":
    main()

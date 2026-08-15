"""
"버핏멍거_가치평가" 시트를 만든다. 첨부 참고자료("워렌버핏과 찰리멍거")에서
도출한 계산들을 코드화한다: 오너 어닝, 그레이엄 내재가치·안전마진, 시장
내재 기대성장률 역산, ROIC vs WACC 스프레드, 다년도 DCF 내재가치 계산기,
경제적 해자 체크리스트(9개 항목).

이 스크립트는 순수 계산(오너 어닝 과거 실적, 그레이엄 공식, DCF 산술)은
전부 수식으로 직접 채우지만, 아래는 스스로 만들어내지 않고 content.json으로
Claude가 제공해야 한다 — 판단이 필요한 가정값이기 때문이다:
  - 할인율(discount_rate_pct), 영구성장률(terminal_growth_pct), WACC 가정
  - 향후 오너 어닝 성장률 가정(연도별)
  - 해자 체크리스트 9개 항목의 평가와 근거
  - 종합 결론 텍스트

사용법:
    python build_valuation_sheet.py <투자분석 포함 xlsx 경로> <content.json 경로>

content.json 스키마는 이 파일 하단 CONTENT_SCHEMA_EXAMPLE 참고.
산출물: 기본적으로 원본 파일에 시트를 추가해 같은 파일명으로 덮어쓴다
(--outdir 지정 시에만 별도 파일 생성).
"""
import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "맑은 고딕"
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="7B3F00")  # 다른 두 시트와 색을 다르게(구분용)
BODY_FONT = Font(name=FONT_NAME, size=10)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="808080")
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

CONTENT_SCHEMA_EXAMPLE = {
    "wacc_pct": 8.0,
    "discount_rate_pct": 10.0,
    "terminal_growth_pct": 3.0,
    "projection_years": 10,
    "owner_earnings_growth_assumptions": [0.10] * 10,
    "moat_checklist": [
        {"item": "경쟁우위 (브랜드가치·특허 등)", "rating": "강함", "note": "..."},
        {"item": "높은 시장 점유율 내지 시장 장악력", "rating": "보통", "note": "..."},
        {"item": "탁월한 경영", "rating": "보통", "note": "..."},
        {"item": "안전마진 (경제위기 극복 능력)", "rating": "보통", "note": "..."},
        {"item": "가격 전가 능력", "rating": "약함", "note": "..."},
        {"item": "신뢰성 (안정적 수치)", "rating": "보통", "note": "..."},
        {"item": "규제 대응 경험", "rating": "-", "note": "해당 없음"},
        {"item": "규모상 우위 (규모의 경제)", "rating": "보통", "note": "..."},
        {"item": "네트워크 효과", "rating": "약함", "note": "..."},
    ],
    "valuation_conclusion": "...",
}


def apply_grid_border(ws, min_row, max_row, min_col, max_col) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def find_row(ws, label: str, col: int = 1) -> int | None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
        if row[0].value == label:
            return row[0].row
    return None


def style_header(ws, row: int, min_col: int, max_col: int):
    for col in range(min_col, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name=FONT_NAME, bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def build_valuation_sheet(src_path: str, content_path: str, outdir: str | None = None) -> str:
    wb = load_workbook(src_path)
    if "지표_연간" not in wb.sheetnames or "투자분석" not in wb.sheetnames:
        print("ERROR: 원본 파일에 '지표_연간'/'투자분석' 시트가 필요합니다. "
              "먼저 dart-financial-extractor로 '--period annual' 파일을 만드세요.", file=sys.stderr)
        sys.exit(1)
    ind = wb["지표_연간"]
    ia = wb["투자분석"]
    content = json.loads(Path(content_path).read_text(encoding="utf-8"))

    if "버핏멍거_가치평가" in wb.sheetnames:
        del wb["버핏멍거_가치평가"]
    ws = wb.create_sheet("버핏멍거_가치평가")
    ws.sheet_view.showGridLines = False

    # --- 연도 헤더(지표_연간과 동일한 기간) ---
    header_row_ind = None
    for r in range(1, 5):
        if ind.cell(row=r, column=1).value == "지표":
            header_row_ind = r
            break
    n = 0
    while ind.cell(row=header_row_ind, column=3 + n).value not in (None, ""):
        n += 1
    period_labels = [ind.cell(row=header_row_ind, column=3 + i).value for i in range(n)]

    row_map = {
        "당기순이익": find_row(ind, "당기순이익", col=2),
        "자본총계": find_row(ind, "자본총계", col=2),
        "부채총계": find_row(ind, "부채총계", col=2),
        "유동부채": find_row(ind, "유동부채", col=2),
        "자산총계": find_row(ind, "자산총계", col=2),
    }
    ia_row = {
        "감가상각비": find_row(ia, "감가상각비"),
        "유형자산의취득": find_row(ia, "유형자산의취득"),
        "종가": find_row(ia, "종가"),
        "시가총액": find_row(ia, "시가총액"),
        "PER(배)": find_row(ia, "PER(배)"),
        "PBR(배)": find_row(ia, "PBR(배)"),
        "ROIC(간이, %)": find_row(ia, "ROIC(간이, %)"),
    }
    missing = [k for k, v in {**row_map, **ia_row}.items() if v is None]

    row = 1
    ws.cell(row=row, column=1, value="버핏-멍거 가치평가").font = TITLE_FONT
    ws.cell(row=row, column=4, value="(금액 단위: 억원, 참고자료: 워렌버핏과 찰리멍거)").font = NOTE_FONT
    row += 2
    if missing:
        ws.cell(row=row, column=1, value=(
            f"※ 아래 항목의 소스 셀을 찾지 못해 일부 계산이 비어 있을 수 있습니다: {', '.join(missing)}"
        )).font = Font(name=FONT_NAME, italic=True, size=9, color="C00000")
        row += 2

    def section_title(text: str):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + n)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        row += 1

    def period_header():
        nonlocal row
        ws.cell(row=row, column=1, value="지표")
        for i, label in enumerate(period_labels):
            ws.cell(row=row, column=3 + i, value=label)
        style_header(ws, row, 1, 2 + n)
        apply_grid_border(ws, row, row, 1, 2 + n)
        row += 1

    def data_row(name: str, formula_fn, fmt="#,##0.0"):
        nonlocal row
        ws.cell(row=row, column=1, value=name)
        for i in range(n):
            f = formula_fn(i)
            c = ws.cell(row=row, column=3 + i)
            if f not in (None, ""):
                c.value = f
            c.number_format = fmt
        apply_grid_border(ws, row, row, 1, 2 + n)
        row += 1
        return row - 1

    def ind_ref(key: str, i: int) -> str | None:
        r = row_map.get(key)
        if not r:
            return None
        col = get_column_letter(3 + i)
        return f"'지표_연간'!{col}{r}"

    def ia_ref(key: str, i: int) -> str | None:
        r = ia_row.get(key)
        if not r:
            return None
        col = get_column_letter(3 + i)
        return f"'투자분석'!{col}{r}"

    # === A. 오너 어닝 (당기순이익 + 감가상각비 - CAPEX) ===
    section_title("A. 오너 어닝 (Owner Earnings) = 당기순이익 + 감가상각비 − |CAPEX|")
    period_header()
    oe_row = data_row(
        "오너 어닝",
        lambda i: (
            f"=IFERROR({ind_ref('당기순이익', i)}+{ia_ref('감가상각비', i)}-ABS({ia_ref('유형자산의취득', i)}),\"\")"
            if ind_ref("당기순이익", i) and ia_ref("감가상각비", i) and ia_ref("유형자산의취득", i) else ""
        ),
    )
    row += 1

    # === B. 그레이엄 내재가치 & 안전마진 (최신 연도 스냅샷) ===
    section_title("B. 그레이엄 내재가치 & 안전마진 (최신 연도 기준)")
    last_i = n - 1
    ws.cell(row=row, column=1, value="항목")
    ws.cell(row=row, column=2, value="값")
    style_header(ws, row, 1, 2)
    apply_grid_border(ws, row, row, 1, 2)
    row += 1

    def snap_row(name: str, formula: str | None, fmt="0.00"):
        nonlocal row
        ws.cell(row=row, column=1, value=name)
        if formula:
            ws.cell(row=row, column=2, value=formula).number_format = fmt
        apply_grid_border(ws, row, row, 1, 2)
        r = row
        row += 1
        return r

    per_ref = ia_ref("PER(배)", last_i)
    pbr_ref = ia_ref("PBR(배)", last_i)
    price_ref = ia_ref("종가", last_i)
    bps_row = snap_row("BPS (주가/PBR)", f"=IFERROR({price_ref}/{pbr_ref},\"\")" if per_ref and pbr_ref else None, "#,##0")
    eps_row = snap_row("EPS (주가/PER)", f"=IFERROR({price_ref}/{per_ref},\"\")" if per_ref else None, "#,##0")
    ni_first = ind_ref("당기순이익", 0)
    ni_last = ind_ref("당기순이익", last_i)
    # CAGR 지수는 "구간 개수"가 아니라 "실제 경과 연수"여야 한다. 연간 모드는
    # 라벨이 "2023" 식이라 연간=구간이지만, 분기 모드는 라벨이 "2023Q1(E)" 식이라
    # 구간(분기) 수를 그대로 쓰면 분기 복리를 연 복리로 착각하는 오류가 생긴다.
    _y_first = re.match(r"(\d{4})", str(period_labels[0]))
    _y_last = re.match(r"(\d{4})", str(period_labels[last_i]))
    years_span = (int(_y_last.group(1)) - int(_y_first.group(1))) if (_y_first and _y_last) else last_i
    years_span = max(years_span, 1)
    cagr_formula = (
        f"=IFERROR(SIGN({ni_last})*SIGN({ni_first})*(({ni_last}/{ni_first})^(1/{years_span})-1),\"\")"
        if ni_first and ni_last and years_span > 0 else None
    )
    cagr_row = snap_row(f"실제 순이익 CAGR ({years_span}개년 환산)", cagr_formula, "0.0%")
    asset_iv_row = snap_row("자산기반 내재가치 = BPS×10", f"=IFERROR(B{bps_row}*10,\"\")" if bps_row else None, "#,##0")
    growth_iv_row = snap_row(
        "성장주 내재가치 = EPS×(8.5+2×기대성장률%)",
        f"=IFERROR(B{eps_row}*(8.5+2*B{cagr_row}*100),\"\")" if eps_row and cagr_row else None,
        "#,##0",
    )
    snap_row("현재 종가", f"={price_ref}" if price_ref else None, "#,##0")
    snap_row(
        "안전마진(자산기반) = (내재가치−종가)/내재가치",
        f"=IFERROR((B{asset_iv_row}-{price_ref})/B{asset_iv_row},\"\")" if price_ref else None,
        "0.0%",
    )
    snap_row(
        "안전마진(성장주기반) = (내재가치−종가)/내재가치",
        f"=IFERROR((B{growth_iv_row}-{price_ref})/B{growth_iv_row},\"\")" if price_ref else None,
        "0.0%",
    )
    ws.cell(row=row, column=1, value=(
        "※ 그레이엄 공식은 참고용 추정치입니다. 안전마진이 20% 이상이면 매수 후보로 보되, "
        "질적 요소(해자, 아래 F섹션)를 반드시 함께 고려하세요."
    )).font = NOTE_FONT
    row += 2

    # === C. 시장 내재 기대성장률 역산 ===
    section_title("C. 시장이 가격에 반영한 기대성장률 역산")
    mkt_g_row = snap_row("시장 내재 기대성장률 = (PER−8.5)/2", f"=IFERROR((({per_ref}-8.5)/2)/100,\"\")" if per_ref else None, "0.0%")
    snap_row("실제 순이익 CAGR (위 B섹션과 동일)", f"=B{cagr_row}" if cagr_row else None, "0.0%")
    snap_row(
        "괴리(실제−시장기대). 양수면 시장이 저평가, 음수면 고평가 신호",
        f"=IFERROR(B{cagr_row}-B{mkt_g_row},\"\")" if cagr_row and mkt_g_row else None,
        "0.0%",
    )
    row += 1

    # === D. ROIC vs 자본비용(WACC) ===
    section_title("D. ROIC vs 자본비용(WACC) — 경제적 해자의 정량적 증거")
    wacc = content.get("wacc_pct", 8.0)
    roic_ref = ia_ref("ROIC(간이, %)", last_i)
    snap_row("ROIC(간이, 최신연도, 투자분석 I섹션 참조)", f"={roic_ref}" if roic_ref else None, "0.0")
    wacc_row = snap_row("자본비용(WACC) 가정", wacc, "0.0")
    ws.cell(row=wacc_row, column=2).number_format = "0.0"
    snap_row(
        "스프레드 = ROIC − WACC (양수면 자본비용을 넘는 초과수익 = 해자의 정량적 근거)",
        f"=IFERROR({roic_ref}-B{wacc_row},\"\")" if roic_ref else None,
        "0.0",
    )
    ws.cell(row=row, column=1, value="※ WACC는 별도 베타 데이터 없이 사용자가 제시한 가정치입니다. 업종 평균(보통 6~10%)을 참고해 조정하세요.").font = NOTE_FONT
    row += 2

    # === E. DCF 내재가치 계산기 (다년도) ===
    section_title("E. DCF 내재가치 계산기 (오너 어닝 기반, 다년도)")
    r_pct = content.get("discount_rate_pct", 10.0)
    g_pct = content.get("terminal_growth_pct", 3.0)
    proj_n = int(content.get("projection_years", 10))
    growth_list = content.get("owner_earnings_growth_assumptions", [0.1] * proj_n)

    ws.cell(row=row, column=1, value="할인율(r)"); ws.cell(row=row, column=2, value=r_pct / 100).number_format = "0.0%"
    r_cell = f"$B${row}"
    row += 1
    ws.cell(row=row, column=1, value="영구성장률(g)"); ws.cell(row=row, column=2, value=g_pct / 100).number_format = "0.0%"
    g_cell = f"$B${row}"
    row += 1
    row += 1

    dcf_header = row
    ws.cell(row=row, column=1, value="연차")
    for i in range(proj_n):
        ws.cell(row=row, column=3 + i, value=i + 1)
    style_header(ws, row, 1, 2 + proj_n)
    apply_grid_border(ws, row, row, 1, 2 + proj_n)
    row += 1

    oe_col_last = get_column_letter(3 + last_i)
    oe_base = f"'{ws.title}'!{oe_col_last}{oe_row}"

    oe_proj_row = row
    ws.cell(row=row, column=1, value="예측 오너 어닝")
    prev_ref = oe_base
    for i in range(proj_n):
        g = growth_list[i] if i < len(growth_list) else (growth_list[-1] if growth_list else 0.1)
        col = get_column_letter(3 + i)
        ws.cell(row=row, column=3 + i, value=f"={prev_ref}*(1+{g})").number_format = "#,##0.0"
        prev_ref = f"{col}{row}"
    apply_grid_border(ws, row, row, 1, 2 + proj_n)
    row += 1

    pv_row = row
    ws.cell(row=row, column=1, value="현재가치(PV)")
    for i in range(proj_n):
        col = get_column_letter(3 + i)
        ws.cell(row=row, column=3 + i, value=f"=IFERROR({col}{oe_proj_row}/(1+{r_cell})^{i + 1},\"\")").number_format = "#,##0.0"
    apply_grid_border(ws, row, row, 1, 2 + proj_n)
    row += 2

    last_proj_col = get_column_letter(3 + proj_n - 1)
    sum_pv_row = row
    ws.cell(row=row, column=1, value="예측기간 PV 합계").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=row, column=3, value=f"=SUM(C{pv_row}:{last_proj_col}{pv_row})").number_format = "#,##0.0"
    row += 1
    tv_row = row
    ws.cell(row=row, column=1, value="영구가치(TV) = 마지막 예측 오너어닝×(1+g)/(r−g)")
    ws.cell(row=row, column=3, value=f"=IFERROR({last_proj_col}{oe_proj_row}*(1+{g_cell})/({r_cell}-{g_cell}),\"\")").number_format = "#,##0.0"
    row += 1
    tv_pv_row = row
    ws.cell(row=row, column=1, value="영구가치의 현재가치")
    ws.cell(row=row, column=3, value=f"=IFERROR(C{tv_row}/(1+{r_cell})^{proj_n},\"\")").number_format = "#,##0.0"
    row += 1
    total_iv_row = row
    ws.cell(row=row, column=1, value="DCF 내재가치 합계 (기업 전체, 억원)").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=row, column=3, value=f"=C{sum_pv_row}+C{tv_pv_row}").number_format = "#,##0.0"
    row += 1

    mktcap_ref = ia_ref("시가총액", last_i)
    shares_row = row
    ws.cell(row=row, column=1, value="상장주식수 역산(시가총액[억원]×1억/종가[원])")
    if mktcap_ref and price_ref:
        ws.cell(row=row, column=3, value=f"=IFERROR({mktcap_ref}*100000000/{price_ref},\"\")").number_format = "#,##0"
    row += 1
    dcf_per_share_row = row
    ws.cell(row=row, column=1, value="DCF 주당 내재가치 (억원/주 → 원 환산)").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=row, column=3, value=f"=IFERROR(C{total_iv_row}/C{shares_row}*100000000,\"\")").number_format = "#,##0"
    row += 1
    ws.cell(row=row, column=1, value="안전마진(DCF) = (DCF주당내재가치−종가)/DCF주당내재가치")
    if price_ref:
        ws.cell(row=row, column=3, value=f"=IFERROR((C{dcf_per_share_row}-{price_ref})/C{dcf_per_share_row},\"\")").number_format = "0.0%"
    row += 2
    ws.cell(row=row, column=1, value=(
        "※ 종가는 원 단위, 시가총액·오너어닝은 억원 단위라 주당 내재가치 계산에서 1억을 곱해 단위를 맞췄습니다."
    )).font = NOTE_FONT
    row += 2

    # === F. 경제적 해자 체크리스트 ===
    section_title("F. 경제적 해자 체크리스트")
    ws.cell(row=row, column=1, value="항목")
    ws.cell(row=row, column=2, value="평가")
    ws.cell(row=row, column=3, value="근거")
    style_header(ws, row, 1, 3)
    apply_grid_border(ws, row, row, 1, 3)
    row += 1
    for item in content.get("moat_checklist", []):
        ws.cell(row=row, column=1, value=item.get("item", ""))
        ws.cell(row=row, column=2, value=item.get("rating", ""))
        c = ws.cell(row=row, column=3, value=item.get("note", ""))
        c.alignment = WRAP
        apply_grid_border(ws, row, row, 1, 3)
        row += 1
    row += 1

    # === G. 종합 결론 ===
    section_title("G. 종합 결론")
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 6, end_column=2 + n)
    cc = ws.cell(row=row, column=1, value=content.get("valuation_conclusion", ""))
    cc.font = BODY_FONT
    cc.alignment = WRAP
    apply_grid_border(ws, row, row + 6, 1, 2 + n)
    row += 8

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    for i in range(max(n, proj_n)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    ws.freeze_panes = "C1"

    if outdir is None:
        outfile = Path(src_path)
    else:
        today = __import__("datetime").date.today().strftime("%Y%m%d")
        outdir_path = Path(outdir)
        outdir_path.mkdir(parents=True, exist_ok=True)
        company = content.get("company_name", "기업")
        suffix = "가치평가"
        outfile = outdir_path / f"{company}_{suffix}_{today}.xlsx"
    wb.save(outfile)
    return str(outfile)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("src_xlsx")
    ap.add_argument("content_json")
    ap.add_argument(
        "--outdir", default=None,
        help="지정하지 않으면(기본값) src_xlsx에 시트를 추가해 같은 파일로 덮어쓴다.",
    )
    args = ap.parse_args()
    outfile = build_valuation_sheet(args.src_xlsx, args.content_json, args.outdir)
    print(json.dumps({"saved": outfile}, ensure_ascii=False))


if __name__ == "__main__":
    main()
